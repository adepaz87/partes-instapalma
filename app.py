#!/usr/bin/env python3
"""
Bot de Partes de Trabajo — Instapalma
Webhook para Twilio WhatsApp
"""

from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from twilio.rest import Client
import json
import os
import smtplib
import base64
import io
import psycopg2
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable, Image as RLImage
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER

app = Flask(__name__)

def fmt_cant(v):
    """Formatea cantidad con coma decimal y 2 decimales. Ej: 2.0 → '2,00'"""
    try:
        f = float(v)
        return f"{f:.2f}".replace(".", ",")
    except:
        return str(v)


# ── Base de datos ──────────────────────────────────────────────────────────────
def get_db():
    return psycopg2.connect(os.environ.get('DATABASE_URL', ''))

def init_db():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS partes (
                id SERIAL PRIMARY KEY,
                numero_parte VARCHAR(20),
                fecha VARCHAR(20),
                operario VARCHAR(100),
                cliente TEXT,
                obra TEXT,
                operarios TEXT,
                albaranes TEXT,
                material_stock TEXT,
                devolucion_almacen TEXT,
                descripcion TEXT,
                terminado TEXT,
                tiempo_restante TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS conversaciones_db (
                numero VARCHAR(50) PRIMARY KEY,
                paso VARCHAR(50),
                datos JSONB,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vehiculos (
                id SERIAL PRIMARY KEY,
                matricula VARCHAR(20),
                marca_modelo VARCHAR(100),
                mes VARCHAR(20),
                km_inicio VARCHAR(20),
                km_fin VARCHAR(20),
                proximo_aceite VARCHAR(20),
                estado_neumaticos TEXT,
                conductores TEXT,
                mantenimientos TEXT,
                observaciones TEXT,
                golpes TEXT,
                operario VARCHAR(100),
                created_at TIMESTAMP DEFAULT NOW(),
                pdf_descargado BOOLEAN DEFAULT FALSE,
                pdf_descargado_at TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vacaciones (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                fecha_inicio VARCHAR(20),
                fecha_fin VARCHAR(20),
                dias_solicitados INTEGER DEFAULT 0,
                fecha_solicitud VARCHAR(20),
                estado VARCHAR(20) DEFAULT 'pendiente',
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS saldo_vacaciones (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100) UNIQUE,
                nombre VARCHAR(100),
                dias_totales INTEGER DEFAULT 23,
                dias_usados INTEGER DEFAULT 0,
                anio INTEGER DEFAULT 2026,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS resumen_mes (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                mes VARCHAR(30),
                horas_extra VARCHAR(20),
                dias_vacaciones VARCHAR(20),
                total_gastos VARCHAR(30),
                foto_url TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Migraciones de columnas
        try:
            cur.execute("ALTER TABLE vacaciones ALTER COLUMN fecha_inicio TYPE VARCHAR(50)")
            cur.execute("ALTER TABLE vacaciones ALTER COLUMN fecha_fin TYPE VARCHAR(50)")
            cur.execute("ALTER TABLE vacaciones ALTER COLUMN fecha_solicitud TYPE VARCHAR(50)")
        except Exception:
            pass
        conn.commit()
        cur.close()
        conn.close()
        print("DB inicializada OK")
    except Exception as e:
        print(f"Error init DB: {e}")

def guardar_parte(datos, numero_operario):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO partes (numero_parte, fecha, operario, cliente, obra, operarios, albaranes, material_stock, devolucion_almacen, descripcion, terminado, tiempo_restante)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            datos.get('numero_parte'),
            datos.get('fecha'),
            numero_operario,
            datos.get('cliente'),
            datos.get('obra'),
            datos.get('operarios'),
            datos.get('albaranes'),
            datos.get('material_stock'),
            datos.get('devolucion_almacen', 'Ninguno'),
            datos.get('descripcion'),
            datos.get('terminado'),
            datos.get('tiempo_restante'),
        ))
        parte_id = cur.fetchone()[0]
        conn.commit()
        cur.close()
        conn.close()
        print(f"Parte guardado en DB OK — id={parte_id}")
        return parte_id
    except Exception as e:
        print(f"Error guardando parte: {e}")
        return None

with app.app_context():
    init_db()

# ── Configuración ─────────────────────────────────────────────────────────────
TWILIO_ACCOUNT_SID = os.environ.get('TWILIO_ACCOUNT_SID', '')
TWILIO_AUTH_TOKEN  = os.environ.get('TWILIO_AUTH_TOKEN', '')
TWILIO_WA_NUMBER   = os.environ.get('TWILIO_WA_NUMBER', 'whatsapp:+19784402048')
SUPERVISOR_EMAIL_1 = 'alberto@adpb.es'
SUPERVISOR_EMAIL_2 = 'adm2@adpb.es'
SUPERVISOR_WA      = os.environ.get('SUPERVISOR_WA', 'whatsapp:+34690875940')
SUPERVISOR_WA_2    = 'whatsapp:+34654893491'
GMAIL_USER         = os.environ.get('GMAIL_USER', '')
GMAIL_APP_PASSWORD = os.environ.get('GMAIL_APP_PASSWORD', '')

# Directorio de operarios: número → nombre
OPERARIOS = {
    '34636606175': 'Jonathan Rodríguez',
    '34616233640': 'Toño Guardia',
    '34666123020': 'Antonio J. Pérez',
    '34689448068': 'Airam',
    '34616908968': 'Moisés',
    '34628380158': 'Petter',
    '34689069588': 'Iker',
    '34690875940': 'Alberto',
    '34606544007': 'Adolfo Castro',
}

def nombre_operario(numero):
    """Devuelve 'Nombre (6XXXXXXXX)' a partir de un número WhatsApp."""
    limpio = numero.replace('whatsapp:','').replace('+','').strip()
    nombre = OPERARIOS.get(limpio, '')
    corto  = limpio[2:] if limpio.startswith('34') else limpio
    return f"{nombre} ({corto})" if nombre else corto

# ── Estado de conversaciones (persistido en DB) ────────────────────────────────
def get_estado(numero):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT paso, datos FROM conversaciones_db WHERE numero=%s", (numero,))
        row = cur.fetchone()
        cur.close(); conn.close()
        if row:
            return {'paso': row[0], 'datos': row[1]}
        return None
    except:
        return None

def set_paso(numero, paso):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO conversaciones_db (numero, paso, datos, updated_at)
            VALUES (%s, %s, '{}'::jsonb, NOW())
            ON CONFLICT (numero) DO UPDATE SET paso=%s, updated_at=NOW()
        """, (numero, paso, paso))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error set_paso: {e}")

def set_dato(numero, clave, valor):
    try:
        import json as _json
        # Siempre serializar a JSON válido para almacenamiento en jsonb
        valor_json = _json.dumps(valor, ensure_ascii=False)
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO conversaciones_db (numero, paso, datos, updated_at)
            VALUES (%s, '', jsonb_build_object(%s, %s::jsonb), NOW())
            ON CONFLICT (numero) DO UPDATE
            SET datos = conversaciones_db.datos || jsonb_build_object(%s, %s::jsonb),
                updated_at = NOW()
        """, (numero, clave, valor_json, clave, valor_json))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error set_dato: {e}")

def borrar_estado(numero):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM conversaciones_db WHERE numero=%s", (numero,))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error borrar_estado: {e}")

def iniciar_parte(numero):
    try:
        conn = get_db(); cur = conn.cursor()
        datos = json.dumps({
            'operario': numero,
            'cliente': '', 'obra': '', 'operarios': '',
            'albaranes': '', 'material_stock': '', 'descripcion': '',
            'terminado': '', 'tiempo_restante': '',
            'fecha': datetime.now().strftime('%d/%m/%Y')
        })
        cur.execute("""
            INSERT INTO conversaciones_db (numero, paso, datos, updated_at)
            VALUES (%s, 'fecha', %s::jsonb, NOW())
            ON CONFLICT (numero) DO UPDATE SET paso='fecha', datos=%s::jsonb, updated_at=NOW()
        """, (numero, datos, datos))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error iniciar_parte: {e}")


MENSAJES_INICIO = ['parte', 'parte de trabajo', 'nuevo parte', 'abrir parte', 'crear parte']
MENSAJES_HOLA   = ['hola', 'menu', 'menú', 'inicio', 'ayuda']

MENU_PRINCIPAL = (
    "👋 *Hola! Soy el bot de Instapalma*\n\n"
    "¿Qué quieres hacer?\n\n"
    "1️⃣ Partes de trabajo\n"
    "2️⃣ Salida de almacén\n"
    "3️⃣ Devolución almacén\n"
    "4️⃣ Consulta\n"
    "5️⃣ Herramienta\n"
    "6️⃣ Vacaciones\n"
    "7️⃣ Resumen fin de mes\n"
    "8️⃣ Vehículos\n\n"
    "_Escribe el número o la palabra clave directamente_"
)

def normalizar(texto):
    return texto.strip().lower()
def es_confirmacion(texto):
    return normalizar(texto) in ['si', 'sí', 'ok', 'vale', 'correcto', 'confirmado', 's', 'yes']
def es_cancelacion(texto):
    return normalizar(texto) in ['no', 'cancelar', 'cancel']


def limpiar_nombre_archivo(texto):
    """Elimina caracteres especiales para nombres de archivo seguros."""
    import unicodedata
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')
    texto = texto.replace(' ', '_').replace('/', '-').replace('\\', '-')
    texto = ''.join(c for c in texto if c.isalnum() or c in '_-.')
    return texto.upper()

def generar_pdf(datos):
    """Genera el PDF del parte y devuelve los bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

    elements = []
    AZUL = colors.HexColor('#1a3a5c')
    GRIS = colors.HexColor('#f5f5f5')

    titulo_style = ParagraphStyle('titulo', fontSize=20, textColor=AZUL,
        alignment=TA_CENTER, spaceAfter=4, fontName='Helvetica-Bold')
    sub_style = ParagraphStyle('sub', fontSize=10, textColor=colors.grey,
        alignment=TA_CENTER, spaceAfter=16)
    sec_style = ParagraphStyle('sec', fontSize=9, textColor=colors.white,
        backColor=AZUL, fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=6, borderPad=4)
    pie_style = ParagraphStyle('pie', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    # Cabecera con logo
    import os as _os
    LOGO_PATH = _os.path.join(_os.path.dirname(__file__), 'logo.jpg')
    if _os.path.exists(LOGO_PATH):
        _logo_ratio = 1024 / 219
        _logo_w = 4 * cm
        _logo_h = _logo_w / _logo_ratio
        logo_img = RLImage(LOGO_PATH, width=_logo_w, height=_logo_h)
    else:
        logo_img = Paragraph("INSTAPALMA", titulo_style)
    cab_title = ParagraphStyle('cab_title', fontName='Helvetica-Bold', fontSize=16,
        textColor=AZUL, alignment=1, leading=20)
    cab = Table([[logo_img, Paragraph('PARTE DE TRABAJO', cab_title)]], colWidths=[5*cm, 12*cm])
    cab.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(cab)
    elements.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=6))
    elements.append(Spacer(1, 0.2*cm))

    # Cabecera — solo fecha
    t_cab = Table([['Fecha', datos['fecha']]],
        colWidths=[3*cm, 14*cm])
    t_cab.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), GRIS),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TEXTCOLOR', (0,0), (0,-1), AZUL),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t_cab)
    elements.append(Spacer(1, 0.3*cm))

    # Cliente y Obra
    t_obra = Table([
        ['Cliente', datos['cliente']],
        ['Obra', datos['obra']],
    ], colWidths=[3*cm, 14*cm])
    t_obra.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), AZUL),
        ('TEXTCOLOR', (0,0), (0,-1), colors.white),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 7),
        ('ROWBACKGROUNDS', (1,0), (1,-1), [colors.white, GRIS]),
    ]))
    elements.append(t_obra)
    elements.append(Spacer(1, 0.3*cm))

    # Operarios
    elements.append(Paragraph("OPERARIOS Y HORAS", sec_style))
    ops_rows = [['Operario', 'Horas']]
    for linea in datos['operarios'].split('\n'):
        linea = linea.strip()
        if not linea:
            continue
        # Separadores: —, -, :, o patrón "NOMBRE Xh" al final
        import re as _re
        m = _re.match(r'^(.+?)\s*[—\-:]\s*(\d[\d.,]*\s*h(?:oras?|rs?)?)$', linea, _re.IGNORECASE)
        if not m:
            m = _re.match(r'^(.+?)\s+(\d[\d.,]*\s*h(?:oras?|rs?)?)$', linea, _re.IGNORECASE)
        if m:
            ops_rows.append([m.group(1).strip(), m.group(2).strip()])
        else:
            ops_rows.append([linea, ''])
    t_ops = Table(ops_rows, colWidths=[13*cm, 4*cm])
    t_ops.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), AZUL),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, GRIS]),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
    ]))
    elements.append(t_ops)
    elements.append(Spacer(1, 0.3*cm))

    # Albaranes — una sola columna
    elements.append(Paragraph("ALBARANES", sec_style))
    alb_texto = datos.get('albaranes', '')
    if normalizar(alb_texto) in ['ninguno', 'no', ''] or not alb_texto:
        alb_contenido = '—'
    else:
        alb_contenido = alb_texto.strip()
    t_alb = Table([['Albaranes'], [alb_contenido]], colWidths=[17*cm])
    t_alb.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), AZUL),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
    ]))
    elements.append(t_alb)
    elements.append(Spacer(1, 0.3*cm))

    # Material de stock
    elements.append(Paragraph("MATERIAL DE STOCK", sec_style))
    mat = datos.get('material_stock', 'Ninguno')
    if normalizar(mat) == 'ninguno' or not mat:
        mat_rows = [['Material', 'Cantidad'], ['—', '—']]
    else:
        mat_rows = [['Material', 'Cantidad']]
        for linea in mat.split('\n'):
            linea = linea.strip()
            if not linea:
                continue
            if '—' in linea:
                parts = linea.split('—', 1)
            elif '-' in linea:
                parts = linea.split('-', 1)
            else:
                parts = [linea, '']
            mat_rows.append([parts[0].strip(), parts[1].strip() if len(parts) > 1 else ''])
    t_mat = Table(mat_rows, colWidths=[10*cm, 7*cm])
    t_mat.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), AZUL),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, GRIS]),
    ]))
    elements.append(t_mat)
    elements.append(Spacer(1, 0.3*cm))

    # Devolución a almacén
    elements.append(Paragraph("DEVOLUCION A ALMACEN", sec_style))
    dev = datos.get('devolucion_almacen', 'Ninguno')
    if normalizar(dev) == 'ninguno' or not dev:
        dev_rows = [['Material', 'Cantidad'], ['—', '—']]
    else:
        dev_rows = [['Material', 'Cantidad']]
        for linea in dev.split('\n'):
            linea = linea.strip()
            if not linea:
                continue
            if '—' in linea:
                parts = linea.split('—', 1)
            elif '-' in linea:
                parts = linea.split('-', 1)
            else:
                parts = [linea, '']
            dev_rows.append([parts[0].strip(), parts[1].strip() if len(parts) > 1 else ''])
    t_dev = Table(dev_rows, colWidths=[10*cm, 7*cm])
    t_dev.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e65100')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, GRIS]),
    ]))
    elements.append(t_dev)
    elements.append(Spacer(1, 0.3*cm))

    # Descripción
    elements.append(Paragraph("DESCRIPCION DE TRABAJOS", sec_style))
    t_desc = Table([[datos.get('descripcion', '')]], colWidths=[17*cm])
    t_desc.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 8),
        ('BACKGROUND', (0,0), (-1,-1), colors.white),
        ('MINROWHEIGHT', (0,0), (-1,-1), 60),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    elements.append(t_desc)
    elements.append(Spacer(1, 0.3*cm))

    # Estado del trabajo
    terminado = datos.get('terminado', '')
    tiempo_restante = datos.get('tiempo_restante', '')
    estado_texto = 'TERMINADO ✓' if normalizar(terminado) in ['sí','si'] else f'EN CURSO — Tiempo restante: {tiempo_restante}'
    VERDE = colors.HexColor('#2e7d32')
    NARANJA = colors.HexColor('#e65100')
    color_estado = VERDE if normalizar(terminado) in ['sí','si'] else NARANJA
    t_estado = Table([[f'ESTADO: {estado_texto}']], colWidths=[17*cm])
    t_estado.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), color_estado),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.white),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))
    elements.append(t_estado)

    elements.append(Spacer(1, 1*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Paragraph(
        f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Instapalma",
        pie_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()

def enviar_email_gmail(datos, numero_operario, pdf_bytes=None):
    """Envía el parte por email con el PDF adjunto usando SMTP + Gmail App Password."""
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    user     = os.environ.get('GMAIL_USER', '')
    password = os.environ.get('GMAIL_APP_PASSWORD', '')
    if not user or not password:
        print("GMAIL no configurado")
        return False
    try:
        destinatarios = ['alberto@adpb.es', 'adm2@adpb.es']
        ops   = datos.get('operarios', 'Ninguno')
        term  = datos.get('terminado', '-')
        trem  = datos.get('tiempo_restante', '')
        estado = f"✅ Terminado" if 'í' in term.lower() or term.lower()=='si' else f"🔄 No terminado — {trem}"

        msg = MIMEMultipart()
        msg['From']    = user
        msg['To']      = ', '.join(destinatarios)
        msg['Subject'] = f"Parte {datos.get('numero_parte')} — {datos.get('cliente')} | {datos.get('obra')} | {datos.get('fecha')}"

        cuerpo = (
            f"Parte de trabajo confirmado.\n\n"
            f"Nº Parte: {datos.get('numero_parte')}\n"
            f"Fecha: {datos.get('fecha')}\n"
            f"Cliente: {datos.get('cliente')}\n"
            f"Obra: {datos.get('obra')}\n"
            f"Operarios:\n{ops}\n"
            f"Albaranes: {datos.get('albaranes','Ninguno')}\n"
            f"Material stock: {datos.get('material_stock','Ninguno')}\n"
            f"Descripción: {datos.get('descripcion','-')}\n"
            f"Estado: {estado}\n\n"
            f"Adjunto el PDF del parte."
        )
        msg.attach(MIMEText(cuerpo, 'plain', 'utf-8'))

        if pdf_bytes:
            part = MIMEBase('application', 'pdf')
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            fecha_pdf = datos.get('fecha', '').replace('/', '-').replace(' ', '')
            obra_pdf = limpiar_nombre_archivo(datos.get('obra', 'obra'))
            ops_raw = datos.get('operarios', '')
            ops_lista = [limpiar_nombre_archivo(l.split('—')[0].split('-')[0].strip().split()[0]) for l in ops_raw.split('\n') if l.strip()]
            ops_pdf = '-'.join(ops_lista) if ops_lista else 'OPERARIOS'
            nombre_pdf = f"{fecha_pdf}-{obra_pdf}-{ops_pdf}.pdf"
            part.add_header('Content-Disposition', f'attachment; filename="{nombre_pdf}"')
            msg.attach(part)

        with smtplib.SMTP('smtp.gmail.com', 587, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(user, password)
            server.sendmail(user, destinatarios, msg.as_bytes())
        print("Email enviado OK")
        return True
    except Exception as e:
        print(f"Error enviando email: {e}")
        return False

def enviar_whatsapp(destino, mensaje, media_url=None):
    try:
        from twilio.rest import Client
        client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)
        # Normalizar formato: asegurar whatsapp:+XXXXX sin espacios
        destino = destino.strip()
        if destino.startswith('whatsapp:'):
            num = destino[len('whatsapp:'):].strip().lstrip('+')
            destino = f'whatsapp:+{num}'
        else:
            destino = 'whatsapp:+' + destino.strip().lstrip('+')
        kwargs = dict(from_=TWILIO_WA_NUMBER, to=destino, body=mensaje)
        if media_url:
            kwargs['media_url'] = [media_url]
        client.messages.create(**kwargs)
        print(f"WA enviado OK a {destino}")
    except Exception as e:
        print(f"Error WA: {e}")

def enviar_via_meta(destino_num, mensaje, media_url=None):
    """Envía mensaje via Meta Cloud API. destino_num: número sin whatsapp: ni +, ej '34690875940'"""
    import requests as _req, os as _os
    token = _os.environ.get('META_TOKEN','')
    phone_id = _os.environ.get('META_PHONE_ID','')
    if not token or not phone_id:
        print('enviar_via_meta: faltan META_TOKEN o META_PHONE_ID')
        return
    num = destino_num.replace('whatsapp:','').replace('+','').strip()
    url = f'https://graph.facebook.com/v19.0/{phone_id}/messages'
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}
    if media_url:
        payload = {
            'messaging_product': 'whatsapp',
            'to': num,
            'type': 'image',
            'image': {'link': media_url, 'caption': mensaje}
        }
    else:
        payload = {
            'messaging_product': 'whatsapp',
            'to': num,
            'type': 'text',
            'text': {'body': mensaje}
        }
    try:
        r = _req.post(url, headers=headers, json=payload, timeout=10)
        if r.status_code == 200:
            print(f'Meta WA enviado OK a {num}')
        else:
            print(f'Error Meta WA: {r.status_code} {r.text[:100]}')
    except Exception as e:
        print(f'Error Meta WA: {e}')

def enviar_supervisor(mensaje, media_url=None):
    """Envía mensaje al supervisor (Alberto) siempre via Meta."""
    enviar_via_meta('34690875940', mensaje, media_url=media_url)

def generar_resumen(datos):
    ops  = datos.get('operarios', 'Ninguno')
    albs = datos.get('albaranes', 'Ninguno')
    mat  = datos.get('material_stock', 'Ninguno')
    dev  = datos.get('devolucion_almacen', 'Ninguno')
    desc = datos.get('descripcion', '-')
    term = datos.get('terminado', '-')
    trem = datos.get('tiempo_restante', '')
    linea_term = f"✅ Sí" if normalizar(term) in ['si','sí'] else f"🔄 No — {trem}" if trem else f"🔄 No"
    return (
        f"📋 *RESUMEN DEL PARTE*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📅 Fecha: {datos['fecha']}\n"
        f"🏢 Cliente: {datos['cliente']}\n"
        f"🔨 Obra: {datos['obra']}\n"
        f"👷 Operarios:\n{ops}\n"
        f"📦 Albaranes: {albs}\n"
        f"🏗️ Material stock: {mat}\n"
        f"📦 Devolución almacén: {dev}\n"
        f"📝 Descripción: {desc}\n"
        f"🏁 Terminado: {linea_term}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"¿Es correcto? Responde *SÍ* para enviar o *NO* para cancelar."
    )

def finalizar_parte(numero, datos):
    ops  = datos.get('operarios', 'Ninguno')
    albs = datos.get('albaranes', 'Ninguno')
    mat  = datos.get('material_stock', 'Ninguno')
    dev  = datos.get('devolucion_almacen', 'Ninguno')
    desc = datos.get('descripcion', '-')
    term = datos.get('terminado', '-')
    trem = datos.get('tiempo_restante', '')
    linea_term = f"Sí" if normalizar(term) in ['si','sí'] else f"No — {trem}" if trem else "No"

    msg_supervisor = (
        f"📋 *PARTE DE TRABAJO*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📅 {datos['fecha']}\n"
        f"📱 Operario: {nombre_operario(numero)}\n"
        f"🏢 Cliente: {datos['cliente']}\n"
        f"🔨 Obra: {datos['obra']}\n"
        f"👷 Operarios:\n{ops}\n"
        f"📦 Albaranes: {albs}\n"
        f"🏗️ Material stock: {mat}\n"
        f"📦 Devolución almacén: {dev}\n"
        f"📝 {desc}\n"
        f"🏁 Terminado: {linea_term}\n"
        f"━━━━━━━━━━━━━━━━━━━━"
    )
    enviar_supervisor(msg_supervisor)

    # Guardar en base de datos y obtener el ID para la URL del PDF
    parte_id = guardar_parte(datos, numero)

    # Enviar PDF a tu WhatsApp y al operario
    BOT_URL = os.environ.get('RAILWAY_PUBLIC_DOMAIN', 'bot-production-66b8.up.railway.app')
    if parte_id:
        pdf_url = f"https://{BOT_URL}/partes/{parte_id}/pdf"
        import unicodedata as _ud
        def _clean(t):
            t2 = _ud.normalize('NFD', str(t))
            return ''.join(c for c in t2 if _ud.category(c) != 'Mn')
        caption = (
            f"📄 *Parte* — {datos['fecha']}\n"
            f"🏢 {_clean(datos['cliente'])} | 🔨 {_clean(datos['obra'])}\n"
            f"🏁 {linea_term}"
        )
        enviar_supervisor(caption, media_url=pdf_url)
        operario_wa = f"whatsapp:{numero}" if not numero.startswith("whatsapp:") else numero
        enviar_whatsapp(operario_wa,
                        f"✅ *Parte confirmado*\nAquí tienes tu copia en PDF:",
                        media_url=pdf_url)

    borrar_estado(numero)


MENSAJES_VEHICULO   = ['vehiculo', 'vehículo', 'coche', 'camion', 'camión', 'furgoneta', 'mantenimiento vehiculo']
MENSAJES_VACACIONES = ['vacaciones', 'vacacion', 'solicitar vacaciones', 'pedir vacaciones', 'dias libres', 'días libres']
MENSAJES_RESUMEN_MES = ['resumen mes', 'resumen del mes', 'resumen mensual', 'cierre mes', 'cierre del mes', 'resumen fin de mes', 'fin de mes']
MENSAJES_STOCK_SALIDA   = ['salida']
MENSAJES_STOCK_DEVOL    = ['devolucion', 'devolución']
MENSAJES_STOCK_CONSULTA = ['consulta']

def iniciar_vacaciones(numero):
    try:
        conn = get_db(); cur = conn.cursor()
        datos = json.dumps({
            'tipo': 'vacaciones',
            'operario': numero,
            'fecha_inicio': '', 'fecha_fin': '',
            'fecha_solicitud': datetime.now().strftime('%d/%m/%Y')
        })
        cur.execute("""
            INSERT INTO conversaciones_db (numero, paso, datos, updated_at)
            VALUES (%s, 'vac_inicio', %s::jsonb, NOW())
            ON CONFLICT (numero) DO UPDATE SET paso='vac_inicio', datos=%s::jsonb, updated_at=NOW()
        """, (numero, datos, datos))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error iniciar_vacaciones: {e}")

def calcular_dias_laborables(fecha_inicio, fecha_fin):
    """Calcula días laborables (lun-vie) entre dos fechas en formato DD/MM/YYYY."""
    try:
        from datetime import datetime as dt, timedelta
        fi = dt.strptime(fecha_inicio, '%d/%m/%Y')
        ff = dt.strptime(fecha_fin, '%d/%m/%Y')
        dias = 0
        current = fi
        while current <= ff:
            if current.weekday() < 5:  # lun-vie
                dias += 1
            current += timedelta(days=1)
        return dias
    except:
        return 0

def get_saldo_vacaciones(numero_operario):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT dias_totales, dias_usados, nombre FROM saldo_vacaciones WHERE operario=%s", (numero_operario,))
        row = cur.fetchone()
        return row  # (dias_totales, dias_usados, nombre) o None
    except:
        return None
    finally:
        try:
            if conn: conn.close()
        except: pass

def guardar_vacacion(datos, numero_operario):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        dias = calcular_dias_laborables(datos.get('fecha_inicio',''), datos.get('fecha_fin',''))
        nombre = datos.get('nombre_operario', nombre_operario(numero_operario))
        cur.execute("""
            INSERT INTO vacaciones (operario, nombre_operario, fecha_inicio, fecha_fin, dias_solicitados, fecha_solicitud, estado)
            VALUES (%s, %s, %s, %s, %s, %s, 'pendiente')
            RETURNING id
        """, (
            numero_operario,
            nombre,
            datos.get('fecha_inicio',''),
            datos.get('fecha_fin',''),
            dias,
            datos.get('fecha_solicitud','')
        ))
        vid = cur.fetchone()[0]
        conn.commit()
        return vid, dias
    except Exception as e:
        print(f"Error guardar_vacacion: {e}")
        if conn: conn.rollback()
        return None, 0
    finally:
        try:
            if conn: conn.close()
        except: pass

def aprobar_rechazar_vacacion(vac_id, estado):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("UPDATE vacaciones SET estado=%s WHERE id=%s RETURNING operario, nombre_operario, fecha_inicio, fecha_fin, dias_solicitados", (estado, vac_id))
        row = cur.fetchone()
        if row and estado == 'aprobada':
            # Descontar días del saldo
            operario, nombre, fi, ff, dias = row
            cur.execute("""
                UPDATE saldo_vacaciones SET dias_usados = dias_usados + %s, updated_at=NOW()
                WHERE operario=%s
            """, (dias or 0, operario))
        conn.commit()
        return row  # (operario, nombre, fecha_inicio, fecha_fin, dias)
    except Exception as e:
        print(f"Error aprobar_rechazar: {e}")
        if conn: conn.rollback()
        return None
    finally:
        try:
            if conn: conn.close()
        except: pass


INIT_VEHICULOS_SQL = """
    CREATE TABLE IF NOT EXISTS vehiculos (
        id SERIAL PRIMARY KEY,
        matricula VARCHAR(20),
        marca_modelo VARCHAR(100),
        mes VARCHAR(20),
        km_inicio VARCHAR(20),
        km_fin VARCHAR(20),
        proximo_aceite VARCHAR(20),
        estado_neumaticos TEXT,
        conductores TEXT,
        mantenimientos TEXT,
        observaciones TEXT,
        golpes TEXT,
        operario VARCHAR(100),
        created_at TIMESTAMP DEFAULT NOW(),
        pdf_descargado BOOLEAN DEFAULT FALSE,
        pdf_descargado_at TIMESTAMP
    )
"""

def iniciar_vehiculo(numero):
    try:
        conn = get_db(); cur = conn.cursor()
        datos = json.dumps({
            'tipo': 'vehiculo',
            'operario': numero,
            'matricula': '', 'marca_modelo': '', 'mes': '',
            'km_inicio': '', 'km_fin': '', 'proximo_aceite': '',
            'estado_neumaticos': '', 'conductores': '',
            'mantenimientos': '', 'observaciones': '', 'golpes': '',
            'fecha': datetime.now().strftime('%d/%m/%Y')
        })
        cur.execute("""
            INSERT INTO conversaciones_db (numero, paso, datos, updated_at)
            VALUES (%s, 'v_matricula', %s::jsonb, NOW())
            ON CONFLICT (numero) DO UPDATE SET paso='v_matricula', datos=%s::jsonb, updated_at=NOW()
        """, (numero, datos, datos))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error iniciar_vehiculo: {e}")

def guardar_vehiculo(datos, numero_operario):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO vehiculos (matricula, marca_modelo, mes, km_inicio, km_fin,
                proximo_aceite, estado_neumaticos, conductores, mantenimientos,
                observaciones, golpes, operario)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """, (
            datos.get('matricula',''), datos.get('marca_modelo',''),
            datos.get('mes',''), datos.get('km_inicio',''), datos.get('km_fin',''),
            datos.get('proximo_aceite',''), datos.get('estado_neumaticos',''),
            datos.get('conductores',''), datos.get('mantenimientos',''),
            datos.get('observaciones',''), datos.get('golpes',''),
            numero_operario
        ))
        vid = cur.fetchone()[0]
        conn.commit()
        return vid
    except Exception as e:
        print(f"Error guardar_vehiculo: {e}")
        if conn:
            conn.rollback()
        return None
    finally:
        try:
            if conn:
                conn.close()
        except:
            pass

def generar_pdf_vehiculo(datos):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    AZUL  = colors.HexColor('#1a3a5c')
    GRIS  = colors.HexColor('#f5f5f5')
    titulo_style = ParagraphStyle('t', fontSize=16, textColor=AZUL, alignment=TA_CENTER, fontName='Helvetica-Bold')
    sec_style    = ParagraphStyle('s', fontSize=9,  textColor=colors.white, backColor=AZUL,
                                  fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=6, borderPad=4)
    pie_style    = ParagraphStyle('p', fontSize=7,  textColor=colors.grey, alignment=TA_CENTER)
    import os as _os
    LOGO_PATH = _os.path.join(_os.path.dirname(__file__), 'logo.jpg')
    if _os.path.exists(LOGO_PATH):
        _lw = 4*cm; _lh = _lw / (1024/219)
        logo_img = RLImage(LOGO_PATH, width=_lw, height=_lh)
    else:
        logo_img = Paragraph("INSTAPALMA", titulo_style)
    cab_t = ParagraphStyle('ct', fontName='Helvetica-Bold', fontSize=16, textColor=AZUL, alignment=1, leading=20)
    cab = Table([[logo_img, Paragraph('PARTE DE VEHÍCULO', cab_t)]], colWidths=[5*cm, 12*cm])
    cab.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0)]))
    elements.append(cab)
    elements.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=6))
    elements.append(Spacer(1, 0.2*cm))

    t_gen = Table([
        ['Matrícula', datos.get('matricula',''), 'Mes', datos.get('mes','')],
        ['Marca/Modelo', datos.get('marca_modelo',''), 'Fecha', datos.get('fecha','')],
    ], colWidths=[3*cm, 6.5*cm, 2.5*cm, 5*cm])
    t_gen.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(0,-1),AZUL),('BACKGROUND',(2,0),(2,-1),AZUL),
        ('TEXTCOLOR',(0,0),(0,-1),colors.white),('TEXTCOLOR',(2,0),(2,-1),colors.white),
        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTNAME',(2,0),(2,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),('PADDING',(0,0),(-1,-1),6),
    ]))
    elements.append(t_gen)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("KILÓMETROS", sec_style))
    t_km = Table([
        ['Km inicio mes', datos.get('km_inicio',''), 'Km fin mes', datos.get('km_fin','')],
        ['Próximo cambio aceite (km)', datos.get('proximo_aceite',''), '', ''],
    ], colWidths=[5*cm, 4.5*cm, 4*cm, 3.5*cm])
    t_km.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(0,-1),GRIS),('BACKGROUND',(2,0),(2,-1),GRIS),
        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTNAME',(2,0),(2,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),('PADDING',(0,0),(-1,-1),6),
        ('SPAN',(1,1),(3,1)),
    ]))
    elements.append(t_km)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("ESTADO NEUMÁTICOS", sec_style))
    t_neum = Table([[datos.get('estado_neumaticos','Ninguno')]], colWidths=[17*cm])
    t_neum.setStyle(TableStyle([
        ('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('PADDING',(0,0),(-1,-1),6),('BACKGROUND',(0,0),(-1,-1),GRIS),
    ]))
    elements.append(t_neum)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("CONDUCTORES", sec_style))
    cond_rows = [['Conductor', 'Desde', 'Hasta']]
    for linea in (datos.get('conductores','') or '').split('\n'):
        linea = linea.strip()
        if not linea: continue
        partes = [p.strip() for p in linea.replace('—','-').split('-',2)]
        if len(partes) >= 3:
            cond_rows.append([partes[0], partes[1], partes[2]])
        elif len(partes) == 2:
            cond_rows.append([partes[0], partes[1], ''])
        else:
            cond_rows.append([linea, '', ''])
    t_cond = Table(cond_rows, colWidths=[8*cm, 4.5*cm, 4.5*cm])
    t_cond.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),AZUL),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),('PADDING',(0,0),(-1,-1),6),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GRIS]),
    ]))
    elements.append(t_cond)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("MANTENIMIENTOS REALIZADOS", sec_style))
    mant_rows = [['Concepto', 'Fecha', 'Km']]
    for linea in (datos.get('mantenimientos','') or '').split('\n'):
        linea = linea.strip()
        if not linea: continue
        partes = [p.strip() for p in linea.replace('—','-').split('-',2)]
        if len(partes) >= 3:
            mant_rows.append([partes[0], partes[1], partes[2]])
        elif len(partes) == 2:
            mant_rows.append([partes[0], partes[1], ''])
        else:
            mant_rows.append([linea, '', ''])
    t_mant = Table(mant_rows, colWidths=[9*cm, 4*cm, 4*cm])
    t_mant.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),AZUL),('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),('PADDING',(0,0),(-1,-1),6),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,GRIS]),
    ]))
    elements.append(t_mant)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("OBSERVACIONES / PRÓXIMOS MANTENIMIENTOS", sec_style))
    t_obs = Table([[datos.get('observaciones','Ninguna')]], colWidths=[17*cm])
    t_obs.setStyle(TableStyle([
        ('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('PADDING',(0,0),(-1,-1),6),
    ]))
    elements.append(t_obs)
    elements.append(Spacer(1,0.3*cm))

    elements.append(Paragraph("GOLPES Y DESPERFECTOS", sec_style))
    t_golp = Table([[datos.get('golpes','Ninguno')]], colWidths=[17*cm])
    t_golp.setStyle(TableStyle([
        ('FONTSIZE',(0,0),(-1,-1),9),('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('PADDING',(0,0),(-1,-1),6),('BACKGROUND',(0,0),(-1,-1),GRIS),
    ]))
    elements.append(t_golp)
    elements.append(Spacer(1,0.5*cm))

    elements.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey, spaceAfter=4))
    elements.append(Paragraph("Instapalma — Parte de Vehículo generado automáticamente", pie_style))
    doc.build(elements)
    return buffer.getvalue()

def generar_resumen_vehiculo(datos):
    return (
        f"📋 *Resumen — Parte de Vehículo*\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🚗 {datos.get('matricula','')} | {datos.get('marca_modelo','')}\n"
        f"📅 Mes: {datos.get('mes','')}\n"
        f"📍 Km inicio: {datos.get('km_inicio','')} | Km fin: {datos.get('km_fin','')}\n"
        f"🔧 Próx. aceite: {datos.get('proximo_aceite','')} km\n"
        f"🔴 Neumáticos: {datos.get('estado_neumaticos','')}\n"
        f"👤 Conductores:\n{datos.get('conductores','')}\n"
        f"🛠️ Mantenimientos:\n{datos.get('mantenimientos','')}\n"
        f"📝 Observaciones: {datos.get('observaciones','')}\n"
        f"⚠️ Golpes: {datos.get('golpes','')}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"¿Es correcto? Responde *SÍ* para enviar o *NO* para cancelar."
    )

def finalizar_vehiculo(numero, datos):
    import threading
    mat = datos.get('matricula','').replace(' ','_').upper()
    mes = datos.get('mes','').replace('/','_').replace(' ','_')
    nombre_pdf = f"{mes}-{mat}-VEHICULO.pdf"
    pdf_bytes = generar_pdf_vehiculo(datos)

    # Guardar en DB y enviar WhatsApp primero (no bloquear)
    vid = guardar_vehiculo(datos, numero)
    BOT_URL = os.environ.get('RAILWAY_PUBLIC_DOMAIN', 'bot-production-66b8.up.railway.app')
    if vid:
        pdf_url = f"https://{BOT_URL}/vehiculos/{vid}/pdf"
        caption = f"Parte Vehiculo - {mat} - {mes}\nKm: {datos.get('km_inicio','')} a {datos.get('km_fin','')}"
        enviar_supervisor(caption, media_url=pdf_url)
        op_wa = numero if numero.startswith("whatsapp:") else f"whatsapp:{numero}"
        enviar_whatsapp(op_wa, "Parte de vehiculo enviado. Aqui tienes tu copia:", media_url=pdf_url)

    borrar_estado(numero)

    # Email en hilo separado para no bloquear la respuesta
    def enviar_email_vehiculo():
        try:
            msg_email = MIMEMultipart()
            msg_email['From']    = GMAIL_USER
            msg_email['To']      = ', '.join([SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2])
            msg_email['Subject'] = f"[VEHICULO] Parte - {mat} - {mes}"
            body_txt = (
                f"Parte de vehiculo generado.\n\n"
                f"Matricula: {datos.get('matricula','')}\n"
                f"Modelo: {datos.get('marca_modelo','')}\n"
                f"Mes: {mes}\n"
                f"Km inicio: {datos.get('km_inicio','')} | Km fin: {datos.get('km_fin','')}\n"
            )
            msg_email.attach(MIMEText(body_txt, 'plain'))
            part = MIMEApplication(pdf_bytes, Name=nombre_pdf)
            part.add_header('Content-Disposition', f'attachment; filename="{nombre_pdf}"')
            msg_email.attach(part)
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as srv:
                srv.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                srv.sendmail(GMAIL_USER, [SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2], msg_email.as_string())
            print("Email vehiculo OK")
        except Exception as e:
            print(f"Error email vehiculo: {e}")

    threading.Thread(target=enviar_email_vehiculo, daemon=True).start()


@app.route('/webhook', methods=['GET'])
def webhook_verify():
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    if mode == 'subscribe' and token == 'instapalma2024':
        return challenge, 200
    return 'Forbidden', 403

@app.route('/webhook', methods=['POST'])
def webhook():
    # Detectar si es Meta Cloud API (JSON) o Twilio (form)
    if request.is_json:
        data = request.get_json()
        try:
            entry = data['entry'][0]['changes'][0]['value']
            msg_obj = entry['messages'][0]
            incoming_msg = msg_obj.get('text', {}).get('body', '').strip()
            numero = 'whatsapp:+' + msg_obj['from']
            media_url = ''
        except (KeyError, IndexError):
            return 'OK', 200
    else:
        incoming_msg = request.form.get('Body', '').strip()
        numero = request.form.get('From', '')
        # Capturar media adjunta (foto enviada por el operario)
        num_media = int(request.form.get('NumMedia', 0))
        media_url = request.form.get('MediaUrl0', '') if num_media > 0 else ''

    use_meta = request.is_json

    class MetaMsg:
        def __init__(self):
            self._body = None
        def body(self, text):
            self._body = text
            enviar_whatsapp(numero, text)

    class MetaResp:
        def __init__(self):
            self._msg = MetaMsg()
        def message(self):
            return self._msg
        def __str__(self):
            return ''

    if use_meta:
        resp = MetaResp()
    else:
        resp = MessagingResponse()
    msg = resp.message()
    estado = get_estado(numero)

    # Comando reset en cualquier momento
    if normalizar(incoming_msg) in ['reset', 'reiniciar', 'restart']:
        borrar_estado(numero)
        msg.body("🔄 Conversación reiniciada. Escribe *parte* para comenzar de nuevo.")
        return str(resp)

    # Detectar arranque vehiculo
    if any(p in normalizar(incoming_msg) for p in MENSAJES_VEHICULO):
        iniciar_vehiculo(numero)
        msg.body(
            "🚗 *Bot de Vehículos — Instapalma*\n\n"
            "Vamos a registrar el parte mensual paso a paso.\n\n"
            "1️⃣ ¿Cuál es la *matrícula* del vehículo?"
        )
        return str(resp) if not use_meta else ('OK', 200)

    # ══════════════════════════════════════════════════════════════════════════
    # HERRAMIENTA — Flujo conversacional
    # ══════════════════════════════════════════════════════════════════════════
    msg_n_herr = normalizar(incoming_msg)
    num_limpio = numero.replace('whatsapp:','').replace('+','').strip()
    nombre_op = OPERARIOS.get(num_limpio, numero)

    BOT_URL_H = os.environ.get('RAILWAY_PUBLIC_DOMAIN', 'bot-production-66b8.up.railway.app')

    def _enviar_pdf_herramienta(seccion):
        """Genera PDF, lo guarda en BD y devuelve la URL pública."""
        import time
        pdf_bytes = generar_pdf_herramienta(seccion)
        ts = int(time.time())
        ref = f"HERR-{seccion.upper()}-{ts}"
        # Guardar en BD (reutiliza tabla stock_albaranes con numero=ref)
        try:
            _c = get_db(); _cur = _c.cursor()
            _cur.execute("""
                INSERT INTO stock_albaranes (numero, pdf_bytes)
                VALUES (%s, %s)
                ON CONFLICT (numero) DO UPDATE SET pdf_bytes=%s
            """, (ref, psycopg2.Binary(pdf_bytes), psycopg2.Binary(pdf_bytes)))
            _c.commit(); _cur.close(); _c.close()
        except Exception as _e:
            print(f"Error guardando PDF herramienta: {_e}")
        return f"https://{BOT_URL_H}/albaran/{ref}.pdf"

    MENU_HERRAMIENTA = (
        "🔧 *Control de Herramienta*\n\n"
        "¿Qué quieres hacer?\n\n"
        "1️⃣ Alta en obra\n"
        "2️⃣ Baja / devolución a almacén\n"
        "3️⃣ Listado almacén (PDF)\n"
        "4️⃣ Listado en obra (PDF)\n"
        "5️⃣ Listado personal (PDF)"
    )

    # ── Arranque: palabra "herramienta" ───────────────────────────────────────
    if msg_n_herr.strip() == 'herramienta':
        set_paso(numero, 'herr_menu')
        msg.body(MENU_HERRAMIENTA)
        return str(resp) if not use_meta else ('OK', 200)

    # ── Comando exclusivo supervisor: nueva herramienta <nombre> ─────────────
    import re as _re_herr
    if num_limpio == '34690875940' and _re_herr.match(r'^nueva herramienta\s+.+', msg_n_herr):
        nombre_nueva = _re_herr.sub(r'^nueva herramienta\s+', '', msg_n_herr).strip()
        set_dato(numero, 'herr_nueva_nombre', nombre_nueva)
        set_paso(numero, 'herr_nueva_cantidad')
        msg.body(f"🔧 *{nombre_nueva.capitalize()}*\n\n¿Cuántas unidades entran al almacén?")
        return str(resp) if not use_meta else ('OK', 200)

    # ── Flujo activo de herramienta ───────────────────────────────────────────
    paso_herr = estado['paso'] if estado else None

    if paso_herr == 'herr_menu':
        op = msg_n_herr.strip()
        if op == '1':
            set_paso(numero, 'herr_alta_nombre')
            msg.body("🔧 *Alta en obra*\n\n¿Qué herramienta vas a sacar?\n_(Escribe el nombre, ej: escalera 4 peldaños)_")
            return str(resp) if not use_meta else ('OK', 200)
        elif op == '2':
            set_paso(numero, 'herr_baja_nombre')
            msg.body("🔙 *Devolución a almacén*\n\n¿Qué herramienta devuelves?\n_(Escribe el nombre)_")
            return str(resp) if not use_meta else ('OK', 200)
        elif op == '3':
            borrar_estado(numero)
            try:
                url = _enviar_pdf_herramienta('almacen')
                msg.body(f"📋 *Listado Almacén:*\n{url}")
            except Exception as e:
                msg.body(f"❌ Error generando PDF: {e}")
            return str(resp) if not use_meta else ('OK', 200)
        elif op == '4':
            borrar_estado(numero)
            try:
                url = _enviar_pdf_herramienta('obra')
                msg.body(f"📋 *Herramienta en Obra:*\n{url}")
            except Exception as e:
                msg.body(f"❌ Error generando PDF: {e}")
            return str(resp) if not use_meta else ('OK', 200)
        elif op == '5':
            borrar_estado(numero)
            try:
                url = _enviar_pdf_herramienta('personal')
                msg.body(f"📋 *Herramienta Personal:*\n{url}")
            except Exception as e:
                msg.body(f"❌ Error generando PDF: {e}")
            return str(resp) if not use_meta else ('OK', 200)
        else:
            msg.body(MENU_HERRAMIENTA)
            return str(resp) if not use_meta else ('OK', 200)

    elif paso_herr == 'herr_alta_nombre':
        set_dato(numero, 'herr_nombre', incoming_msg.strip())
        set_paso(numero, 'herr_alta_obra')
        msg.body(f"📍 ¿En qué obra o lugar se va a usar?\n_(Escribe el nombre de la obra)_")
        return str(resp) if not use_meta else ('OK', 200)

    elif paso_herr == 'herr_alta_obra':
        datos_herr = estado.get('datos', {}) if estado else {}
        nombre_herr = datos_herr.get('herr_nombre', incoming_msg.strip())
        obra_herr = incoming_msg.strip()
        borrar_estado(numero)
        ok, respuesta = herramienta_alta_obra(nombre_herr, obra_herr, numero, nombre_op)
        msg.body(respuesta)
        if ok:
            enviar_supervisor(f"🏗️ *Alta herramienta en obra*\n👷 {nombre_op}\n🔧 {nombre_herr}\n📍 {obra_herr}")
        return str(resp) if not use_meta else ('OK', 200)

    elif paso_herr == 'herr_baja_nombre':
        nombre_herr = incoming_msg.strip()
        set_dato(numero, 'herr_nombre', nombre_herr)
        set_paso(numero, 'herr_baja_obra')
        msg.body(f"📍 ¿De qué obra viene?\n_(Escribe el nombre de la obra)_")
        return str(resp) if not use_meta else ('OK', 200)

    elif paso_herr == 'herr_baja_obra':
        datos_herr = estado.get('datos', {}) if estado else {}
        nombre_herr = datos_herr.get('herr_nombre', '')
        obra_herr = incoming_msg.strip()
        borrar_estado(numero)
        ok, respuesta = herramienta_baja_obra(nombre_herr, None, numero, nombre_op)
        msg.body(respuesta)
        if ok:
            enviar_supervisor(f"🔙 *Devolución herramienta → almacén*\n👷 {nombre_op}\n🔧 {nombre_herr}\n📍 Obra: {obra_herr}")
        return str(resp) if not use_meta else ('OK', 200)

    elif paso_herr == 'herr_nueva_cantidad':
        datos_herr = estado.get('datos', {}) if estado else {}
        nombre_nueva = datos_herr.get('herr_nueva_nombre', '')
        try:
            cantidad = int(incoming_msg.strip())
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            msg.body("⚠️ Escribe un número válido de unidades.")
            return str(resp) if not use_meta else ('OK', 200)
        borrar_estado(numero)
        # Insertar o sumar al stock
        try:
            _c = get_db(); _cur = _c.cursor()
            _cur.execute("""
                INSERT INTO herramienta (nombre, tipo, stock_almacen)
                VALUES (%s, 'almacen', %s)
                ON CONFLICT (nombre) DO UPDATE
                  SET stock_almacen = herramienta.stock_almacen + EXCLUDED.stock_almacen
                RETURNING stock_almacen
            """, (nombre_nueva, cantidad))
            stock_total = _cur.fetchone()[0]
            _c.commit(); _cur.close(); _c.close()
            msg.body(f"✅ *{nombre_nueva.capitalize()}* añadida al almacén.\n📦 Stock actual: {int(stock_total)} ud.")
            enviar_supervisor(f"📦 *Nueva herramienta en almacén*\n🔧 {nombre_nueva.capitalize()}\n➕ {cantidad} ud. añadidas\n📦 Total: {int(stock_total)} ud.")
        except Exception as _e:
            msg.body(f"❌ Error al guardar: {_e}")
        return str(resp) if not use_meta else ('OK', 200)

    # ── Almacén: Listado PDF de stock ─────────────────────────────────────────
    msg_n = normalizar(incoming_msg)
    if msg_n.startswith('listado stock') or msg_n.startswith('pdf stock') or msg_n == 'listado':
        import re as _re
        familia_filtro = None
        # Extraer familia si viene: "listado stock cables", "pdf stock tubos"
        m_fam = _re.search(r'(?:listado stock|pdf stock|listado)\s+(.+)', msg_n)
        if m_fam:
            f_raw = m_fam.group(1).strip()
            if f_raw and f_raw not in ['todo', 'todos', 'completo']:
                familia_filtro = f_raw
        import threading as _th
        def _gen_listado():
            try:
                titulo = f"LISTADO STOCK — {familia_filtro.upper()}" if familia_filtro else "LISTADO COMPLETO DE STOCK"
                pdf_bytes = generar_pdf_stock(titulo=titulo, familia_filtro=familia_filtro)
                import hashlib as _hs
                ts = datetime.now().strftime('%Y%m%d%H%M%S')
                ref = f"STOCK-{ts}"
                subir_pdf_albaran(pdf_bytes, ref)
                pdf_url = f"https://bot-production-66b8.up.railway.app/albaran/{ref}.pdf"
                texto = f"📊 *{titulo}*\n\nListado generado."
                enviar_supervisor(texto, media_url=pdf_url)
            except Exception as e:
                enviar_supervisor(f"❌ Error generando listado: {e}")
        _th.Thread(target=_gen_listado, daemon=True).start()
        msg.body("⏳ Generando PDF de stock, te lo envío en un momento...")
        return str(resp) if not use_meta else ('OK', 200)

    # ── Almacén: Salida ───────────────────────────────────────────────────────
    if normalizar(incoming_msg).strip() in MENSAJES_STOCK_SALIDA:
        num_limpio = numero.replace('whatsapp:','').replace('+','').strip()
        nombre_conocido = OPERARIOS.get(num_limpio, '')
        if nombre_conocido:
            set_dato(numero, 'nombre_operario', nombre_conocido)
        set_dato(numero, 'stock_lineas', [])
        set_paso(numero, 'stock_salida_obra')
        msg.body(
            "📤 *SALIDA DE ALMACÉN*\n\n"
            "¿Para qué *obra o cliente* es esta salida?\n"
            "_Ejemplo: Ayuntamiento Los Llanos — Calle Real_"
        )
        return str(resp) if not use_meta else ('OK', 200)

    # ── Almacén: Devolución ───────────────────────────────────────────────────
    if normalizar(incoming_msg).strip() in MENSAJES_STOCK_DEVOL:
        num_limpio = numero.replace('whatsapp:','').replace('+','').strip()
        nombre_conocido = OPERARIOS.get(num_limpio, '')
        if nombre_conocido:
            set_dato(numero, 'nombre_operario', nombre_conocido)
        set_dato(numero, 'stock_lineas', [])
        set_paso(numero, 'stock_devol_obra')
        msg.body(
            "📥 *DEVOLUCIÓN A ALMACÉN*\n\n"
            "¿De qué obra procede el material?\n"
            "_Escribe el nombre de la obra_"
        )
        return str(resp) if not use_meta else ('OK', 200)

    # ── Almacén: Consulta de stock ────────────────────────────────────────────
    # ── Paso: consulta_menu (qué consultar) ──────────────────────────────────
    if paso_herr == 'consulta_menu':
        op = msg_n_herr.strip()
        if op == '1':
            # Stock Almacén → submenu buscar/PDF
            set_paso(numero, 'consulta_stock_almacen')
            msg.body(
                "📦 *Stock Almacén*\n\n"
                "1️⃣ Buscar por artículo\n"
                "2️⃣ Listado completo en PDF"
            )
            return str(resp) if not use_meta else ('OK', 200)
        elif op == '2':
            # Stock Herramienta → PDF directo
            borrar_estado(numero)
            try:
                url = _enviar_pdf_herramienta('almacen')
                msg.body(f"🔧 *Stock de herramienta:*\n{url}")
            except Exception as e:
                msg.body(f"❌ Error generando PDF: {e}")
            return str(resp) if not use_meta else ('OK', 200)
        else:
            msg.body(
                "🔍 *CONSULTA*\n\n"
                "1️⃣ Stock Almacén\n"
                "2️⃣ Stock Herramienta"
            )
            return str(resp) if not use_meta else ('OK', 200)

    # ── Paso: consulta_stock_almacen (buscar artículo / PDF) ─────────────────
    if paso_herr == 'consulta_stock_almacen':
        op = msg_n_herr.strip()
        if op == '1':
            borrar_estado(numero)
            set_paso(numero, 'stock_consulta')
            msg.body("🔍 ¿Qué artículo quieres consultar?\n_Escribe el nombre o parte de él_")
            return str(resp) if not use_meta else ('OK', 200)
        elif op == '2':
            borrar_estado(numero)
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.lib import colors
                from reportlab.lib.units import cm
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import ParagraphStyle
                from reportlab.lib.enums import TA_CENTER
                import io as _io
                from datetime import datetime as _dt
                _c = get_db(); _cur = _c.cursor()
                _cur.execute("""
                    SELECT nombre, stock_actual, unidad, familia
                    FROM stock_materiales
                    ORDER BY familia, nombre
                """)
                rows = _cur.fetchall()
                _cur.close(); _c.close()
                AZUL = colors.HexColor('#1a3a5c')
                GRIS = colors.HexColor('#f5f5f5')
                buffer = _io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4,
                    rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
                elements = []
                t_style = ParagraphStyle('T', fontSize=15, textColor=AZUL, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=6)
                pie_style = ParagraphStyle('P', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
                normal = ParagraphStyle('N', fontSize=9)
                fecha_str = _dt.now().strftime('%d/%m/%Y %H:%M')
                elements.append(Paragraph("INSTAPALMA — STOCK ALMACÉN ELÉCTRICO", t_style))
                elements.append(Paragraph(f"Generado: {fecha_str}", ParagraphStyle('F', fontSize=8, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=10)))
                elements.append(Spacer(1, 0.3*cm))
                if rows:
                    filas = [['Artículo', 'Familia', 'Stock', 'Ud.']]
                    for r in rows:
                        stock_val = r[1]
                        if stock_val is None:
                            stock_str = '0'
                        else:
                            stock_f = float(stock_val)
                            stock_str = (f"{stock_f:.2f}".rstrip('0').rstrip('.') if stock_f != int(stock_f) else str(int(stock_f))).replace('.', ',')
                        filas.append([r[0] or '', r[3] or '', stock_str, r[2] or ''])
                    t = Table(filas, colWidths=[8*cm, 3*cm, 2*cm, 2*cm])
                    t.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), AZUL),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0,0), (-1,-1), 8),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
                        ('PADDING', (0,0), (-1,-1), 5),
                        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, GRIS]),
                        ('ALIGN', (1,0), (1,-1), 'CENTER'),
                    ]))
                    elements.append(t)
                else:
                    elements.append(Paragraph("No hay artículos en el inventario.", normal))
                elements.append(Spacer(1, 0.5*cm))
                elements.append(Paragraph(f"Instapalma · {fecha_str}", pie_style))
                doc.build(elements)
                pdf_bytes = buffer.getvalue()
                import uuid as _uuid
                fname = f"stock_almacen_{_dt.now().strftime('%Y%m%d_%H%M%S')}_{_uuid.uuid4().hex[:6]}"
                try:
                    _c2 = get_db(); _cur2 = _c2.cursor()
                    _cur2.execute("""
                        INSERT INTO stock_albaranes (numero, pdf_bytes)
                        VALUES (%s, %s)
                        ON CONFLICT (numero) DO UPDATE SET pdf_bytes=%s
                    """, (fname, psycopg2.Binary(pdf_bytes), psycopg2.Binary(pdf_bytes)))
                    _c2.commit(); _cur2.close(); _c2.close()
                except Exception as _db_e:
                    print(f"Error guardando PDF almacén: {_db_e}")
                url = f"https://bot-production-66b8.up.railway.app/albaran/{fname}.pdf"
                msg.body(f"📦 *Listado completo de almacén:*\n{url}")
            except Exception as e:
                msg.body(f"❌ Error generando PDF: {e}")
            return str(resp) if not use_meta else ('OK', 200)
        else:
            msg.body(
                "📦 *Stock Almacén*\n\n"
                "1️⃣ Buscar por artículo\n"
                "2️⃣ Listado completo en PDF"
            )
            return str(resp) if not use_meta else ('OK', 200)



    # ── Paso: herr_stock_familia (el usuario elige familia) ───────────────────
    if paso_herr == 'herr_stock_familia':
        datos_h = estado.get('datos', {}) if estado else {}
        familias = datos_h.get('herr_familias', [])
        entrada = msg_n_herr.strip()
        # Puede ser número o texto
        familia_sel = None
        if entrada.isdigit():
            idx = int(entrada) - 1
            if 0 <= idx < len(familias):
                familia_sel = familias[idx]
        else:
            # Buscar coincidencia parcial
            for f in familias:
                if entrada in f or f in entrada:
                    familia_sel = f
                    break
        if not familia_sel:
            borrar_estado(numero)
            msg.body("⚠️ Familia no reconocida. Escribe `consulta` para volver a intentarlo.")
            return str(resp) if not use_meta else ('OK', 200)
        borrar_estado(numero)
        # Generar listado de esa familia
        try:
            _c = get_db(); _cur = _c.cursor()
            _cur.execute("""
                SELECT nombre, stock_almacen, observaciones
                FROM herramienta
                WHERE LOWER(nombre) LIKE %s
                  AND tipo != 'personal'
                ORDER BY nombre
            """, (f"%{familia_sel}%",))
            rows = _cur.fetchall()
            _cur.close(); _c.close()
            if not rows:
                msg.body(f"⚠️ No hay herramienta de la familia *{familia_sel.capitalize()}* en almacén.")
            else:
                lineas = [f"🔧 *{r[0]}*: {r[1]} ud." + (f"\n   _{r[2]}_" if r[2] else "") for r in rows]
                msg.body(f"📦 *Familia: {familia_sel.capitalize()}*\n\n" + "\n".join(lineas))
        except Exception as e:
            msg.body(f"❌ Error: {e}")
        return str(resp) if not use_meta else ('OK', 200)

    if normalizar(incoming_msg).strip() in MENSAJES_STOCK_CONSULTA:
        # Mostrar menú de consulta
        set_paso(numero, 'consulta_menu')
        msg.body(
            "🔍 *CONSULTA*\n\n"
            "1️⃣ Material de almacén\n"
            "2️⃣ Stock de herramienta"
        )
        return str(resp) if not use_meta else ('OK', 200)

    if False and normalizar(incoming_msg).strip() in []:  # bloque original desactivado
        # Intentar extraer el material de la misma frase
        msg_norm = normalizar(incoming_msg)
        busqueda = msg_norm
        for p in MENSAJES_STOCK_CONSULTA:
            busqueda = busqueda.replace(p, '').strip()
        if busqueda:
            mat, err = buscar_material_msg(busqueda)
            if err:
                if isinstance(err, tuple) and err[0] == 'RETALES':
                    candidatos = err[1]
                    lineas = []
                    for c in candidatos:
                        precio = float(c[5]) if len(c) > 5 and c[5] else 0
                        precio_txt = f" — {precio:.2f} €/ud".replace(".",",") if precio > 0 else ""
                        lineas.append(f"• *{c[1]}*: {c[3]} {c[2]}{precio_txt}")
                    msg.body("🔍 *Resultados encontrados:*\n\n" + "\n".join(lineas))
                else:
                    msg.body(err if isinstance(err, str) else str(err))
            else:
                stock = mat[3]; minimo = mat[4]; unidad = mat[2]; nombre_mat = mat[1]
                precio = float(mat[5]) if len(mat) > 5 and mat[5] else 0
                alerta = "\n⚠️ *Stock por debajo del mínimo*" if stock <= minimo and minimo > 0 else ""
                precio_txt = f"\nPrecio unitario: *{precio:.2f} €*" if precio > 0 else ""
                msg.body(f"🔍 *{nombre_mat}*\nStock actual: *{stock} {unidad}*\nStock mínimo: {minimo} {unidad}{precio_txt}{alerta}")
        else:
            set_paso(numero, 'stock_consulta')
            msg.body("🔍 ¿Qué material quieres consultar?\n_Escribe el nombre o parte de él_")
        return str(resp) if not use_meta else ('OK', 200)

    # Detectar resumen fin de mes
    if any(p in normalizar(incoming_msg) for p in MENSAJES_RESUMEN_MES):
        num_limpio = numero.replace('whatsapp:','').replace('+','').strip()
        nombre_conocido = OPERARIOS.get(num_limpio, '')
        if nombre_conocido:
            set_dato(numero, 'nombre_operario', nombre_conocido)
            set_paso(numero, 'resumen_mes')
        else:
            set_paso(numero, 'resumen_nombre')
        msg.body(
            "📊 *RESUMEN FIN DE MES — Instapalma*\n\n"
            + (f"Hola {nombre_conocido}! 👷\n\n" if nombre_conocido else "")
            + "1️⃣ ¿De qué *mes* es el resumen?\n_Ejemplo: Junio 2026_"
            if nombre_conocido else
            "📊 *RESUMEN FIN DE MES — Instapalma*\n\n"
            "Para empezar, ¿cuál es tu *nombre completo*?"
        )
        return str(resp) if not use_meta else ('OK', 200)

    # Detectar solicitud de vacaciones
    if any(p in normalizar(incoming_msg) for p in MENSAJES_VACACIONES):
        iniciar_vacaciones(numero)
        # Ver si el número está registrado
        num_limpio = numero.replace('whatsapp:','').replace('+','').strip()
        nombre_conocido = OPERARIOS.get(num_limpio, '')
        if nombre_conocido:
            set_dato(numero, 'nombre_operario', nombre_conocido)
            set_paso(numero, 'vac_inicio')
            saldo = get_saldo_vacaciones(numero.replace('whatsapp:','').replace('+','').strip())
            saldo_txt = f"\n📊 Días disponibles: *{saldo[0] - saldo[1]}* de {saldo[0]}" if saldo else ""
            msg.body(
                f"🌴 *Solicitud de Vacaciones — Instapalma*\n"
                f"Hola {nombre_conocido}!{saldo_txt}\n\n"
                f"1️⃣ ¿Cuál es la *fecha de inicio*?\n_Ejemplo: 14/07/2026_"
            )
        else:
            set_paso(numero, 'vac_nombre')
            msg.body(
                "🌴 *Solicitud de Vacaciones — Instapalma*\n\n"
                "Para continuar, ¿cuál es tu *nombre completo*?"
            )
        return str(resp) if not use_meta else ('OK', 200)

    # Gestión de aprobación por Alberto (APROBAR/RECHAZAR ID)
    if numero in [SUPERVISOR_WA, 'whatsapp:+34690875940']:
        msg_norm = normalizar(incoming_msg)
        import re as _re
        m = _re.match(r'^(aprobar|rechazar)\s+(\d+)$', msg_norm)
        if m:
            accion = m.group(1)
            vac_id = int(m.group(2))
            estado_nuevo = 'aprobada' if accion == 'aprobar' else 'rechazada'
            row = aprobar_rechazar_vacacion(vac_id, estado_nuevo)
            if row:
                op_num, op_nombre_bd, fi, ff, dias = row
                op_nombre = op_nombre_bd or nombre_operario(op_num)
                op_wa = op_num if op_num.startswith('whatsapp:') else f"whatsapp:{op_num}"
                num_limpio = op_num.replace('whatsapp:','').replace('+','').strip()
                saldo = get_saldo_vacaciones(num_limpio)
                if estado_nuevo == 'aprobada':
                    saldo_txt = f"\n📊 Te quedan *{saldo[0] - saldo[1]}* días de vacaciones." if saldo else ""
                    enviar_whatsapp(op_wa, f"✅ *Vacaciones aprobadas*\nTus vacaciones del {fi} al {ff} ({dias} días) han sido aprobadas por Alberto. ¡Disfrútalas! 🌴{saldo_txt}")
                    msg.body(f"✅ Vacaciones de {op_nombre} ({fi} – {ff}, {dias} días) aprobadas.")
                else:
                    enviar_whatsapp(op_wa, f"❌ *Vacaciones no aprobadas*\nTu solicitud del {fi} al {ff} no ha sido aprobada. Contacta con Alberto.")
                    msg.body(f"❌ Vacaciones de {op_nombre} rechazadas. El operario ha sido notificado.")
            else:
                msg.body(f"No encontré la solicitud #{vac_id}. Verifica el número.")
            return str(resp) if not use_meta else ('OK', 200)

    # ── Menú principal: "hola" ────────────────────────────────────────────────
    if normalizar(incoming_msg).strip() in MENSAJES_HOLA:
        set_paso(numero, 'menu_principal')
        msg.body(MENU_PRINCIPAL)
        return str(resp) if not use_meta else ('OK', 200)

    # ── Selección del menú principal ──────────────────────────────────────────
    if estado and estado.get('paso') == 'menu_principal':
        op = msg_n_herr.strip()
        borrar_estado(numero)
        if op == '1':
            iniciar_parte(numero)
            msg.body("👷 *Bot de Partes de Trabajo — Instapalma*\n\nVamos a crear tu parte paso a paso.\n\n1️⃣ ¿Cuál es la *fecha* del parte?\n\n_Escribe en formato DD/MM/YYYY o escribe *hoy* para usar la fecha de hoy ({})_".format(datetime.now().strftime('%d/%m/%Y')))
        elif op == '2':
            num_limpio2 = numero.replace('whatsapp:','').replace('+','').strip()
            nombre_conocido2 = OPERARIOS.get(num_limpio2, '')
            if nombre_conocido2:
                set_dato(numero, 'nombre_operario', nombre_conocido2)
            set_dato(numero, 'stock_lineas', [])
            set_paso(numero, 'stock_salida_obra')
            msg.body("📤 *SALIDA DE ALMACÉN*\n\n¿Para qué *obra o cliente* es esta salida?\n_Ejemplo: Ayuntamiento Los Llanos — Calle Real_")
        elif op == '3':
            num_limpio2 = numero.replace('whatsapp:','').replace('+','').strip()
            nombre_conocido2 = OPERARIOS.get(num_limpio2, '')
            if nombre_conocido2:
                set_dato(numero, 'nombre_operario', nombre_conocido2)
            set_dato(numero, 'stock_lineas', [])
            set_paso(numero, 'stock_devol_obra')
            msg.body("📥 *DEVOLUCIÓN A ALMACÉN*\n\n¿De qué obra procede el material?\n_Escribe el nombre de la obra_")
        elif op == '4':
            set_paso(numero, 'consulta_menu')
            msg.body("🔍 *CONSULTA*\n\n1️⃣ Material de almacén\n2️⃣ Stock de herramienta")
        elif op == '5':
            set_paso(numero, 'herr_menu')
            msg.body(MENU_HERRAMIENTA)
        elif op == '6':
            iniciar_vacaciones(numero)
            num_limpio2 = numero.replace('whatsapp:','').replace('+','').strip()
            nombre_conocido2 = OPERARIOS.get(num_limpio2, '')
            if nombre_conocido2:
                set_dato(numero, 'nombre_operario', nombre_conocido2)
                set_paso(numero, 'vac_inicio')
            else:
                set_paso(numero, 'vac_nombre')
            msg.body("🏖️ *Solicitud de Vacaciones*\n\n¿Cuál es tu nombre completo?" if not nombre_conocido2 else f"🏖️ *Solicitud de Vacaciones*\n\nHola {nombre_conocido2}! ¿Desde qué fecha?\n_Ejemplo: 01/08/2026_")
        elif op == '7':
            num_limpio2 = numero.replace('whatsapp:','').replace('+','').strip()
            nombre_conocido2 = OPERARIOS.get(num_limpio2, '')
            if nombre_conocido2:
                set_dato(numero, 'nombre_operario', nombre_conocido2)
                set_paso(numero, 'resumen_mes')
            else:
                set_paso(numero, 'resumen_nombre')
            msg.body("📊 *RESUMEN FIN DE MES*\n\n¿De qué mes es el resumen?\n_Ejemplo: Junio 2026_" if nombre_conocido2 else "📊 *RESUMEN FIN DE MES*\n\n¿Cuál es tu nombre completo?")
        elif op == '8':
            set_paso(numero, 'vehiculo_menu')
            msg.body("🚗 *Vehículos*\n\nEscribe *vehiculo* para acceder al módulo de mantenimiento.")
        else:
            msg.body(MENU_PRINCIPAL)
        return str(resp) if not use_meta else ('OK', 200)

    if not estado:
        if any(p in normalizar(incoming_msg) for p in MENSAJES_INICIO):
            iniciar_parte(numero)
            msg.body(
                "👷 *Bot de Partes de Trabajo — Instapalma*\n\n"
                "Vamos a crear tu parte paso a paso.\n\n"
                "1️⃣ ¿Cuál es la *fecha* del parte?\n\n"
                "_Escribe en formato DD/MM/YYYY o escribe *hoy* para usar la fecha de hoy ({})_".format(datetime.now().strftime('%d/%m/%Y'))
            )
        else:
            msg.body("Hola 👋 Escribe *hola* para ver el menú o *parte* para crear un parte de trabajo.")
        return str(resp)

    paso = estado['paso']
    datos = estado['datos']

    if paso == 'fecha':
        texto = incoming_msg.strip()
        if normalizar(texto) in ['hoy', 'today']:
            fecha_val = datetime.now().strftime('%d/%m/%Y')
        else:
            import re as _re
            # Acepta DD/MM/YYYY, DD-MM-YYYY, DD/MM/YY
            m = _re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$', texto)
            if m:
                d, mo, y = m.group(1), m.group(2), m.group(3)
                if len(y) == 2:
                    y = '20' + y
                try:
                    datetime.strptime(f"{d.zfill(2)}/{mo.zfill(2)}/{y}", '%d/%m/%Y')
                    fecha_val = f"{d.zfill(2)}/{mo.zfill(2)}/{y}"
                except ValueError:
                    msg.body("⚠️ Fecha no válida. Escribe la fecha en formato *DD/MM/YYYY* o escribe *hoy*.")
                    return str(resp)
            else:
                msg.body(
                    f"1️⃣ ¿Cuál es la *fecha* del parte?\n\n"
                    f"Escribe la fecha en formato *DD/MM/YYYY*\n"
                    f"_(o escribe *hoy* para usar la fecha de hoy: {datetime.now().strftime('%d/%m/%Y')})_"
                )
                return str(resp)
        set_dato(numero, 'fecha', fecha_val)
        set_paso(numero, 'cliente')
        msg.body("2️⃣ ¿Cuál es el *cliente*?")

    elif paso == 'cliente':
        set_dato(numero, 'cliente', incoming_msg.upper())
        set_paso(numero, 'obra')
        msg.body("3️⃣ ¿Cuál es la *obra*?")

    elif paso == 'obra':
        set_dato(numero, 'obra', incoming_msg.upper())
        set_paso(numero, 'operarios')
        msg.body(
            "4️⃣ *Operarios y horas*\n\n"
            "Escribe cada operario en una línea:\n"
            "_Ejemplo:_\n"
            "JORGE GARCIA — 8h\n"
            "ANTONIO — 6h"
        )

    elif paso == 'operarios':
        import re
        lineas = [l.strip() for l in incoming_msg.strip().split('\n') if l.strip()]
        errores = []
        for linea in lineas:
            # Debe tener texto (nombre) Y número de horas
            tiene_nombre = bool(re.search(r'[a-záéíóúüñA-ZÁÉÍÓÚÜÑ]', linea))
            tiene_horas  = bool(re.search(r'\d+[\.,]?\d*\s*h(?:oras?|rs?)?', linea, re.IGNORECASE))
            if not tiene_nombre or not tiene_horas:
                errores.append(linea)
        if not lineas:
            msg.body(
                "⚠️ No has escrito ningún operario.\n\n"
                "Escribe uno por línea:\n"
                "_NOMBRE — 8h_"
            )
        elif errores:
            lista = '\n'.join(f'• {l}' for l in errores)
            msg.body(
                f"⚠️ Formato incorrecto en:\n{lista}\n\n"
                "Cada línea debe tener *nombre* y *horas*:\n"
                "_JUAN GARCÍA — 8h_\n"
                "_ANTONIO — 6.5h_\n\n"
                "Escribe de nuevo la lista completa:"
            )
        else:
            set_dato(numero, 'operarios', incoming_msg)
            set_paso(numero, 'albaranes')
            msg.body(
                "5️⃣ *Albaranes*\n\n"
                "Escribe los albaranes como quieras, uno por línea.\n"
                "_Ejemplo:_\n"
                "DIEXFE 012604\n"
                "COELCA 9981\n\n"
                "Si no hay, escribe: *ninguno*"
            )

    elif paso == 'albaranes':
        if normalizar(incoming_msg) in ['ninguno', 'no', 'n']:
            set_dato(numero, 'albaranes', 'Ninguno')
        else:
            set_dato(numero, 'albaranes', incoming_msg)
        set_paso(numero, 'material_stock')
        msg.body(
            "6️⃣ *Material de stock* utilizado\n\n"
            "Escribe el material, uno por línea:\n"
            "_Ejemplo:_\n"
            "Cable 2.5mm² — 20m\n"
            "Caja superficie — 2ud\n\n"
            "Si no hay, escribe: *ninguno*"
        )

    elif paso == 'material_stock':
        val = incoming_msg if normalizar(incoming_msg) != 'ninguno' else 'Ninguno'
        set_dato(numero, 'material_stock', val)
        set_paso(numero, 'devolucion_almacen')
        msg.body(
            "7️⃣ *Devolución a Almacén*\n\n"
            "¿Devuelves algún material al almacén?\n"
            "_Ejemplo: Cable 2.5mm² — 10m sobrantes_\n\n"
            "Si no hay, escribe: *ninguno*"
        )

    elif paso == 'devolucion_almacen':
        val = incoming_msg if normalizar(incoming_msg) != 'ninguno' else 'Ninguno'
        set_dato(numero, 'devolucion_almacen', val)
        set_paso(numero, 'descripcion')
        msg.body("8️⃣ *Descripción* de los trabajos realizados:")

    elif paso == 'descripcion':
        set_dato(numero, 'descripcion', incoming_msg)
        set_paso(numero, 'terminado')
        msg.body("9️⃣ ¿El trabajo está *terminado*?\n\nResponde *SÍ* o *NO*")

    elif paso == 'terminado':
        if normalizar(incoming_msg) in ['si', 'sí', 's', 'yes']:
            set_dato(numero, 'terminado', 'Sí')
            set_dato(numero, 'tiempo_restante', '')
            set_paso(numero, 'confirmar')
            msg.body(generar_resumen(get_estado(numero)['datos']))
        elif normalizar(incoming_msg) in ['no', 'n']:
            set_dato(numero, 'terminado', 'No')
            set_paso(numero, 'tiempo_restante')
            msg.body("🔟 ¿Cuánto tiempo queda para terminarlo?\n\n_Ejemplo: 2 días, media jornada, 3 horas..._")
        else:
            msg.body("Responde *SÍ* si está terminado o *NO* si falta trabajo.")

    elif paso == 'tiempo_restante':
        set_dato(numero, 'tiempo_restante', incoming_msg)
        set_paso(numero, 'confirmar')
        msg.body(generar_resumen(get_estado(numero)['datos']))

    elif paso == 'confirmar':
        if es_confirmacion(incoming_msg):
            finalizar_parte(numero, datos)
            msg.body(
                "✅ *Parte enviado correctamente.*\n\n"
                "Se ha notificado al supervisor por WhatsApp y email con PDF. ¡Gracias!"
            )
        elif es_cancelacion(incoming_msg):
            borrar_estado(numero)
            msg.body("❌ Parte cancelado. Escribe *parte* para crear uno nuevo.")
        else:
            msg.body("Responde *SÍ* para confirmar y enviar, o *NO* para cancelar.")

    # ── Flujo vehículos ────────────────────────────────────────────────────
    elif paso == 'v_matricula':
        set_dato(numero, 'matricula', incoming_msg.upper())
        set_paso(numero, 'v_modelo')
        msg.body("2️⃣ *Marca y modelo* del vehículo:\n_Ejemplo: Ford Transit 2020_")

    elif paso == 'v_modelo':
        set_dato(numero, 'marca_modelo', incoming_msg)
        set_paso(numero, 'v_mes')
        msg.body("3️⃣ ¿A qué *mes* corresponde este parte?\n_Ejemplo: Junio 2026_")

    elif paso == 'v_mes':
        set_dato(numero, 'mes', incoming_msg)
        set_paso(numero, 'v_km_inicio')
        msg.body("4️⃣ *Km al inicio del mes*:")

    elif paso == 'v_km_inicio':
        set_dato(numero, 'km_inicio', incoming_msg)
        set_paso(numero, 'v_km_fin')
        msg.body("5️⃣ *Km al final del mes*:")

    elif paso == 'v_km_fin':
        set_dato(numero, 'km_fin', incoming_msg)
        set_paso(numero, 'v_aceite')
        msg.body("6️⃣ ¿A cuántos km es el *próximo cambio de aceite*?\n_Ejemplo: 85000_")

    elif paso == 'v_aceite':
        set_dato(numero, 'proximo_aceite', incoming_msg)
        set_paso(numero, 'v_neumaticos')
        msg.body(
            "7️⃣ *Estado de los neumáticos*\n\n"
            "_Ejemplo: Delanteros OK, traseros con desgaste_\n"
            "Si están bien, escribe: *OK*"
        )

    elif paso == 'v_neumaticos':
        set_dato(numero, 'estado_neumaticos', incoming_msg)
        set_paso(numero, 'v_conductores')
        msg.body(
            "8️⃣ *Conductores del mes*\n\n"
            "Escribe uno por línea con fechas:\n"
            "_NOMBRE - desde - hasta_\n"
            "Ejemplo:\n"
            "JUAN GARCÍA - 01/06 - 15/06\n"
            "PEDRO MARTÍN - 16/06 - 30/06\n\n"
            "Si solo hay uno todo el mes escribe el nombre."
        )

    elif paso == 'v_conductores':
        set_dato(numero, 'conductores', incoming_msg)
        set_paso(numero, 'v_mantenimientos')
        msg.body(
            "9️⃣ *Mantenimientos realizados*\n\n"
            "Escribe uno por línea:\n"
            "_CONCEPTO - FECHA - KM_\n"
            "Ejemplo:\n"
            "Cambio aceite - 10/06/2026 - 82000\n"
            "Revisión frenos - 20/06/2026 - 82500\n\n"
            "Si no hay, escribe: *ninguno*"
        )

    elif paso == 'v_mantenimientos':
        val = incoming_msg if normalizar(incoming_msg) != 'ninguno' else 'Ninguno'
        set_dato(numero, 'mantenimientos', val)
        set_paso(numero, 'v_observaciones')
        msg.body(
            "🔟 *Observaciones y próximos mantenimientos*\n\n"
            "Escribe lo que necesites.\n"
            "Si no hay, escribe: *ninguno*"
        )

    elif paso == 'v_observaciones':
        val = incoming_msg if normalizar(incoming_msg) != 'ninguno' else 'Ninguno'
        set_dato(numero, 'observaciones', val)
        set_paso(numero, 'v_golpes')
        msg.body(
            "1️⃣1️⃣ *Golpes y desperfectos*\n\n"
            "Describe cualquier daño visible.\n"
            "Si no hay, escribe: *ninguno*"
        )

    elif paso == 'v_golpes':
        val = incoming_msg if normalizar(incoming_msg) != 'ninguno' else 'Ninguno'
        set_dato(numero, 'golpes', val)
        set_paso(numero, 'v_confirmar')
        msg.body(generar_resumen_vehiculo(get_estado(numero)['datos']))

    elif paso == 'v_confirmar':
        if es_confirmacion(incoming_msg):
            finalizar_vehiculo(numero, datos)
            msg.body("✅ *Parte de vehículo enviado.* Se ha notificado al supervisor por WhatsApp y email. ¡Gracias!")
        elif es_cancelacion(incoming_msg):
            borrar_estado(numero)
            msg.body("❌ Parte cancelado. Escribe *vehiculo* para crear uno nuevo.")
        else:
            msg.body("Responde *SÍ* para confirmar y enviar, o *NO* para cancelar.")

    # ── Flujo Resumen Fin de Mes ──────────────────────────────────────────────
    elif paso == 'resumen_nombre':
        set_dato(numero, 'nombre_operario', incoming_msg)
        set_paso(numero, 'resumen_mes')
        msg.body(f"👋 Hola *{incoming_msg}*!\n\n1️⃣ ¿De qué *mes* es el resumen?\n_Ejemplo: Junio 2026_")

    elif paso == 'resumen_mes':
        set_dato(numero, 'mes', incoming_msg)
        set_paso(numero, 'resumen_horas')
        msg.body("2️⃣ ¿Cuántas *horas extra* has realizado este mes?\n_Ejemplo: 8_")

    elif paso == 'resumen_horas':
        set_dato(numero, 'horas_extra', incoming_msg)
        set_paso(numero, 'resumen_vacaciones')
        msg.body("3️⃣ ¿Cuántos *días de vacaciones* has disfrutado este mes?\n_Ejemplo: 5_")

    elif paso == 'resumen_vacaciones':
        set_dato(numero, 'dias_vacaciones', incoming_msg)
        set_paso(numero, 'resumen_gastos')
        msg.body("4️⃣ ¿Cuál es el *total de gastos* del mes? (en €)\n_Ejemplo: 127.50_")

    elif paso == 'resumen_gastos':
        set_dato(numero, 'total_gastos', incoming_msg)
        set_paso(numero, 'resumen_foto')
        msg.body(
            "5️⃣ Adjunta la *foto del justificante* 📎\n"
            "_Es obligatoria para enviar el resumen._"
        )

    elif paso == 'resumen_foto':
        # Foto obligatoria — si no hay imagen, pedir de nuevo
        foto_url = media_url if media_url else ''
        if not foto_url:
            msg.body(
                "⚠️ Necesito que adjuntes la *foto del justificante*.\n"
                "Por favor envía la imagen para continuar."
            )
        else:
            set_dato(numero, 'foto_url', foto_url)
            set_paso(numero, 'resumen_confirmar')
            datos_r = get_estado(numero)['datos']
            msg.body(
                f"📊 *RESUMEN FIN DE MES*\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"📅 Mes: {datos_r.get('mes','')}\n"
                f"👷 {datos_r.get('nombre_operario','')}\n"
                f"⏱ Horas extra: {datos_r.get('horas_extra','0')}\n"
                f"🌴 Días vacaciones: {datos_r.get('dias_vacaciones','0')}\n"
                f"💶 Total gastos: {datos_r.get('total_gastos','0')} €\n"
                f"📎 Foto adjunta ✅\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"¿Es correcto? Responde *SÍ* o *NO*"
            )

    elif paso == 'resumen_confirmar':
        if es_confirmacion(incoming_msg):
            datos_r = get_estado(numero)['datos']
            borrar_estado(numero)
            finalizar_resumen_mes(numero, datos_r)
            msg.body("✅ Resumen enviado. Te llegará el PDF en un momento.")
        elif es_cancelacion(incoming_msg):
            borrar_estado(numero)
            msg.body("❌ Cancelado. Escribe *resumen mes* para empezar de nuevo.")
        else:
            msg.body("Responde *SÍ* para confirmar y enviar, o *NO* para cancelar.")

    # ── Flujo Almacén: Salida ─────────────────────────────────────────────────
    elif paso == 'stock_salida_obra':
        set_dato(numero, 'stock_obra', incoming_msg)
        set_paso(numero, 'stock_salida_material')
        msg.body(
            "📦 ¿Qué *material* retiras?\n"
            "_Escribe el nombre del material_"
        )

    elif paso == 'stock_salida_material':
        if normalizar(incoming_msg) in ['listo', 'fin', 'terminar', 'acabar', 'ya', 'eso es todo']:
            # Pasar a confirmación
            datos_s = get_estado(numero)['datos']
            lineas = datos_s.get('stock_lineas', [])
            if not lineas:
                msg.body("No has añadido ningún material. Dime qué material retiras.")
            else:
                set_paso(numero, 'stock_salida_confirmar')
                resumen = '\n'.join([f"• {l['material']} — {fmt_cant(l['cantidad'])} {l['unidad']}" for l in lineas])
                msg.body(
                    f"📤 *RESUMEN SALIDA*\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"🏢 Obra: {datos_s.get('stock_obra','')}\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"{resumen}\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"¿Es correcto? Responde *SÍ* o *NO*"
                )
        else:
            # ── Selección por número si hay candidatos múltiples guardados ──
            datos_s = get_estado(numero)['datos']
            candidatos_prev = datos_s.get('stock_multiples_candidatos', [])
            mat = None; err = None
            if candidatos_prev and incoming_msg.strip().isdigit():
                idx = int(incoming_msg.strip()) - 1
                if 0 <= idx < len(candidatos_prev):
                    c = candidatos_prev[idx]
                    mat = tuple(c)
                    set_dato(numero, 'stock_multiples_candidatos', [])
                else:
                    err = f"❌ Número fuera de rango. Elige entre 1 y {len(candidatos_prev)}."
            else:
                set_dato(numero, 'stock_multiples_candidatos', [])
                mat, err = buscar_material_msg(incoming_msg)
            if err:
                if isinstance(err, tuple) and err[0] == 'RETALES':
                    # Es un grupo de retales — mostrar lista y elegir directamente
                    candidatos = err[1]
                    candidatos_serial = [[c[0], c[1], c[2], float(c[3]), float(c[4])] for c in candidatos]
                    set_dato(numero, 'stock_retales_candidatos', candidatos_serial)
                    set_dato(numero, 'stock_mat_busqueda', incoming_msg)
                    # Construir sugerencia con todos los metros disponibles
                    sugerencia_previa = sugerir_retales(candidatos, 9999)
                    set_dato(numero, 'stock_retales_sugerencia', sugerencia_previa)
                    set_paso(numero, 'stock_salida_retales_elegir')
                    retales_info = sugerencia_previa['retales']
                    total_m = sugerencia_previa['total_disponible']
                    lista_r = '\n'.join([f"  *{i}.* {r['nombre']} — {fmt_cant(r['metros'])} m" for i, r in enumerate(retales_info, 1)])
                    msg.body(
                        f"📏 *Retales disponibles de {incoming_msg}:*\n{lista_r}\n"
                        f"Total: *{fmt_cant(total_m)} m*\n\n"
                        f"¿Cuál coges? Responde el *número* o el nombre."
                    )
                elif isinstance(err, tuple) and err[0] == 'MULTIPLES':
                    candidatos = err[1]
                    texto = err[2]
                    candidatos_serial = [[c[0], c[1], c[2], float(c[3]), float(c[4]), float(c[5]) if len(c) > 5 and c[5] else 0] for c in candidatos]
                    set_dato(numero, 'stock_multiples_candidatos', candidatos_serial)
                    set_paso(numero, 'stock_salida_material')
                    msg.body(texto)
                else:
                    set_paso(numero, 'stock_salida_material')
                    msg.body(err if isinstance(err, str) else str(err))
            else:
                set_dato(numero, 'stock_mat_tmp', {'id': mat[0], 'nombre': mat[1], 'unidad': mat[2], 'stock': float(mat[3]), 'precio': float(mat[5]) if len(mat) > 5 and mat[5] else 0})
                set_paso(numero, 'stock_salida_cantidad')
                msg.body(
                    f"📦 *{mat[1]}*\n"
                    f"Stock disponible: *{mat[3]} {mat[2]}*\n\n"
                    f"¿Qué *cantidad* retiras?"
                )

    elif paso == 'stock_salida_retales_elegir':
        datos_s = get_estado(numero)['datos']
        sugerencia = datos_s.get('stock_retales_sugerencia', {})
        retales_disp = sugerencia.get('retales', [])
        print(f"[DEBUG retales_elegir] sugerencia keys={list(sugerencia.keys()) if sugerencia else 'VACIO'}, retales_disp len={len(retales_disp)}, datos_keys={list(datos_s.keys()) if datos_s else 'NONE'}")
        # Si no hay sugerencia guardada, reconstruirla desde candidatos
        if not retales_disp:
            candidatos_raw = datos_s.get('stock_retales_candidatos', [])
            if candidatos_raw:
                candidatos = [tuple(c) for c in candidatos_raw]
                sugerencia = sugerir_retales(candidatos, 9999)
                set_dato(numero, 'stock_retales_sugerencia', sugerencia)
                retales_disp = sugerencia.get('retales', [])
                print(f"[DEBUG retales_elegir] Reconstruido: retales_disp len={len(retales_disp)}")
        try:
            import re as _re
            txt_orig = incoming_msg.strip()

            # Detectar si viene con metros incluidos al final: "M20 15m", "1 20m"
            metros_inline = None
            patron_metros = _re.search(r'[,\s]+(\d+[.,]?\d*)\s*m(?:etros?)?\s*$', txt_orig, _re.IGNORECASE)
            if patron_metros:
                metros_inline = float(patron_metros.group(1).replace(',', '.'))
                txt_busq = txt_orig[:patron_metros.start()].strip().upper()
            else:
                txt_busq = txt_orig.upper()

            # Intentar selección por número de lista
            seleccionados_idx = []
            nums_str = _re.findall(r'\b(\d+)\b', txt_busq)
            seleccionados_idx = [int(n)-1 for n in nums_str if 0 < int(n) <= len(retales_disp)]

            # Si no, intentar por descripción textual
            if not seleccionados_idx:
                for i, r in enumerate(retales_disp):
                    nombre_norm = r['nombre'].upper()
                    palabras = [p for p in nombre_norm.split() if len(p) > 2]
                    if any(p in txt_busq for p in palabras):
                        seleccionados_idx.append(i)

            if not seleccionados_idx:
                lista_r = '\n'.join([f"  *{i+1}.* {r['nombre']} — {fmt_cant(r['metros'])} m" for i, r in enumerate(retales_disp)])
                msg.body(f"No encontré ese retal. Disponibles:\n{lista_r}\n\nResponde el *número* o escribe parte del nombre (ej: *M20*).")
            elif len(seleccionados_idx) == 1:
                idx = seleccionados_idx[0]
                r = retales_disp[idx]
                if metros_inline is not None:
                    # Ya viene con metros en el mensaje → procesar directamente
                    if metros_inline <= 0:
                        msg.body("Los metros deben ser mayor que 0.")
                    elif metros_inline > r['metros']:
                        msg.body(
                            f"⚠️ Solo hay *{fmt_cant(r['metros'])} m* de *{r['nombre']}*. "
                            f"¿Cuántos metros necesitas? (máx. {fmt_cant(r['metros'])})"
                        )
                        set_dato(numero, 'stock_retal_elegido_idx', idx)
                        set_paso(numero, 'stock_salida_retales_metros_usar')
                    else:
                        lineas = datos_s.get('stock_lineas', [])
                        lineas.append({
                            'material': r['nombre'],
                            'cantidad': metros_inline,
                            'unidad': r['unidad'],
                            'material_id': r['id'],
                            'delta_stock': -metros_inline,
                            'precio': r.get('precio', 0)
                        })
                        set_dato(numero, 'stock_lineas', lineas)
                        set_paso(numero, 'stock_salida_material')
                        msg.body(
                            f"✅ Añadido: *{r['nombre']}* — {fmt_cant(metros_inline)} m\n\n"
                            f"¿Otro material? Escribe el nombre o di *listo* para terminar."
                        )
                else:
                    # Preguntar metros
                    set_dato(numero, 'stock_retal_elegido_idx', idx)
                    set_paso(numero, 'stock_salida_retales_metros_usar')
                    msg.body(
                        f"📐 *{r['nombre']}* — {fmt_cant(r['metros'])} m disponibles\n\n"
                        f"¿Cuántos metros necesitas?"
                    )
            else:
                # Varios retales: preguntar metros para cada uno secuencialmente
                # Por ahora usar todos sus metros disponibles y añadir directamente
                lineas = datos_s.get('stock_lineas', [])
                for idx in seleccionados_idx:
                    r = retales_disp[idx]
                    lineas.append({
                        'material': r['nombre'],
                        'cantidad': r['metros'],
                        'unidad': r['unidad'],
                        'material_id': r['id'],
                        'delta_stock': -r['metros'],
                        'precio': r.get('precio', 0)
                    })
                set_dato(numero, 'stock_lineas', lineas)
                set_paso(numero, 'stock_salida_material')
                añadidos = ', '.join([retales_disp[i]['nombre'] for i in seleccionados_idx])
                msg.body(
                    f"✅ Añadido: *{añadidos}*\n\n"
                    f"¿Otro material? Escribe el nombre o di *listo* para terminar."
                )
        except:
            msg.body(f"Responde el número del retal. Ejemplo: *1* o *1,2*")

    elif paso == 'stock_salida_retales_metros_usar':
        try:
            import re as _re
            txt_mu = incoming_msg.strip().replace(',', '.')
            try:
                metros_usar = float(txt_mu)
            except:
                nums_e = _re.findall(r'\b(\d+(?:\.\d+)?)\b', txt_mu)
                metros_usar = float(nums_e[-1]) if nums_e else None
                if metros_usar is None:
                    raise ValueError("sin número")
            datos_s = get_estado(numero)['datos']
            sugerencia = datos_s.get('stock_retales_sugerencia', {})
            retales_disp = sugerencia.get('retales', [])
            idx = int(datos_s.get('stock_retal_elegido_idx', 0))
            r = retales_disp[idx]
            if metros_usar <= 0:
                msg.body("Los metros deben ser mayor que 0.")
            elif metros_usar > r['metros']:
                msg.body(
                    f"⚠️ Solo hay *{fmt_cant(r['metros'])} m* de este retal. "
                    f"¿Cuántos metros necesitas? (máx. {fmt_cant(r['metros'])})"
                )
            else:
                lineas = datos_s.get('stock_lineas', [])
                lineas.append({
                    'material': r['nombre'],
                    'cantidad': metros_usar,
                    'unidad': r['unidad'],
                    'material_id': r['id'],
                    'delta_stock': -metros_usar,
                    'precio': r.get('precio', 0)
                })
                set_dato(numero, 'stock_lineas', lineas)
                set_paso(numero, 'stock_salida_material')
                msg.body(
                    f"✅ Añadido: *{r['nombre']}* — {fmt_cant(metros_usar)} m\n\n"
                    f"¿Otro material? Escribe el nombre o di *listo* para terminar."
                )
        except:
            msg.body("Escribe los metros que necesitas. Ejemplo: *15* o *7,5*")

    elif paso == 'stock_salida_cantidad':
        try:
            cantidad = float(incoming_msg.replace(',','.'))
            datos_s = get_estado(numero)['datos']
            mat_tmp = datos_s.get('stock_mat_tmp', {})
            if cantidad <= 0:
                msg.body("La cantidad debe ser mayor que 0.")
            elif cantidad > mat_tmp.get('stock', 0):
                msg.body(f"⚠️ Solo hay *{mat_tmp.get('stock',0)} {mat_tmp.get('unidad','')}* disponibles. Indica una cantidad menor.")
            else:
                lineas = datos_s.get('stock_lineas', [])
                lineas.append({'material': mat_tmp['nombre'], 'cantidad': cantidad, 'unidad': mat_tmp['unidad'], 'material_id': mat_tmp['id'], 'precio': mat_tmp.get('precio', 0)})
                set_dato(numero, 'stock_lineas', lineas)
                set_paso(numero, 'stock_salida_material')
                msg.body(
                    f"✅ Añadido: *{mat_tmp['nombre']}* — {cantidad} {mat_tmp['unidad']}\n\n"
                    f"¿Otro material? Escribe el nombre o di *listo* para terminar."
                )
        except:
            msg.body("Escribe solo el número de la cantidad. Ejemplo: *25* o *0.5*")

    elif paso == 'stock_salida_confirmar':
        if es_confirmacion(incoming_msg):
            datos_s = get_estado(numero)['datos']
            lineas = datos_s.get('stock_lineas', [])
            obra = datos_s.get('stock_obra', '')
            nombre_op = datos_s.get('nombre_operario', nombre_operario(numero))
            borrar_estado(numero)
            import threading as _th
            def _procesar_salida():
                numero_alb = siguiente_numero_albaran()
                from datetime import datetime as dt
                fecha_str = dt.now().strftime('%d/%m/%Y %H:%M')
                # Crear albarán en DB
                aid = crear_albaran(numero_alb, numero, nombre_op, obra, lineas)
                # Ajustar stock y registrar movimientos
                alertas = []
                for l in lineas:
                    # Si es un retal, el delta de stock es -1 (sacar el retal entero); si no, -cantidad
                    delta = l.get('delta_stock', -l['cantidad'])
                    r = ajustar_stock(l['material_id'], delta)
                    registrar_movimiento('salida', l['material_id'], l['material'], l['cantidad'],
                        l['unidad'], numero, nombre_op, obra, aid)
                    if r and r[0] <= r[1] and r[1] > 0:
                        alertas.append(f"⚠️ {r[2]}: quedan {r[0]} {r[3]} (mínimo {r[1]})")
                # Generar PDF
                pdf_bytes_sal = generar_pdf_albaran({
                    'numero': numero_alb, 'nombre_operario': nombre_op,
                    'obra': obra, 'lineas': lineas, 'fecha': fecha_str
                })
                pdf_url = subir_pdf_albaran(pdf_bytes_sal, numero_alb)
                # Enviar al operario
                op_wa = numero if numero.startswith('whatsapp:') else f'whatsapp:+{numero.lstrip("+")}'
                resumen_txt = '\n'.join([f"• {l['material']} — {fmt_cant(l['cantidad'])} {l['unidad']}" for l in lineas])
                texto = (
                    f"✅ *Albarán {numero_alb}*\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"🏢 {obra}\n{resumen_txt}\n"
                    f"━━━━━━━━━━━━━━━━━━━━"
                    + (f"\n📄 PDF: {pdf_url}" if pdf_url else "")
                )
                enviar_whatsapp(op_wa, texto)
                # Enviar al supervisor
                enviar_supervisor(f"📤 *Salida almacén — {numero_alb}*\n👷 {nombre_op}\n🏢 {obra}\n{resumen_txt}"
                    + (f"\n📄 PDF: {pdf_url}" if pdf_url else ""))
                # Email con PDF adjunto
                try:
                    nombre_pdf_sal = f"Albaran_Salida_{numero_alb.replace('/','_').replace('-','_')}.pdf"
                    msg_email_sal = MIMEMultipart()
                    msg_email_sal['From']    = GMAIL_USER
                    msg_email_sal['To']      = ', '.join([SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2])
                    msg_email_sal['Subject'] = f"[SALIDA ALMACÉN] {numero_alb} — {nombre_op} — {obra}"
                    body_sal = (
                        f"Salida de almacén registrada.\n\n"
                        f"Albarán: {numero_alb}\n"
                        f"Operario: {nombre_op}\n"
                        f"Obra: {obra}\n"
                        f"Fecha: {fecha_str}\n\n"
                        f"Materiales:\n" +
                        '\n'.join([f"  - {l['material']}: {fmt_cant(l['cantidad'])} {l['unidad']}" for l in lineas])
                    )
                    msg_email_sal.attach(MIMEText(body_sal, 'plain'))
                    part_sal = MIMEApplication(pdf_bytes_sal, Name=nombre_pdf_sal)
                    part_sal.add_header('Content-Disposition', f'attachment; filename="{nombre_pdf_sal}"')
                    msg_email_sal.attach(part_sal)
                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as srv:
                        srv.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                        srv.sendmail(GMAIL_USER, [SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2], msg_email_sal.as_string())
                    print(f"Email salida almacén OK — {numero_alb}")
                except Exception as e_mail_sal:
                    print(f"Error email salida almacén: {e_mail_sal}")
                # Alertas de stock bajo
                for a in alertas:
                    enviar_supervisor(a)
            _th.Thread(target=_procesar_salida, daemon=True).start()
            msg.body(f"✅ Salida registrada. Te envío el albarán en un momento.")
        elif es_cancelacion(incoming_msg):
            borrar_estado(numero)
            msg.body("❌ Cancelado.")
        else:
            msg.body("Responde *SÍ* para confirmar o *NO* para cancelar.")

    # ── Flujo Almacén: Devolución ─────────────────────────────────────────────
    elif paso == 'stock_devol_obra':
        # Guarda la obra de procedencia y pasa a pedir material
        set_dato(numero, 'stock_devol_obra_txt', incoming_msg.strip())
        set_paso(numero, 'stock_devol_material')
        msg.body(
            f"✅ Obra: *{incoming_msg.strip()}*\n\n"
            f"¿Qué material devuelves?\n"
            f"_Escribe nombre y cantidad. Ej: Cable RV-K 3x2.5 — 12 m_\n"
            f"_O escribe el nombre y te pido la cantidad_\n\n"
            f"Di *listo* cuando hayas añadido todo."
        )

    elif paso == 'stock_devol_material':
        # Flujo simplificado: texto libre, sin buscar en BD ni retales
        if normalizar(incoming_msg) in ['listo', 'fin', 'terminar', 'acabar', 'ya', 'eso es todo']:
            datos_d = get_estado(numero)['datos']
            lineas = datos_d.get('stock_lineas', [])
            if not lineas:
                msg.body("No has añadido ningún material. Dime qué devuelves.")
            else:
                set_paso(numero, 'stock_devol_confirmar')
                obra_proc = datos_d.get('stock_devol_obra_txt', 'Sin especificar')
                resumen = '\n'.join([f"• {l['material']} — {fmt_cant(l['cantidad'])} {l['unidad']}" for l in lineas])
                msg.body(
                    f"📥 *RESUMEN DEVOLUCIÓN*\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"🏗️ Obra: {obra_proc}\n"
                    f"{resumen}\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"¿Es correcto? Responde *SÍ* o *NO*"
                )
        else:
            # Intentar parsear "nombre — cantidad unidad" en una sola línea
            import re as _re
            texto_limpio = incoming_msg.strip()
            # Patrón: texto — número [unidad]  o  texto: número [unidad]
            m = _re.match(r'^(.+?)[\s\-–:]+(\d+[\.,]?\d*)\s*([a-zA-Záéíóúüñ²³/]+)?$', texto_limpio)
            if m:
                nombre_mat = m.group(1).strip()
                cantidad = float(m.group(2).replace(',','.'))
                unidad = m.group(3).strip() if m.group(3) else 'ud'
                datos_d = get_estado(numero)['datos']
                lineas = datos_d.get('stock_lineas', [])
                lineas.append({'material': nombre_mat, 'cantidad': cantidad, 'unidad': unidad, 'material_id': None})
                set_dato(numero, 'stock_lineas', lineas)
                msg.body(
                    f"✅ *{nombre_mat}* — {fmt_cant(cantidad)} {unidad}\n\n"
                    f"¿Otro material? Escribe nombre o di *listo*."
                )
            else:
                # Solo el nombre — pedir cantidad
                set_dato(numero, 'stock_mat_tmp', {'nombre': texto_limpio, 'unidad': 'ud'})
                set_paso(numero, 'stock_devol_cantidad')
                msg.body(f"📥 *{texto_limpio}*\n¿Qué *cantidad* devuelves? _(número y unidad, ej: 12 m)_")

    elif paso == 'stock_devol_cantidad':
        import re as _re
        texto_cant = incoming_msg.strip()
        m = _re.match(r'^(\d+[\.,]?\d*)\s*([a-zA-Záéíóúüñ²³/]+)?$', texto_cant)
        if m:
            cantidad = float(m.group(1).replace(',','.'))
            unidad = m.group(2).strip() if m.group(2) else 'ud'
            datos_d = get_estado(numero)['datos']
            mat_tmp = datos_d.get('stock_mat_tmp', {})
            lineas = datos_d.get('stock_lineas', [])
            lineas.append({'material': mat_tmp.get('nombre',''), 'cantidad': cantidad, 'unidad': unidad, 'material_id': None})
            set_dato(numero, 'stock_lineas', lineas)
            set_paso(numero, 'stock_devol_material')
            msg.body(
                f"✅ *{mat_tmp.get('nombre','')}* — {fmt_cant(cantidad)} {unidad}\n\n"
                f"¿Otro material? Escribe nombre o di *listo*."
            )
        else:
            msg.body("Escribe la cantidad. Ej: *12 m* o *5 ud*")

    elif paso == 'stock_devol_confirmar':
        if es_confirmacion(incoming_msg):
            datos_d = get_estado(numero)['datos']
            lineas = datos_d.get('stock_lineas', [])
            nombre_op = datos_d.get('nombre_operario', nombre_operario(numero))
            obra_proc = datos_d.get('stock_devol_obra_txt', 'Sin especificar')
            borrar_estado(numero)
            import threading as _th
            def _procesar_devolucion():
                numero_alb = siguiente_numero_albaran()
                from datetime import datetime as dt
                fecha_str = dt.now().strftime('%d/%m/%Y %H:%M')
                obra_label = f"DEVOLUCIÓN — {obra_proc}"
                aid = crear_albaran(numero_alb, numero, nombre_op, obra_label, lineas)
                for l in lineas:
                    mat_id = l.get('material_id')
                    # Si no tiene material_id, buscar en BD por nombre exacto/aproximado
                    if not mat_id:
                        r = get_material_by_nombre(l['material'])
                        if isinstance(r, tuple):
                            mat_id = r[0]
                        elif isinstance(r, list) and len(r) == 1:
                            mat_id = r[0][0]
                    # Si sigue sin encontrarse, crear el artículo en BD
                    if not mat_id:
                        try:
                            conn_c = get_db(); cur_c = conn_c.cursor()
                            cur_c.execute(
                                "INSERT INTO stock_materiales (nombre, unidad, stock_actual, stock_minimo) "
                                "VALUES (%s, %s, 0, 0) RETURNING id",
                                (l['material'], l['unidad'])
                            )
                            mat_id = cur_c.fetchone()[0]
                            conn_c.commit(); cur_c.close(); conn_c.close()
                            print(f"Artículo creado en BD: {l['material']} (id={mat_id})")
                        except Exception as e_ins:
                            print(f"Error creando artículo {l['material']}: {e_ins}")
                    # Ajustar stock y registrar movimiento
                    if mat_id:
                        ajustar_stock(mat_id, l['cantidad'])
                        registrar_movimiento('devolucion', mat_id, l['material'], l['cantidad'],
                            l['unidad'], numero, nombre_op, obra_label, aid)
                pdf_bytes_dev = generar_pdf_albaran({
                    'numero': numero_alb, 'nombre_operario': nombre_op,
                    'obra': obra_label, 'lineas': lineas, 'fecha': fecha_str, 'tipo': 'devolucion'
                })
                pdf_url = subir_pdf_albaran(pdf_bytes_dev, numero_alb)
                op_wa = numero if numero.startswith('whatsapp:') else f'whatsapp:+{numero.lstrip("+")}'
                resumen_txt = '\n'.join([f"• {l['material']} — {fmt_cant(l['cantidad'])} {l['unidad']}" for l in lineas])
                texto = (
                    f"✅ *Albarán devolución {numero_alb}*\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"📥 Material devuelto al almacén\n"
                    f"🏗️ Obra: {obra_proc}\n"
                    f"{resumen_txt}\n"
                    f"━━━━━━━━━━━━━━━━━━━━"
                )
                enviar_whatsapp(op_wa, texto, media_url=pdf_url if pdf_url else None)
                enviar_supervisor(f"📥 *Devolución almacén — {numero_alb}*\n👷 {nombre_op}\n🏗️ {obra_proc}\n{resumen_txt}",
                    media_url=pdf_url if pdf_url else None)
                # Email con PDF adjunto
                try:
                    nombre_pdf_dev = f"Albaran_Devolucion_{numero_alb.replace('/','_').replace('-','_')}.pdf"
                    msg_email_dev = MIMEMultipart()
                    msg_email_dev['From']    = GMAIL_USER
                    msg_email_dev['To']      = ', '.join([SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2])
                    msg_email_dev['Subject'] = f"[DEVOLUCIÓN ALMACÉN] {numero_alb} — {nombre_op} — {obra_proc}"
                    body_dev = (
                        f"Devolución a almacén registrada.\n\n"
                        f"Albarán: {numero_alb}\n"
                        f"Operario: {nombre_op}\n"
                        f"Obra procedencia: {obra_proc}\n"
                        f"Fecha: {fecha_str}\n\n"
                        f"Materiales devueltos:\n" +
                        '\n'.join([f"  - {l['material']}: {fmt_cant(l['cantidad'])} {l['unidad']}" for l in lineas])
                    )
                    msg_email_dev.attach(MIMEText(body_dev, 'plain'))
                    part_dev = MIMEApplication(pdf_bytes_dev, Name=nombre_pdf_dev)
                    part_dev.add_header('Content-Disposition', f'attachment; filename="{nombre_pdf_dev}"')
                    msg_email_dev.attach(part_dev)
                    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as srv:
                        srv.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                        srv.sendmail(GMAIL_USER, [SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2], msg_email_dev.as_string())
                    print(f"Email devolución OK — {numero_alb}")
                except Exception as e_mail:
                    print(f"Error email devolución: {e_mail}")
            _th.Thread(target=_procesar_devolucion, daemon=True).start()
            msg.body("✅ Devolución registrada. Stock actualizado. Te envío el albarán en un momento.")
        elif es_cancelacion(incoming_msg):
            borrar_estado(numero)
            msg.body("❌ Cancelado.")
        else:
            msg.body("Responde *SÍ* para confirmar o *NO* para cancelar.")

    # ── Flujo Almacén: Consulta ───────────────────────────────────────────────
    elif paso == 'stock_consulta':
        mat, err = buscar_material_msg(incoming_msg)
        if err:
            # Caso retales: mostrar lista formateada
            if isinstance(err, tuple) and err[0] == 'RETALES':
                candidatos = err[1]
                borrar_estado(numero)
                lineas = []
                for c in candidatos:
                    precio = float(c[5]) if len(c) > 5 and c[5] else 0
                    precio_txt = f" — {precio:.2f} €/ud".replace(".",",") if precio > 0 else ""
                    def _fmt_r(v):
                        f2 = float(v)
                        return (f"{f2:.3f}".rstrip('0').rstrip('.')).replace('.', ',')
                    lineas.append(f"• *{c[1]}*: {_fmt_r(c[3])} {c[2]}{precio_txt}")
                msg.body("🔍 *Resultados encontrados:*\n\n" + "\n".join(lineas) + "\n\nEscribe el nombre más completo para más detalle.")
            # Varios resultados normales (ya formateados como string)
            elif isinstance(err, str) and err.startswith("🔍 Encontré varios"):
                borrar_estado(numero)
                msg.body(err + "\n\nEscribe *Consulta* para intentarlo de nuevo.")
            else:
                borrar_estado(numero)
                msg.body(f"{err}\n\nEscribe *Consulta* para intentarlo de nuevo.")
        else:
            borrar_estado(numero)
            stock = mat[3]; minimo = mat[4]; unidad = mat[2]; nombre_mat = mat[1]
            precio = float(mat[5]) if len(mat) > 5 and mat[5] else 0
            alerta = "\n⚠️ *Stock por debajo del mínimo*" if stock <= minimo and minimo > 0 else ""
            precio_txt = f"\nPrecio unitario: *{precio:.2f} €*".replace(".",",") if precio > 0 else ""
            def _fmt(v):
                f = float(v)
                return (f"{f:.3f}".rstrip('0').rstrip('.')).replace('.', ',')
            msg.body(f"🔍 *{nombre_mat}*\nStock actual: *{_fmt(stock)} {unidad}*\nStock mínimo: {_fmt(minimo)} {unidad}{precio_txt}{alerta}")

    # ── Flujo vacaciones ──────────────────────────────────────────────────────
    elif paso == 'vac_nombre':
        set_dato(numero, 'nombre_operario', incoming_msg)
        set_paso(numero, 'vac_inicio')
        msg.body(
            f"👋 Hola *{incoming_msg}*\n\n"
            f"1️⃣ ¿Cuál es la *fecha de inicio* de tus vacaciones?\n_Ejemplo: 14/07/2026_"
        )

    elif paso == 'vac_inicio':
        set_dato(numero, 'fecha_inicio', incoming_msg)
        set_paso(numero, 'vac_fin')
        msg.body("2️⃣ ¿Cuál es la *fecha de fin*?\n_Ejemplo: 21/07/2026_")

    elif paso == 'vac_fin':
        set_dato(numero, 'fecha_fin', incoming_msg)
        set_paso(numero, 'vac_confirmar')
        datos_vac = get_estado(numero)['datos']
        fi = datos_vac.get('fecha_inicio','')
        ff = incoming_msg
        dias = calcular_dias_laborables(fi, ff)
        num_limpio = numero.replace('whatsapp:','').replace('+','').strip()
        saldo = get_saldo_vacaciones(num_limpio)
        saldo_txt = f"\n📊 Días disponibles tras solicitud: *{saldo[0] - saldo[1] - dias}* de {saldo[0]}" if saldo else ""
        op_nombre = datos_vac.get('nombre_operario', nombre_operario(numero))
        msg.body(
            f"🌴 *RESUMEN SOLICITUD DE VACACIONES*\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"👷 {op_nombre}\n"
            f"📅 Inicio: {fi}\n"
            f"📅 Fin: {ff}\n"
            f"📆 Días laborables: *{dias}*{saldo_txt}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"¿Confirmas la solicitud? Responde *SÍ* o *NO*"
        )

    elif paso == 'vac_confirmar':
        if es_confirmacion(incoming_msg):
            datos_vac = get_estado(numero)['datos']
            vac_id, dias = guardar_vacacion(datos_vac, numero)
            borrar_estado(numero)
            fi = datos_vac.get('fecha_inicio','')
            ff = datos_vac.get('fecha_fin','')
            op_nombre = datos_vac.get('nombre_operario', nombre_operario(numero))
            # Notificar a Alberto
            if vac_id:
                enviar_supervisor(
                    f"🌴 *SOLICITUD DE VACACIONES #{vac_id}*\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"👷 {op_nombre}\n"
                    f"📅 Del {fi} al {ff}\n"
                    f"📆 {dias} días laborables\n"
                    f"━━━━━━━━━━━━━━━━━━━━\n"
                    f"✅ *APROBAR {vac_id}*\n"
                    f"❌ *RECHAZAR {vac_id}*"
                )
                msg.body(f"✅ Solicitud #{vac_id} enviada. Pendiente de aprobación por Alberto.")
            else:
                msg.body("✅ Solicitud enviada. Pendiente de aprobación por Alberto.")
        elif es_cancelacion(incoming_msg):
            borrar_estado(numero)
            msg.body("❌ Solicitud cancelada. Escribe *vacaciones* para empezar de nuevo.")
        else:
            msg.body("Responde *SÍ* para confirmar o *NO* para cancelar.")

    if use_meta:
        return 'OK', 200
    return str(resp)

CSS_BASE = """
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f0f2f5; color: #333; }
  header { background: #1a3a5c; color: white; padding: 20px 30px; display: flex; align-items: center; justify-content: space-between; }
  header h1 { font-size: 22px; font-weight: 700; }
  header p { font-size: 13px; opacity: .75; margin-top: 2px; }
  .stats { display: flex; gap: 16px; padding: 20px 30px; flex-wrap: wrap; }
  .stat { background: white; border-radius: 10px; padding: 16px 24px; box-shadow: 0 1px 4px rgba(0,0,0,.08); }
  .stat .num { font-size: 28px; font-weight: 700; color: #1a3a5c; }
  .stat .lbl { font-size: 12px; color: #888; margin-top: 2px; }
  .wrap { padding: 0 30px 30px; overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,.08); font-size: 13px; }
  th { background: #1a3a5c; color: white; padding: 12px 10px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: .5px; white-space: nowrap; }
  td { padding: 10px; border-bottom: 1px solid #f0f0f0; vertical-align: top; }
  tr:last-child td { border-bottom: none; }
  tr.clickable:hover td { background: #eef3fa; cursor: pointer; }
  .badge-ok { background:#2e7d32; color:white; padding:3px 10px; border-radius:10px; font-size:11px; white-space:nowrap; }
  .badge-curso { background:#e65100; color:white; padding:3px 10px; border-radius:10px; font-size:11px; white-space:nowrap; }
  .empty { text-align: center; padding: 60px; color: #aaa; font-size: 15px; }
  .back { display:inline-block; margin:20px 30px 0; color:#1a3a5c; text-decoration:none; font-weight:600; font-size:14px; }
  .back:hover { text-decoration:underline; }
  .ficha { max-width: 800px; margin: 24px auto; background: white; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,.1); overflow: hidden; }
  .ficha-header { background: #1a3a5c; color: white; padding: 20px 28px; }
  .ficha-header h2 { font-size: 20px; }
  .ficha-header p { font-size: 13px; opacity: .75; margin-top: 4px; }
  .ficha-body { padding: 28px; }
  .campo { margin-bottom: 18px; }
  .campo label { display:block; font-size:11px; text-transform:uppercase; letter-spacing:.5px; color:#888; margin-bottom:4px; }
  .campo .val { font-size:15px; color:#222; white-space:pre-line; }
  .grid2 { display:grid; grid-template-columns:1fr 1fr; gap:16px; }
  .estado-ok { background:#e8f5e9; border:1px solid #a5d6a7; border-radius:8px; padding:14px 20px; text-align:center; font-weight:700; color:#2e7d32; font-size:16px; }
  .estado-curso { background:#fff3e0; border:1px solid #ffcc80; border-radius:8px; padding:14px 20px; text-align:center; font-weight:700; color:#e65100; font-size:16px; }
  .btn-pdf { display:inline-block; margin-top:20px; background:#1a3a5c; color:white; padding:10px 22px; border-radius:8px; text-decoration:none; font-size:14px; font-weight:600; }
  .btn-pdf:hover { background:#14304f; }
  @media (max-width:600px) { .grid2 { grid-template-columns:1fr; } .wrap { padding:0 12px 20px; } header h1 { font-size:17px; } }
"""

def get_parte_by_id(parte_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, numero_parte, fecha, operario, cliente, obra, operarios, albaranes, material_stock, devolucion_almacen, descripcion, terminado, tiempo_restante, created_at FROM partes WHERE id=%s", (parte_id,))
        r = cur.fetchone()
        cur.close(); conn.close()
        return r
    except:
        return None

@app.route('/partes', methods=['GET'])
def listar_partes():
    if request.args.get('format') == 'json':
        try:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("SELECT id, numero_parte, fecha, operario, cliente, obra, operarios, albaranes, material_stock, devolucion_almacen, descripcion, terminado, tiempo_restante, created_at FROM partes ORDER BY TO_DATE(fecha, 'DD/MM/YYYY') DESC, created_at DESC LIMIT 200")
            rows = cur.fetchall()
            cur.close(); conn.close()
            partes = [{'id':r[0],'numero_parte':r[1],'fecha':r[2],'operario':r[3],'cliente':r[4],'obra':r[5],'operarios':r[6],'albaranes':r[7],'material_stock':r[8],'devolucion_almacen':r[9],'descripcion':r[10],'terminado':r[11],'tiempo_restante':r[12],'created_at':str(r[13])} for r in rows]
            return {'partes': partes, 'total': len(partes)}, 200
        except Exception as e:
            return {'error': str(e)}, 500

    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT id, numero_parte, fecha, operario, cliente, obra, terminado, tiempo_restante, created_at, pdf_descargado, pdf_descargado_at FROM partes ORDER BY TO_DATE(fecha, 'DD/MM/YYYY') DESC, created_at DESC LIMIT 200")
        rows = cur.fetchall()
        cur.close(); conn.close()
    except:
        rows = []

    filas = ""
    for r in rows:
        terminado = r[6] or ''
        es_ok = 'í' in terminado.lower() or terminado.lower() == 'si'
        badge = '<span class="badge-ok">✓ Terminado</span>' if es_ok else f'<span class="badge-curso">🔄 {r[7] or "En curso"}</span>'
        operario_limpio = nombre_operario(r[3] or '')
        descargado = r[9]
        desc_at = str(r[10])[:16].replace('T',' ') if r[10] else ''
        if descargado:
            pdf_badge = f'<span title="Descargado {desc_at}" style="color:#2e7d32;font-size:18px" title="{desc_at}">⬇️</span>'
        else:
            pdf_badge = '<span style="color:#ccc;font-size:18px">—</span>'
        filas += f'<tr class="clickable" onclick="window.location=\'/partes/{r[0]}\'">' \
                 f'<td>{r[2] or ""}</td>' \
                 f'<td style="font-size:11px;color:#666">{operario_limpio}</td>' \
                 f'<td><strong>{r[4] or ""}</strong></td>' \
                 f'<td>{r[5] or ""}</td>' \
                 f'<td>{badge}</td>' \
                 f'<td style="text-align:center">{pdf_badge}</td>' \
                 f'</tr>'

    total = len(rows)
    n_ok = sum(1 for r in rows if r[6] and ('í' in r[6].lower() or r[6].lower()=='si'))
    n_curso = sum(1 for r in rows if r[6] and 'no' in r[6].lower())
    n_pdf = sum(1 for r in rows if r[9])

    tabla = "<p class='empty'>No hay partes registrados aún.</p>" if not rows else f"""
  <table>
    <thead><tr>
      <th>Fecha</th><th>Operario</th><th>Cliente</th><th>Obra</th><th>Estado</th><th>PDF</th>
    </tr></thead>
    <tbody>{filas}</tbody>
  </table>"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Partes de Trabajo — Instapalma</title>
<style>{CSS_BASE}</style>
</head>
<body>
<header>
  <div><h1>⚡ Partes de Trabajo — Instapalma</h1><p>Panel de control · Haz clic en un parte para ver el detalle</p></div>
</header>
<div class="stats">
  <div class="stat"><div class="num">{total}</div><div class="lbl">Total partes</div></div>
  <div class="stat"><div class="num">{n_ok}</div><div class="lbl">Terminados</div></div>
  <div class="stat"><div class="num">{n_curso}</div><div class="lbl">En curso</div></div>
  <div class="stat"><div class="num">{n_pdf}</div><div class="lbl">PDFs descargados</div></div>
</div>
<div class="wrap">{tabla}</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/partes/<int:parte_id>', methods=['GET'])
def ver_parte(parte_id):
    r = get_parte_by_id(parte_id)
    if not r:
        return "<p style='padding:40px;font-family:sans-serif'>Parte no encontrado.</p>", 404

    terminado = r[11] or ''
    es_ok = 'í' in terminado.lower() or terminado.lower() == 'si'
    estado_html = f'<div class="estado-ok">✅ TRABAJO TERMINADO</div>' if es_ok \
        else f'<div class="estado-curso">🔄 EN CURSO — Tiempo restante: {r[12] or "no especificado"}</div>'
    operario_limpio = nombre_operario(r[3] or '')

    html = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Parte {r[1]} — Instapalma</title>
<style>{CSS_BASE}</style>
</head>
<body>
<header>
  <div><h1>⚡ Instapalma — Ficha de Parte</h1><p>Detalle completo del parte de trabajo</p></div>
</header>
<a class="back" href="/partes">← Volver al listado</a>
<div class="ficha">
  <div class="ficha-header">
    <h2>Parte — {r[2] or ''}</h2>
    <p>Fecha: {r[2] or '—'} &nbsp;·&nbsp; Registrado: {str(r[13])[:16] if r[13] else '—'}</p>
  </div>
  <div class="ficha-body">
    <div class="grid2">
      <div class="campo"><label>Cliente</label><div class="val">{r[4] or '—'}</div></div>
      <div class="campo"><label>Obra</label><div class="val">{r[5] or '—'}</div></div>
      <div class="campo"><label>Operario</label><div class="val">{operario_limpio}</div></div>
    </div>
    <div class="campo"><label>Operarios y horas</label><div class="val">{r[6] or '—'}</div></div>
    <div class="campo"><label>Albaranes</label><div class="val">{r[7] or '—'}</div></div>
    <div class="campo"><label>Material de stock</label><div class="val">{r[8] or '—'}</div></div>
    <div class="campo"><label>Devolución a almacén</label><div class="val">{r[9] or '—'}</div></div>
    <div class="campo"><label>Descripción de trabajos</label><div class="val">{r[10] or '—'}</div></div>
    {estado_html}
    <a class="btn-pdf" href="/partes/{parte_id}/pdf">⬇ Descargar PDF</a>
  </div>
</div>
</body></html>"""
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/partes/<int:parte_id>/pdf', methods=['GET'])
def descargar_pdf(parte_id):
    r = get_parte_by_id(parte_id)
    if not r:
        return "Parte no encontrado", 404
    datos = {
        'numero_parte': r[1], 'fecha': r[2], 'cliente': r[4], 'obra': r[5],
        'operarios': r[6] or '', 'albaranes': r[7] or '', 'material_stock': r[8] or '',
        'devolucion_almacen': r[9] or 'Ninguno',
        'descripcion': r[10] or '', 'terminado': r[11] or '', 'tiempo_restante': r[12] or ''
    }
    pdf_bytes = generar_pdf(datos)
    fecha_pdf2 = (r[2] or '').replace('/', '-').replace(' ', '')
    obra_pdf2 = limpiar_nombre_archivo(r[5] or 'obra')
    ops_raw2 = r[6] or ''
    ops_lista2 = [limpiar_nombre_archivo(l.split('—')[0].split('-')[0].strip().split()[0]) for l in ops_raw2.split('\n') if l.strip()]
    ops_pdf2 = '-'.join(ops_lista2) if ops_lista2 else 'OPERARIOS'
    nombre = f"{fecha_pdf2}-{obra_pdf2}-{ops_pdf2}.pdf"
    # Marcar como descargado
    try:
        from datetime import datetime as dt
        conn2 = get_db(); cur2 = conn2.cursor()
        cur2.execute("UPDATE partes SET pdf_descargado=TRUE, pdf_descargado_at=%s WHERE id=%s",
                    (dt.utcnow(), parte_id))
        conn2.commit(); cur2.close(); conn2.close()
    except Exception:
        pass
    from flask import Response
    return Response(pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nombre}"'})

@app.route('/albaran/<numero_pdf>', methods=['GET'])
def servir_albaran_pdf(numero_pdf):
    """Sirve el PDF de un albarán almacenado en BD."""
    from flask import Response
    try:
        # Quitar extensión .pdf si viene en la URL
        clave = numero_pdf
        if clave.endswith('.pdf'):
            clave = clave[:-4]
        conn = get_db()
        cur = conn.cursor()
        # Intentar búsqueda exacta primero, luego con _ reemplazado por /
        cur.execute("""
            SELECT pdf_bytes, numero FROM stock_albaranes
            WHERE numero = %s
               OR numero = REPLACE(%s, '_', '/')
               OR REPLACE(numero, '/', '_') = %s
            LIMIT 1
        """, (clave, clave, clave))
        row = cur.fetchone()
        cur.close(); conn.close()
        if not row or not row[0]:
            return "Albarán no encontrado", 404
        pdf_data = bytes(row[0])
        nombre = f"Albaran_{row[1].replace('/','_')}.pdf"
        return Response(pdf_data, mimetype='application/pdf',
            headers={'Content-Disposition': f'inline; filename="{nombre}"',
                     'Content-Type': 'application/pdf'})
    except Exception as e:
        print(f"Error servir albaran PDF: {e}")
        return "Error interno", 500


    try:
        datos = request.get_json()
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO partes (numero_parte, fecha, operario, cliente, obra, operarios, albaranes, material_stock, devolucion_almacen, descripcion, terminado, tiempo_restante)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            datos.get('numero_parte'), datos.get('fecha'), datos.get('operario'),
            datos.get('cliente'), datos.get('obra'), datos.get('operarios'),
            datos.get('albaranes'), datos.get('material_stock'), datos.get('devolucion_almacen','Ninguno'),
            datos.get('descripcion'), datos.get('terminado'), datos.get('tiempo_restante')
        ))
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/reset-conversacion', methods=['POST'])
def admin_reset_conv():
    try:
        datos = request.get_json()
        numero = datos.get('numero', '')
        conn = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM conversaciones_db WHERE numero=%s", (numero,))
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'msg': f'Conversación {numero} borrada'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/reset-herramienta', methods=['POST'])
def admin_reset_herramienta():
    """Vacía herramienta_obra y herramienta para recarga limpia."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("DELETE FROM herramienta_obra")
        cur.execute("DELETE FROM herramienta")
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'msg': 'Tablas herramienta vaciadas'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/fix-herramienta-nombre', methods=['POST'])
def admin_fix_herramienta_nombre():
    """Rellena herramienta_nombre en herramienta_obra donde sea NULL, usando la tabla herramienta."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            UPDATE herramienta_obra ho
            SET herramienta_nombre = h.nombre
            FROM herramienta h
            WHERE ho.herramienta_id = h.id
              AND (ho.herramienta_nombre IS NULL OR ho.herramienta_nombre = '')
        """)
        rows = cur.rowcount
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'reparados': rows}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/carga-almacen-herramienta', methods=['POST'])
def admin_carga_almacen_herramienta():
    """Carga stock de almacén. Body: lista de {nombre, stock, observaciones}
    Si la herramienta ya existe (viene de herramienta_obra), actualiza solo el stock.
    Si no existe, la crea."""
    try:
        items = request.get_json()
        conn = get_db(); cur = conn.cursor()
        actualizadas = 0
        for item in items:
            nombre = item['nombre']
            stock  = int(item.get('stock', 0))
            obs    = item.get('observaciones', '')
            cur.execute("""
                INSERT INTO herramienta (nombre, tipo, stock_almacen, observaciones)
                VALUES (%s, 'almacen', %s, %s)
                ON CONFLICT (nombre) DO UPDATE
                  SET stock_almacen = EXCLUDED.stock_almacen,
                      observaciones = EXCLUDED.observaciones
            """, (nombre, stock, obs))
            actualizadas += 1
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'actualizadas': actualizadas}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/carga-personal-herramienta', methods=['POST'])
def admin_carga_personal_herramienta():
    """Carga herramienta personal. Body: lista de {propietario, articulo, tipo}
    tipo: 'herramienta' | 'epi'
    """
    try:
        items = request.get_json()
        conn = get_db(); cur = conn.cursor()
        insertadas = 0
        for item in items:
            propietario = item['propietario']
            articulo    = item['articulo']
            tipo        = item.get('tipo', 'herramienta').lower()
            cur.execute("""
                INSERT INTO herramienta_personal (propietario, articulo, tipo, fecha_alta)
                VALUES (%s, %s, %s, NOW())
            """, (propietario, articulo, tipo))
            insertadas += 1
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'insertadas': insertadas}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/carga-herramienta', methods=['POST'])
def admin_carga_herramienta():
    """Carga masiva de herramienta. Body: lista de {nombre, obra, responsable, fecha_alta, cantidad}"""
    try:
        items = request.get_json()
        conn = get_db(); cur = conn.cursor()
        insertadas = 0
        for item in items:
            nombre   = item['nombre']
            obra     = item.get('obra', '')
            resp     = item.get('responsable') or ''
            fecha    = item.get('fecha_alta', '2026-01-01')
            cantidad = int(item.get('cantidad', 1))
            cur.execute("""
                INSERT INTO herramienta (nombre, tipo, stock_almacen)
                VALUES (%s, 'almacen', 0)
                ON CONFLICT (nombre) DO NOTHING
            """, (nombre,))
            cur.execute("SELECT id FROM herramienta WHERE nombre=%s", (nombre,))
            herr_id = cur.fetchone()[0]
            for _ in range(cantidad):
                cur.execute("""
                    INSERT INTO herramienta_obra (herramienta_id, herramienta_nombre, operario, nombre_operario, obra, fecha_alta)
                    VALUES (%s, %s, %s, %s, %s, %s)
                """, (herr_id, nombre, resp, resp or '—', obra, fecha))
                insertadas += 1
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'insertadas': insertadas}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/truncate-partes', methods=['POST'])
def admin_truncate():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("TRUNCATE TABLE partes RESTART IDENTITY")
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'msg': 'Tabla partes vaciada'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/insert-partes', methods=['POST'])
def admin_insert_partes():
    try:
        datos = request.json  # lista de partes
        conn = get_db(); cur = conn.cursor()
        count = 0
        for p in datos:
            cur.execute(
                """INSERT INTO partes (numero_parte, fecha, operario, cliente, obra, operarios,
                    albaranes, material_stock, devolucion_almacen, descripcion, terminado, tiempo_restante)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (
                    p.get('numero_parte',''),
                    p.get('fecha',''),
                    p.get('operario',''),
                    p.get('cliente',''),
                    p.get('obra',''),
                    p.get('operarios',''),
                    p.get('albaranes',''),
                    p.get('material_stock',''),
                    p.get('devolucion_almacen',''),
                    p.get('descripcion',''),
                    p.get('terminado',''),
                    p.get('tiempo_restante','')
                )
            )
            count += 1
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'insertados': count}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/truncate-albaranes', methods=['POST'])
def admin_truncate_albaranes():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("TRUNCATE TABLE stock_albaranes RESTART IDENTITY")
        cur.execute("TRUNCATE TABLE stock_movimientos RESTART IDENTITY")
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'msg': 'Albaranes y movimientos borrados'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/truncate-resumenes', methods=['POST'])
def admin_truncate_resumenes():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("TRUNCATE TABLE resumen_mes RESTART IDENTITY")
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'msg': 'Resúmenes de fin de mes borrados'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/truncate-vehiculos', methods=['POST'])
def admin_truncate_vehiculos():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("TRUNCATE TABLE vehiculos RESTART IDENTITY CASCADE")
        conn.commit(); cur.close(); conn.close()
        return {'status': 'ok', 'msg': 'Registros de vehículos borrados'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/schema', methods=['GET'])
def admin_schema():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT table_name, column_name, data_type
            FROM information_schema.columns
            WHERE table_name IN ('herramienta','herramienta_obra','herramienta_personal','herramienta_epis','stock_materiales')
            ORDER BY table_name, ordinal_position
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        result = {}
        for t, c, d in rows:
            result.setdefault(t, []).append(f"{c} ({d})")
        return result, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/stock-materiales', methods=['GET'])
def admin_stock_materiales():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, nombre, unidad, stock_actual, stock_minimo, familia FROM stock_materiales ORDER BY familia, nombre")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return {'count': len(rows), 'items': [{'id':r[0],'nombre':r[1],'unidad':r[2],'stock':float(r[3]) if r[3] else 0,'minimo':float(r[4]) if r[4] else 0,'familia':r[5]} for r in rows]}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/herr-debug', methods=['GET'])
def admin_herr_debug():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT nombre, tipo, stock_almacen FROM herramienta ORDER BY nombre")
        stock = [{'nombre': r[0], 'tipo': r[1], 'stock_almacen': r[2]} for r in cur.fetchall()]
        cur.execute("SELECT herramienta_nombre, obra, nombre_operario, activo, fecha_alta FROM herramienta_obra ORDER BY fecha_alta DESC LIMIT 30")
        obra = [{'herramienta': r[0], 'obra': r[1], 'operario': r[2], 'activo': r[3], 'fecha': str(r[4])} for r in cur.fetchall()]
        cur.close(); conn.close()
        return {'stock': stock, 'obra': obra}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/admin/albaranes-lista', methods=['GET'])
def admin_albaranes_lista():
    try:
        conn = get_db(); cur = conn.cursor()
        # Estructura de la tabla
        cur.execute("""
            SELECT column_name, data_type, column_default
            FROM information_schema.columns
            WHERE table_name='stock_albaranes'
            ORDER BY ordinal_position
        """)
        cols = cur.fetchall()
        cur.execute("SELECT numero, length(pdf_bytes), created_at FROM stock_albaranes ORDER BY created_at DESC LIMIT 20")
        rows = cur.fetchall()
        cur.close(); conn.close()
        return {
            'columns': [{'name': c[0], 'type': c[1], 'default': c[2]} for c in cols],
            'count': len(rows),
            'items': [{'numero': r[0], 'bytes': r[1], 'created_at': str(r[2])} for r in rows]
        }, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/migrate', methods=['GET'])
def migrate():
    try:
        conn = get_db()
        cur = conn.cursor()
        # Columnas tabla partes
        for col, tipo in [('material_stock','TEXT'), ('devolucion_almacen','TEXT'), ('terminado','TEXT'), ('tiempo_restante','TEXT'), ('pdf_descargado','BOOLEAN DEFAULT FALSE'), ('pdf_descargado_at','TIMESTAMP')]:
            try:
                cur.execute(f"ALTER TABLE partes ADD COLUMN IF NOT EXISTS {col} {tipo}")
                conn.commit()
            except Exception:
                conn.rollback()
        # Crear tabla vehiculos si no existe
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vehiculos (
                id SERIAL PRIMARY KEY,
                matricula VARCHAR(20),
                marca_modelo VARCHAR(100),
                mes VARCHAR(20),
                km_inicio VARCHAR(20),
                km_fin VARCHAR(20),
                proximo_aceite VARCHAR(20),
                estado_neumaticos TEXT,
                conductores TEXT,
                mantenimientos TEXT,
                observaciones TEXT,
                golpes TEXT,
                operario VARCHAR(100),
                created_at TIMESTAMP DEFAULT NOW(),
                pdf_descargado BOOLEAN DEFAULT FALSE,
                pdf_descargado_at TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS vacaciones (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                fecha_inicio VARCHAR(20),
                fecha_fin VARCHAR(20),
                dias_solicitados INTEGER DEFAULT 0,
                fecha_solicitud VARCHAR(20),
                estado VARCHAR(20) DEFAULT 'pendiente',
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Añadir columnas si la tabla ya existía sin ellas
        for col, tipo in [('nombre_operario','VARCHAR(100)'), ('dias_solicitados','INTEGER DEFAULT 0')]:
            try:
                cur.execute(f"ALTER TABLE vacaciones ADD COLUMN IF NOT EXISTS {col} {tipo}")
                conn.commit()
            except Exception:
                conn.rollback()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS saldo_vacaciones (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100) UNIQUE,
                nombre VARCHAR(100),
                dias_totales INTEGER DEFAULT 23,
                dias_usados INTEGER DEFAULT 0,
                anio INTEGER DEFAULT 2026,
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS resumen_mes (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                mes VARCHAR(30),
                horas_extra VARCHAR(20),
                dias_vacaciones VARCHAR(20),
                total_gastos VARCHAR(30),
                foto_url TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit()
        cur.close()
        conn.close()
        return {'status': 'migración OK, tablas vehiculos, vacaciones, saldo_vacaciones y resumen_mes creadas'}, 200
    except Exception as e:
        return {'error': str(e)}, 500

@app.route('/')
def dashboard():
    try:
        conn = get_db(); cur = conn.cursor()
        # Partes
        cur.execute("SELECT COUNT(*) FROM stock_albaranes WHERE numero LIKE 'ALB-%'")
        total_partes = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM stock_albaranes WHERE numero LIKE 'ALB-%' AND pdf_bytes IS NOT NULL")
        partes_terminados = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM stock_albaranes WHERE numero LIKE 'ALB-%' AND DATE(created_at) = CURRENT_DATE")
        partes_hoy = cur.fetchone()[0]
        cur.execute("SELECT numero, nombre_operario, obra, pdf_bytes IS NOT NULL, created_at FROM stock_albaranes WHERE numero LIKE 'ALB-%' ORDER BY created_at DESC LIMIT 8")
        ultimos_partes = cur.fetchall()
        # Stock almacén
        cur.execute("SELECT COUNT(*) FROM stock_materiales")
        total_mat = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM stock_materiales WHERE stock_actual <= stock_minimo AND stock_minimo > 0")
        mat_bajos = cur.fetchone()[0]
        cur.execute("SELECT nombre, stock_actual, unidad, stock_minimo FROM stock_materiales WHERE stock_actual <= stock_minimo AND stock_minimo > 0 ORDER BY nombre LIMIT 6")
        mat_alerta = cur.fetchall()
        # Herramienta
        cur.execute("SELECT COUNT(*) FROM herramienta_obra WHERE activo=TRUE")
        herr_obra = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM herramienta WHERE tipo='almacen' AND stock_almacen > 0")
        herr_almacen = cur.fetchone()[0]
        cur.execute("SELECT herramienta_nombre, nombre_operario, obra, fecha_alta FROM herramienta_obra WHERE activo=TRUE ORDER BY fecha_alta DESC LIMIT 8")
        herr_en_obra = cur.fetchall()
        # Vacaciones
        cur.execute("SELECT COUNT(*) FROM vacaciones WHERE estado='pendiente'")
        vac_pendientes = (cur.fetchone() or [0])[0]
        # Movimientos almacén
        cur.execute("SELECT tipo, material_nombre, cantidad, unidad, nombre_operario, obra, created_at FROM stock_movimientos ORDER BY created_at DESC LIMIT 6")
        movimientos = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        return f"<h1>Error: {e}</h1>", 500

    filas_partes = ''
    for p in ultimos_partes:
        num, op, obra, term, cat = p
        badge = '<span style="background:#2e7d32;color:white;padding:2px 8px;border-radius:8px;font-size:11px">✓ PDF generado</span>' if term else '<span style="background:#e65100;color:white;padding:2px 8px;border-radius:8px;font-size:11px">🔄 Sin PDF</span>'
        fecha = cat.strftime('%d/%m %H:%M') if cat else '—'
        pdf_link = f'<a href="/albaran/{num}.pdf" target="_blank" style="color:#1a3a5c;font-weight:700;font-size:16px">📄</a>'
        filas_partes += f'<tr><td><b>{num}</b></td><td style="font-size:12px;color:#666">{op or "—"}</td><td>{obra or "—"}</td><td>{badge}</td><td style="font-size:12px;color:#888">{fecha}</td><td style="text-align:center">{pdf_link}</td></tr>'

    filas_herr = ''
    for h in herr_en_obra:
        nombre, op, obra, falta = h
        fecha = falta.strftime('%d/%m/%Y') if falta else '—'
        filas_herr += f'<tr><td><b>{nombre}</b></td><td style="font-size:12px;color:#666">{op or "—"}</td><td>{obra or "—"}</td><td style="font-size:12px;color:#888">{fecha}</td></tr>'

    alertas_stock = ''
    for m in mat_alerta:
        nombre, stock, unidad, minimo = m
        alertas_stock += f'<div style="display:flex;justify-content:space-between;align-items:center;padding:9px 0;border-bottom:1px solid #f0f0f0"><span style="font-size:13px">{nombre}</span><span style="background:#fff3e0;border:1px solid #ffcc80;color:#e65100;padding:2px 10px;border-radius:8px;font-size:12px;white-space:nowrap">{str(stock).replace(".",",")} / {str(minimo).replace(".",",")} {unidad}</span></div>'
    if not alertas_stock:
        alertas_stock = '<p style="color:#2e7d32;text-align:center;padding:20px;font-size:13px">✅ Todo el stock sobre mínimos</p>'

    filas_mov = ''
    for mv in movimientos:
        tipo, mat, cant, unidad, op, obra, cat = mv
        color = '#e8f5e9' if tipo == 'entrada' else '#fff3e0'
        icono = '📥' if tipo == 'entrada' else '📤'
        fecha = cat.strftime('%d/%m %H:%M') if cat else '—'
        filas_mov += f'<tr style="background:{color}"><td>{icono} {tipo.capitalize()}</td><td><b>{mat}</b></td><td style="text-align:center">{str(cant).replace(".",",")} {unidad}</td><td style="font-size:12px;color:#666">{op or "—"}</td><td style="font-size:12px;color:#666">{obra or "—"}</td><td style="font-size:12px;color:#888">{fecha}</td></tr>'

    warn_stock = 'warn' if mat_bajos > 0 else 'ok'
    warn_vac = 'warn' if vac_pendientes > 0 else ''

    return f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Dashboard — Instapalma</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f0f2f5;color:#333}}
  header{{background:#1a3a5c;color:white;padding:18px 28px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}}
  header h1{{font-size:20px;font-weight:700}} header p{{font-size:12px;opacity:.7;margin-top:2px}}
  .nav{{display:flex;gap:8px;flex-wrap:wrap}}
  .nav a{{background:rgba(255,255,255,.15);color:white;padding:7px 14px;border-radius:20px;text-decoration:none;font-size:12px;font-weight:600;transition:.2s}}
  .nav a:hover{{background:rgba(255,255,255,.3)}}
  .kpis{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:14px;padding:20px 28px}}
  .kpi{{background:white;border-radius:12px;padding:16px 20px;box-shadow:0 1px 4px rgba(0,0,0,.08);border-top:3px solid #1a3a5c}}
  .kpi.warn{{border-top-color:#e65100}} .kpi.ok{{border-top-color:#2e7d32}}
  .kpi .num{{font-size:30px;font-weight:800;color:#1a3a5c}}
  .kpi.warn .num{{color:#e65100}} .kpi.ok .num{{color:#2e7d32}}
  .kpi .lbl{{font-size:11px;color:#888;margin-top:3px;text-transform:uppercase;letter-spacing:.4px}}
  .grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px;padding:0 28px 28px}}
  .card{{background:white;border-radius:12px;box-shadow:0 1px 4px rgba(0,0,0,.08);overflow:hidden}}
  .card-full{{grid-column:1/-1}}
  .card-head{{background:#1a3a5c;color:white;padding:12px 18px;display:flex;align-items:center;justify-content:space-between}}
  .card-head h2{{font-size:14px;font-weight:700}}
  .card-head a{{color:rgba(255,255,255,.8);font-size:12px;text-decoration:none}} .card-head a:hover{{color:white}}
  table{{width:100%;border-collapse:collapse;font-size:13px}}
  th{{background:#f7f8fa;color:#888;padding:9px 12px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.4px;border-bottom:1px solid #eee}}
  td{{padding:9px 12px;border-bottom:1px solid #f5f5f5;vertical-align:middle}}
  tr:last-child td{{border-bottom:none}} tr:hover td{{background:#fafbff}}
  .alertas{{padding:12px 18px}}
  @media(max-width:700px){{.grid{{grid-template-columns:1fr}}.card-full{{grid-column:1}}.kpis{{padding:14px}}.grid{{padding:0 14px 14px}}header{{padding:14px}}}}
</style>
</head>
<body>
<header>
  <div><h1>⚡ Instapalma — Dashboard</h1><p>Panel de control general</p></div>
  <nav class="nav">
    <a href="/partes">📋 Partes</a>
    <a href="/almacen">📦 Almacén</a>
    <a href="/herramienta">🔧 Herramienta</a>
    <a href="/vacaciones">🏖️ Vacaciones</a>
    <a href="/vehiculos">🚐 Vehículos</a>
    <a href="/resumenes">📊 Resúmenes</a>
  </nav>
</header>

<div class="kpis">
  <div class="kpi"><div class="num">{partes_hoy}</div><div class="lbl">Partes hoy</div></div>
  <div class="kpi"><div class="num">{total_partes}</div><div class="lbl">Total partes</div></div>
  <div class="kpi ok"><div class="num">{partes_terminados}</div><div class="lbl">Terminados</div></div>
  <div class="kpi"><div class="num">{total_mat}</div><div class="lbl">Artículos almacén</div></div>
  <div class="kpi {warn_stock}"><div class="num">{mat_bajos}</div><div class="lbl">Stock bajo mínimo</div></div>
  <div class="kpi"><div class="num">{herr_obra}</div><div class="lbl">Herramienta en obra</div></div>
  <div class="kpi ok"><div class="num">{herr_almacen}</div><div class="lbl">Herramienta almacén</div></div>
  <div class="kpi {warn_vac}"><div class="num">{vac_pendientes}</div><div class="lbl">Vacaciones pendientes</div></div>
</div>

<div class="grid">

  <div class="card">
    <div class="card-head"><h2>📋 Últimos partes</h2><a href="/partes">Ver todos →</a></div>
    <table>
      <thead><tr><th>Nº</th><th>Operario</th><th>Obra</th><th>Estado</th><th>Fecha</th><th></th></tr></thead>
      <tbody>{filas_partes or "<tr><td colspan='6' style='text-align:center;color:#aaa;padding:24px'>Sin partes</td></tr>"}</tbody>
    </table>
  </div>

  <div class="card">
    <div class="card-head"><h2>🔧 Herramienta en obra</h2><a href="/herramienta">Ver todo →</a></div>
    <table>
      <thead><tr><th>Herramienta</th><th>Operario</th><th>Obra</th><th>Desde</th></tr></thead>
      <tbody>{filas_herr or "<tr><td colspan='4' style='text-align:center;color:#aaa;padding:24px'>Nada en obra</td></tr>"}</tbody>
    </table>
  </div>

  <div class="card">
    <div class="card-head"><h2>⚠️ Alertas stock</h2><a href="/almacen">Ver almacén →</a></div>
    <div class="alertas">{alertas_stock}</div>
  </div>

  <div class="card card-full">
    <div class="card-head"><h2>📦 Últimos movimientos almacén</h2><a href="/almacen">Ver almacén →</a></div>
    <table>
      <thead><tr><th>Tipo</th><th>Material</th><th>Cantidad</th><th>Operario</th><th>Obra</th><th>Fecha</th></tr></thead>
      <tbody>{filas_mov or "<tr><td colspan='6' style='text-align:center;color:#aaa;padding:24px'>Sin movimientos</td></tr>"}</tbody>
    </table>
  </div>

</div>
</body></html>'''

@app.route('/health', methods=['GET'])
def health():
    return {
        'status': 'ok',
        'service': 'partes-instapalma',
        'zapia_notify_url': 'SET' if os.environ.get('ZAPIA_NOTIFY_URL') else 'NOT SET',
        'zapia_notify_token': 'SET' if os.environ.get('ZAPIA_NOTIFY_TOKEN') else 'NOT SET',
        'supervisor_wa': os.environ.get('SUPERVISOR_WA', 'default'),
        'gmail_user': os.environ.get('GMAIL_USER', 'NOT SET'),
        'gmail_password': 'SET' if os.environ.get('GMAIL_APP_PASSWORD') else 'NOT SET'
    }, 200

@app.route('/vehiculos')
def panel_vehiculos():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, matricula, marca_modelo, mes, km_inicio, km_fin, operario, created_at FROM vehiculos ORDER BY created_at DESC")
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        rows = []
    filas = ""
    for r in rows:
        filas += f"""<tr class="clickable" onclick="window.location='/vehiculos/{r[0]}'">
            <td>{r[0]}</td><td><b>{r[1]}</b></td><td>{r[2]}</td><td>{r[3]}</td>
            <td>{r[4]}</td><td>{r[5]}</td><td>{nombre_operario(r[6] or '')}</td>
            <td>{str(r[7])[:10]}</td>
            <td><a href='/vehiculos/{r[0]}/pdf' onclick='event.stopPropagation()'>📄 PDF</a></td>
        </tr>"""
    html = f"""<!DOCTYPE html><html><head><meta charset='utf-8'>
    <title>Vehículos — Instapalma</title>
    <style>{CSS_BASE}</style></head><body>
    <header><div><h1>🚗 Partes de Vehículos</h1><p>Instapalma</p></div></header>
    <div class='wrap'><table>
    <thead><tr><th>#</th><th>Matrícula</th><th>Modelo</th><th>Mes</th>
    <th>Km inicio</th><th>Km fin</th><th>Operario</th><th>Fecha</th><th>PDF</th></tr></thead>
    <tbody>{''.join([filas]) if rows else "<tr><td colspan=9 class='empty'>Sin registros</td></tr>"}</tbody>
    </table></div>
    <div style='padding:0 30px'><a href='/partes' class='back'>← Ver Partes de Trabajo</a></div>
    </body></html>"""
    from flask import Response
    return Response(html, mimetype='text/html')

@app.route('/vehiculos/<int:v_id>/pdf')
def descargar_pdf_vehiculo(v_id):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT * FROM vehiculos WHERE id=%s", (v_id,))
        r = cur.fetchone(); cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500
    if not r:
        return "Vehículo no encontrado", 404
    datos = {
        'matricula': r[1], 'marca_modelo': r[2], 'mes': r[3],
        'km_inicio': r[4], 'km_fin': r[5], 'proximo_aceite': r[6],
        'estado_neumaticos': r[7], 'conductores': r[8], 'mantenimientos': r[9],
        'observaciones': r[10], 'golpes': r[11],
        'fecha': str(r[13])[:10] if r[13] else ''
    }
    pdf_bytes = generar_pdf_vehiculo(datos)
    mat = (r[1] or 'VEHICULO').replace(' ','_').upper()
    mes = (r[3] or '').replace('/','_').replace(' ','_')
    nombre = f"{mes}-{mat}-VEHICULO.pdf"
    try:
        from datetime import datetime as dt
        conn2 = get_db(); cur2 = conn2.cursor()
        cur2.execute("UPDATE vehiculos SET pdf_descargado=TRUE, pdf_descargado_at=%s WHERE id=%s", (dt.utcnow(), v_id))
        conn2.commit(); cur2.close(); conn2.close()
    except Exception:
        pass
    from flask import Response
    return Response(pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nombre}"'})


# ── Panel Vacaciones ──────────────────────────────────────────────────────────
@app.route('/vacaciones')
def panel_vacaciones():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, nombre_operario, operario, fecha_inicio, fecha_fin, dias_solicitados, fecha_solicitud, estado, created_at FROM vacaciones ORDER BY created_at DESC")
        rows = cur.fetchall()
        cur.execute("SELECT operario, nombre, dias_totales, dias_usados FROM saldo_vacaciones ORDER BY nombre")
        saldos = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500

    filas_vac = ''
    for r in rows:
        vid, nombre, op, fi, ff, dias, fsol, estado, cat = r
        badge = {'aprobada':'badge-ok','rechazada':'badge-no','pendiente':'badge-pend'}.get(estado,'badge-pend')
        filas_vac += f"<tr><td>{vid}</td><td>{nombre or op}</td><td>{fi}</td><td>{ff}</td><td>{dias or '-'}</td><td>{fsol}</td><td><span class='{badge}'>{estado.upper()}</span></td></tr>"
    if not filas_vac:
        filas_vac = "<tr><td colspan=7 class='empty'>Sin solicitudes</td></tr>"

    filas_saldo = ''
    for s in saldos:
        op, nombre, tot, usados = s
        restantes = tot - usados
        color = 'color:#c62828' if restantes < 5 else 'color:#2e7d32'
        filas_saldo += f"<tr><td>{nombre}</td><td>{tot}</td><td>{usados}</td><td style='{color};font-weight:700'>{restantes}</td><td><a href='/vacaciones/saldo/{op}/editar' style='color:#1a3a5c;font-size:12px'>✏️ Editar</a></td></tr>"
    if not filas_saldo:
        filas_saldo = "<tr><td colspan=5 class='empty'>Sin saldos configurados</td></tr>"

    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>Vacaciones — Instapalma</title>
    <style>{CSS_BASE}
    .badge-ok{{background:#2e7d32;color:white;padding:3px 10px;border-radius:10px;font-size:11px}}
    .badge-no{{background:#c62828;color:white;padding:3px 10px;border-radius:10px;font-size:11px}}
    .badge-pend{{background:#e65100;color:white;padding:3px 10px;border-radius:10px;font-size:11px}}
    </style></head><body>
    <header><div><h1>🌴 Vacaciones</h1><p>Instapalma</p></div>
    <a href='/vacaciones/saldo/nuevo' style='background:white;color:#1a3a5c;padding:8px 16px;border-radius:8px;font-weight:700;text-decoration:none;font-size:13px'>+ Añadir saldo</a>
    </header>
    <div class='wrap' style='padding-top:20px'>
    <h3 style='color:#1a3a5c;margin-bottom:12px'>Saldo de Vacaciones por Operario</h3>
    <table><thead><tr><th>Operario</th><th>Total días</th><th>Usados</th><th>Restantes</th><th></th></tr></thead>
    <tbody>{filas_saldo}</tbody></table>
    <h3 style='color:#1a3a5c;margin:24px 0 12px'>Solicitudes</h3>
    <table><thead><tr><th>#</th><th>Operario</th><th>Inicio</th><th>Fin</th><th>Días</th><th>Solicitado</th><th>Estado</th></tr></thead>
    <tbody>{filas_vac}</tbody></table>
    </div>
    <div style='padding:0 30px'><a href='/partes' class='back'>← Ver Partes de Trabajo</a></div>
    </body></html>"""

@app.route('/vacaciones/saldo/nuevo', methods=['GET','POST'])
@app.route('/vacaciones/saldo/<path:op>/editar', methods=['GET','POST'])
def editar_saldo(op=None):
    from flask import request as req
    if req.method == 'POST':
        nombre = req.form.get('nombre','')
        operario = req.form.get('operario','')
        dias_totales = int(req.form.get('dias_totales', 23))
        dias_usados = int(req.form.get('dias_usados', 0))
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute("""
                INSERT INTO saldo_vacaciones (operario, nombre, dias_totales, dias_usados, anio, updated_at)
                VALUES (%s, %s, %s, %s, 2026, NOW())
                ON CONFLICT (operario) DO UPDATE SET nombre=%s, dias_totales=%s, dias_usados=%s, updated_at=NOW()
            """, (operario, nombre, dias_totales, dias_usados, nombre, dias_totales, dias_usados))
            conn.commit(); cur.close(); conn.close()
        except Exception as e:
            return f"Error: {e}", 500
        from flask import redirect
        return redirect('/vacaciones')

    # GET — cargar datos si es edición
    datos = {'operario': op or '', 'nombre': '', 'dias_totales': 23, 'dias_usados': 0}
    if op:
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute("SELECT operario, nombre, dias_totales, dias_usados FROM saldo_vacaciones WHERE operario=%s", (op,))
            r = cur.fetchone()
            cur.close(); conn.close()
            if r:
                datos = {'operario': r[0], 'nombre': r[1], 'dias_totales': r[2], 'dias_usados': r[3]}
        except: pass
    titulo = 'Editar saldo' if op else 'Nuevo saldo'
    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>{titulo} — Instapalma</title>
    <style>{CSS_BASE} label{{display:block;margin-bottom:4px;font-size:13px;font-weight:600;color:#555}}
    input{{width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px;margin-bottom:16px;box-sizing:border-box}}
    .btn{{background:#1a3a5c;color:white;padding:12px 28px;border:none;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer}}</style></head><body>
    <header><div><h1>🌴 {titulo}</h1><p>Instapalma</p></div></header>
    <div style='max-width:500px;margin:30px auto;background:white;padding:28px;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.1)'>
    <form method='POST'>
    <label>Número operario (ej: 34636606175)</label><input name='operario' value='{datos["operario"]}' required>
    <label>Nombre</label><input name='nombre' value='{datos["nombre"]}' required>
    <label>Días totales 2026</label><input name='dias_totales' type='number' value='{datos["dias_totales"]}' required>
    <label>Días ya usados</label><input name='dias_usados' type='number' value='{datos["dias_usados"]}' required>
    <button class='btn' type='submit'>Guardar</button>
    </form>
    </div>
    <div style='padding:0 30px'><a href='/vacaciones' class='back'>← Volver</a></div>
    </body></html>"""


# ── Resumen Fin de Mes ────────────────────────────────────────────────────────

def guardar_resumen_mes(datos, numero_operario):
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        nombre = datos.get('nombre_operario', nombre_operario(numero_operario))
        cur.execute("""
            INSERT INTO resumen_mes (operario, nombre_operario, mes, horas_extra, dias_vacaciones, total_gastos, foto_url, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            RETURNING id
        """, (
            numero_operario,
            nombre,
            datos.get('mes', ''),
            datos.get('horas_extra', '0'),
            datos.get('dias_vacaciones', '0'),
            datos.get('total_gastos', '0'),
            datos.get('foto_url', ''),
        ))
        rid = cur.fetchone()[0]
        conn.commit()
        return rid
    except Exception as e:
        print(f"Error guardar_resumen_mes: {e}")
        if conn: conn.rollback()
        return None
    finally:
        try:
            if conn: conn.close()
        except: pass


def generar_pdf_resumen_mes(datos):
    """Genera PDF del resumen de fin de mes y devuelve bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    AZUL = colors.HexColor('#1a3a5c')
    GRIS = colors.HexColor('#f5f5f5')

    titulo_style = ParagraphStyle('titulo', fontSize=20, textColor=AZUL,
        alignment=TA_CENTER, spaceAfter=4, fontName='Helvetica-Bold')
    pie_style = ParagraphStyle('pie', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    import os as _os
    LOGO_PATH = _os.path.join(_os.path.dirname(__file__), 'logo.jpg')
    if _os.path.exists(LOGO_PATH):
        _logo_ratio = 1024 / 219
        _logo_w = 4 * cm
        _logo_h = _logo_w / _logo_ratio
        logo_img = RLImage(LOGO_PATH, width=_logo_w, height=_logo_h)
    else:
        logo_img = Paragraph("INSTAPALMA", titulo_style)

    cab_title_style = ParagraphStyle('cab_title', fontName='Helvetica-Bold', fontSize=16,
        textColor=AZUL, alignment=1, leading=20)
    cab = Table([[logo_img, Paragraph('RESUMEN FIN DE MES', cab_title_style)]], colWidths=[5*cm, 12*cm])
    cab.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(cab)
    elements.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=10))
    elements.append(Spacer(1, 0.3*cm))

    filas = [
        ['Mes',                   datos.get('mes','')],
        ['Operario',              datos.get('nombre_operario','')],
        ['Horas extra',           datos.get('horas_extra','0')],
        ['Días vacaciones',       datos.get('dias_vacaciones','0')],
        ['Total gastos',          f"{datos.get('total_gastos','0')} €"],
    ]
    t = Table(filas, colWidths=[5*cm, 12*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), AZUL),
        ('TEXTCOLOR', (0,0), (0,-1), colors.white),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (1,0), (1,-1), [colors.white, GRIS]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.5*cm))

    # Foto si existe
    foto_url = datos.get('foto_url', '')
    if foto_url:
        try:
            import urllib.request as _ur
            import tempfile, os as _os
            suffix = '.jpg'
            if '.png' in foto_url.lower(): suffix = '.png'
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            _ur.urlretrieve(foto_url, tmp.name)
            sec_style = ParagraphStyle('sec', fontSize=10, textColor=colors.white,
                backColor=AZUL, fontName='Helvetica-Bold', spaceAfter=4, spaceBefore=6, borderPad=4)
            elements.append(Paragraph('JUSTIFICANTE / FOTO', sec_style))
            elements.append(Spacer(1, 0.2*cm))
            img_w = 14*cm
            from PIL import Image as PILImg
            with PILImg.open(tmp.name) as im:
                w, h = im.size
                img_h = img_w * h / w
                if img_h > 18*cm:
                    img_h = 18*cm
                    img_w = img_h * w / h
            foto_img = RLImage(tmp.name, width=img_w, height=img_h)
            elements.append(foto_img)
            _os.unlink(tmp.name)
        except Exception as e:
            print(f"Error foto PDF: {e}")

    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph("Instapalma · Resumen generado automáticamente", pie_style))

    doc.build(elements)
    return buffer.getvalue()


def subir_pdf_resumen_mes(pdf_bytes, rid):
    """Guarda el PDF en la BD y devuelve la URL pública del bot (sin Cloudinary)."""
    try:
        BOT_URL = os.environ.get('RAILWAY_PUBLIC_DOMAIN', 'bot-production-66b8.up.railway.app')
        ref = f"RESUMEN-MES-{rid}"
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO stock_albaranes (numero, pdf_bytes)
            VALUES (%s, %s)
            ON CONFLICT (numero) DO UPDATE SET pdf_bytes = %s
        """, (ref, psycopg2.Binary(pdf_bytes), psycopg2.Binary(pdf_bytes)))
        conn.commit(); cur.close(); conn.close()
        return f"https://{BOT_URL}/albaran/{ref}.pdf"
    except Exception as e:
        print(f"Error guardar PDF resumen mes: {e}")
        return ''


def finalizar_resumen_mes(numero, datos):
    """Guarda, genera PDF, envía por WhatsApp (URL) y por email con adjunto."""
    import threading
    def _enviar():
        rid = guardar_resumen_mes(datos, numero)
        nombre_op = datos.get('nombre_operario', nombre_operario(numero))
        mes = datos.get('mes','')
        horas = datos.get('horas_extra','0')
        dias_vac = datos.get('dias_vacaciones','0')
        gastos = datos.get('total_gastos','0')

        texto_wa = (
            f"📊 *RESUMEN FIN DE MES #{rid}*\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📅 Mes: {mes}\n"
            f"👷 {nombre_op}\n"
            f"⏱ Horas extra: {horas}\n"
            f"🌴 Días vacaciones: {dias_vac}\n"
            f"💶 Total gastos: {gastos} €\n"
            f"━━━━━━━━━━━━━━━━━━━━"
        )

        # Generar PDF
        pdf_bytes_data = generar_pdf_resumen_mes(datos)
        pdf_url = subir_pdf_resumen_mes(pdf_bytes_data, rid)

        # Enviar por WhatsApp al supervisor con enlace al PDF
        enviar_supervisor(texto_wa + ("\n📄 PDF: " + pdf_url if pdf_url else ""))
        # Enviar al operario
        op_wa = numero if numero.startswith('whatsapp:') else f'whatsapp:+{numero.lstrip("+")}'
        enviar_whatsapp(op_wa, f"✅ Resumen de {mes} enviado correctamente." + (f"\n📄 Tu copia: {pdf_url}" if pdf_url else ""))

        # Enviar por email con PDF adjunto
        try:
            nombre_pdf = f"Resumen_Mes_{nombre_op.replace(' ','_')}_{mes.replace(' ','_')}.pdf"
            msg_email = MIMEMultipart()
            msg_email['From']    = GMAIL_USER
            msg_email['To']      = ', '.join([SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2])
            msg_email['Subject'] = f"[RESUMEN MES] {nombre_op} — {mes}"
            body_txt = (
                f"Resumen de fin de mes recibido.\n\n"
                f"Operario: {nombre_op}\n"
                f"Mes: {mes}\n"
                f"Horas extra: {horas}\n"
                f"Días vacaciones: {dias_vac}\n"
                f"Total gastos: {gastos} €\n"
            )
            msg_email.attach(MIMEText(body_txt, 'plain'))
            part = MIMEApplication(pdf_bytes_data, Name=nombre_pdf)
            part.add_header('Content-Disposition', f'attachment; filename="{nombre_pdf}"')
            msg_email.attach(part)
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as srv:
                srv.login(GMAIL_USER, GMAIL_APP_PASSWORD)
                srv.sendmail(GMAIL_USER, [SUPERVISOR_EMAIL_1, SUPERVISOR_EMAIL_2], msg_email.as_string())
            print(f"Email resumen mes OK — {nombre_op} {mes}")
        except Exception as e:
            print(f"Error email resumen mes: {e}")

    t = threading.Thread(target=_enviar, daemon=True)
    t.start()


# ── Panel web Resumen Fin de Mes ──────────────────────────────────────────────
@app.route('/resumenes')
def panel_resumenes():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            SELECT id, nombre_operario, operario, mes, horas_extra, dias_vacaciones, total_gastos, foto_url, created_at
            FROM resumen_mes ORDER BY created_at DESC
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500

    filas = ''
    for r in rows:
        rid, nombre, op, mes, horas, dias_vac, gastos, foto_url, cat = r
        foto_link = f"<a href='{foto_url}' target='_blank'>📎 Ver</a>" if foto_url else '-'
        fecha = str(cat)[:10] if cat else ''
        filas += (
            f"<tr>"
            f"<td>{rid}</td>"
            f"<td>{nombre or op}</td>"
            f"<td>{mes}</td>"
            f"<td>{horas}</td>"
            f"<td>{dias_vac}</td>"
            f"<td>{gastos} €</td>"
            f"<td>{foto_link}</td>"
            f"<td>{fecha}</td>"
            f"<td><a href='/resumenes/{rid}/pdf' style='color:#1a3a5c;font-size:12px'>⬇ PDF</a></td>"
            f"</tr>"
        )
    if not filas:
        filas = "<tr><td colspan=9 class='empty'>Sin resúmenes aún</td></tr>"

    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'>
    <title>Resumen Fin de Mes — Instapalma</title>
    <style>{CSS_BASE}</style></head><body>
    <header><div><h1>📊 Resumen Fin de Mes</h1><p>Instapalma</p></div></header>
    <div class='wrap' style='padding-top:20px'>
    <table><thead><tr>
    <th>#</th><th>Operario</th><th>Mes</th><th>H. Extra</th><th>Días Vac.</th><th>Gastos</th><th>Foto</th><th>Fecha</th><th>PDF</th>
    </tr></thead><tbody>{filas}</tbody></table>
    </div>
    <div style='padding:0 30px'><a href='/partes' class='back'>← Ver Partes de Trabajo</a></div>
    </body></html>"""


@app.route('/resumenes/<int:rid>/pdf')
def pdf_resumen(rid):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT nombre_operario, operario, mes, horas_extra, dias_vacaciones, total_gastos, foto_url FROM resumen_mes WHERE id=%s", (rid,))
        r = cur.fetchone(); cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500
    if not r:
        return "No encontrado", 404
    datos = {
        'nombre_operario': r[0] or r[1],
        'mes': r[2], 'horas_extra': r[3],
        'dias_vacaciones': r[4], 'total_gastos': r[5],
        'foto_url': r[6] or ''
    }
    pdf_bytes = generar_pdf_resumen_mes(datos)
    nombre_fichero = f"RESUMEN-{(r[2] or 'MES').replace(' ','_').upper()}-{(r[0] or '').replace(' ','_').upper()}.pdf"
    from flask import Response
    return Response(pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nombre_fichero}"'})


    return Response(pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nombre_fichero}"'})


# ══════════════════════════════════════════════════════════════════════════════
# GESTIÓN DE STOCK DE ALMACÉN
# ══════════════════════════════════════════════════════════════════════════════

MENSAJES_STOCK_SALIDA    = ['salida']
MENSAJES_STOCK_DEVOL     = ['devolucion', 'devolución']
MENSAJES_STOCK_CONSULTA  = ['consulta']

# ── DB ────────────────────────────────────────────────────────────────────────

def init_stock_db():
    conn = None
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS stock_materiales (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(200) UNIQUE,
                unidad VARCHAR(30) DEFAULT 'ud',
                stock_actual NUMERIC(12,3) DEFAULT 0,
                stock_minimo NUMERIC(12,3) DEFAULT 0,
                precio_unitario NUMERIC(12,4) DEFAULT 0,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Migración: añadir columna precio_unitario si no existe
        cur.execute("""
            DO $$ BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='stock_materiales' AND column_name='precio_unitario'
                ) THEN
                    ALTER TABLE stock_materiales ADD COLUMN precio_unitario NUMERIC(12,4) DEFAULT 0;
                END IF;
            END $$;
        """)
        # Migración: añadir columna familia si no existe
        cur.execute("""
            DO $$ BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM information_schema.columns
                    WHERE table_name='stock_materiales' AND column_name='familia'
                ) THEN
                    ALTER TABLE stock_materiales ADD COLUMN familia VARCHAR(100) DEFAULT 'General';
                END IF;
            END $$;
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS stock_movimientos (
                id SERIAL PRIMARY KEY,
                tipo VARCHAR(20),
                material_id INTEGER REFERENCES stock_materiales(id),
                material_nombre VARCHAR(200),
                cantidad NUMERIC(12,3),
                unidad VARCHAR(30),
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                obra VARCHAR(200),
                albaran_id INTEGER,
                notas TEXT,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS stock_albaranes (
                id SERIAL PRIMARY KEY,
                numero VARCHAR(100) UNIQUE,
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                obra VARCHAR(200),
                lineas JSONB,
                pdf_url TEXT,
                pdf_bytes BYTEA,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Migración: añadir columna pdf_bytes si no existe
        cur.execute("""
            ALTER TABLE stock_albaranes ADD COLUMN IF NOT EXISTS pdf_bytes BYTEA
        """)
        # Migración: ampliar numero de VARCHAR(30) a VARCHAR(100)
        cur.execute("""
            ALTER TABLE stock_albaranes ALTER COLUMN numero TYPE VARCHAR(100)
        """)
        conn.commit(); cur.close(); conn.close()
        print("Stock DB OK")
    except Exception as e:
        print(f"Error init_stock_db: {e}")
        if conn:
            try: conn.rollback()
            except: pass
    finally:
        try:
            if conn: conn.close()
        except: pass

init_stock_db()

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO HERRAMIENTA — Control de herramienta, EPIs y asignaciones de obra
# ═══════════════════════════════════════════════════════════════════════════════

def init_herramienta_db():
    conn = None
    try:
        conn = get_db(); cur = conn.cursor()
        # Tabla maestra de herramienta
        cur.execute("""
            CREATE TABLE IF NOT EXISTS herramienta (
                id SERIAL PRIMARY KEY,
                nombre VARCHAR(200) UNIQUE NOT NULL,
                tipo VARCHAR(30) DEFAULT 'general',  -- general | personal | epi
                stock_almacen INTEGER DEFAULT 0,
                propietario VARCHAR(100) DEFAULT NULL,  -- para tipo=personal
                observaciones TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        # Herramienta en obra (asignaciones activas)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS herramienta_obra (
                id SERIAL PRIMARY KEY,
                herramienta_id INTEGER REFERENCES herramienta(id),
                herramienta_nombre VARCHAR(200),
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                obra VARCHAR(200),
                fecha_alta TIMESTAMP DEFAULT NOW(),
                fecha_baja TIMESTAMP DEFAULT NULL,
                activo BOOLEAN DEFAULT TRUE
            )
        """)
        # EPIs por persona
        cur.execute("""
            CREATE TABLE IF NOT EXISTS herramienta_epis (
                id SERIAL PRIMARY KEY,
                operario VARCHAR(100),
                nombre_operario VARCHAR(100),
                epi VARCHAR(200),
                talla VARCHAR(30) DEFAULT '',
                fecha_entrega DATE DEFAULT NOW(),
                observaciones TEXT DEFAULT ''
            )
        """)
        # Herramienta y EPIs asignados a operario (tabla unificada personal)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS herramienta_personal (
                id SERIAL PRIMARY KEY,
                propietario VARCHAR(150) NOT NULL,
                articulo VARCHAR(300) NOT NULL,
                tipo VARCHAR(20) DEFAULT 'herramienta',  -- herramienta | epi
                fecha_alta TIMESTAMP DEFAULT NOW()
            )
        """)
        conn.commit(); cur.close(); conn.close()
        print("Herramienta DB OK")
    except Exception as e:
        print(f"Error init_herramienta_db: {e}")
        if conn:
            try: conn.rollback()
            except: pass

init_herramienta_db()

def buscar_herramienta(nombre_buscado):
    """Busca herramienta por nombre (coincidencia parcial). Devuelve (id, nombre, tipo, stock_almacen) o None."""
    import unicodedata as _ud
    def _norm(t):
        return ''.join(c for c in _ud.normalize('NFD', t.lower()) if _ud.category(c) != 'Mn')
    nb = _norm(nombre_buscado)
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, nombre, tipo, stock_almacen, propietario FROM herramienta ORDER BY nombre")
        rows = cur.fetchall(); cur.close(); conn.close()
        # Primero exact match
        for r in rows:
            if _norm(r[1]) == nb:
                return r
        # Luego substring
        for r in rows:
            if nb in _norm(r[1]) or _norm(r[1]) in nb:
                return r
        return None
    except Exception as e:
        print(f"Error buscar_herramienta: {e}")
        return None

def herramienta_alta_obra(nombre_herr, obra, operario, nombre_op):
    """Mueve 1 unidad de almacén → obra. Devuelve (ok, mensaje)."""
    herr = buscar_herramienta(nombre_herr)
    if not herr:
        return False, f"❌ No encuentro *{nombre_herr}* en el inventario.\n_Escribe *herramienta almacen* para ver el listado._"
    hid, hnombre, htipo, hstock, _ = herr
    if htipo == 'personal':
        return False, f"⚠️ *{hnombre}* es herramienta personal, no se puede asignar a obra desde aquí."
    if hstock < 1:
        return False, f"⚠️ No hay stock de *{hnombre}* en almacén (disponible: {hstock})."
    try:
        conn = get_db(); cur = conn.cursor()
        # Descontar almacén
        cur.execute("UPDATE herramienta SET stock_almacen = stock_almacen - 1, updated_at=NOW() WHERE id=%s", (hid,))
        # Registrar en obra
        cur.execute("""
            INSERT INTO herramienta_obra (herramienta_id, herramienta_nombre, operario, nombre_operario, obra, activo)
            VALUES (%s,%s,%s,%s,%s,TRUE)
        """, (hid, hnombre, operario, nombre_op, obra))
        conn.commit(); cur.close(); conn.close()
        return True, f"✅ *{hnombre}* asignada a obra *{obra}*\nStock almacén: {hstock-1} ud."
    except Exception as e:
        return False, f"❌ Error: {e}"

def herramienta_baja_obra(nombre_herr, obra, operario, nombre_op):
    """Devuelve herramienta de obra → almacén. Devuelve (ok, mensaje)."""
    herr = buscar_herramienta(nombre_herr)
    if not herr:
        return False, f"❌ No encuentro *{nombre_herr}* en el inventario."
    hid, hnombre, htipo, hstock, _ = herr
    try:
        conn = get_db(); cur = conn.cursor()
        # Buscar asignación activa (por herramienta y obra, o solo herramienta si obra no especificada)
        if obra:
            import unicodedata as _ud
            def _norm(t): return ''.join(c for c in _ud.normalize('NFD', t.lower()) if _ud.category(c) != 'Mn')
            cur.execute("""
                SELECT id FROM herramienta_obra
                WHERE herramienta_id=%s AND activo=TRUE AND LOWER(obra) LIKE %s
                ORDER BY fecha_alta DESC LIMIT 1
            """, (hid, f'%{_norm(obra)}%'))
        else:
            cur.execute("""
                SELECT id FROM herramienta_obra
                WHERE herramienta_id=%s AND activo=TRUE
                ORDER BY fecha_alta DESC LIMIT 1
            """, (hid,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return False, f"⚠️ No hay ninguna *{hnombre}* activa en obra *{obra or 'ninguna'}*."
        # Cerrar asignación
        cur.execute("UPDATE herramienta_obra SET activo=FALSE, fecha_baja=NOW() WHERE id=%s", (row[0],))
        # Sumar al almacén
        cur.execute("UPDATE herramienta SET stock_almacen = stock_almacen + 1, updated_at=NOW() WHERE id=%s", (hid,))
        conn.commit(); cur.close(); conn.close()
        return True, f"✅ *{hnombre}* devuelta de obra *{obra or ''}* al almacén\nStock almacén: {hstock+1} ud."
    except Exception as e:
        return False, f"❌ Error: {e}"

def generar_pdf_herramienta(seccion='todo'):
    """Genera PDF de herramienta. seccion: 'almacen'|'obra'|'personal'|'epis'|'todo'"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import io as _io
    from datetime import datetime as _dt

    AZUL = colors.HexColor('#1a3a5c')
    GRIS = colors.HexColor('#f5f5f5')
    NARANJA = colors.HexColor('#e67e22')
    VERDE = colors.HexColor('#27ae60')

    buffer = _io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []

    titulo_style = ParagraphStyle('T', fontSize=16, textColor=AZUL, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=4)
    sec_style = ParagraphStyle('S', fontSize=10, textColor=colors.white, backColor=AZUL,
        fontName='Helvetica-Bold', spaceAfter=2, spaceBefore=8, borderPad=4)
    pie_style = ParagraphStyle('P', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)
    normal = ParagraphStyle('N', fontSize=9)

    fecha_str = _dt.now().strftime('%d/%m/%Y %H:%M')
    titulos_map = {
        'almacen': 'HERRAMIENTA EN ALMACÉN',
        'obra': 'HERRAMIENTA EN OBRA',
        'personal': 'HERRAMIENTA PERSONAL',
        'epis': 'EPIs POR TRABAJADOR',
        'todo': 'INVENTARIO COMPLETO DE HERRAMIENTA',
    }
    elements.append(Paragraph(f"INSTAPALMA — {titulos_map.get(seccion,'HERRAMIENTA')}", titulo_style))
    elements.append(Paragraph(f"Generado: {fecha_str}", ParagraphStyle('f', fontSize=8, alignment=TA_CENTER, textColor=colors.grey)))
    elements.append(Spacer(1, 0.4*cm))

    try:
        conn = get_db(); cur = conn.cursor()

        # ── ALMACÉN ───────────────────────────────────────────────────────────
        if seccion in ('almacen', 'todo'):
            cur.execute("""
                SELECT nombre, tipo, stock_almacen, observaciones
                FROM herramienta WHERE tipo != 'personal'
                ORDER BY tipo, nombre
            """)
            rows = cur.fetchall()
            elements.append(Paragraph("📦 ALMACÉN", sec_style))
            elements.append(Spacer(1, 0.1*cm))
            if rows:
                filas = [['Herramienta', 'Tipo', 'Stock', 'Obs.']]
                for r in rows:
                    filas.append([r[0], r[1].capitalize(), str(r[2]), r[3] or ''])
                t = Table(filas, colWidths=[8*cm, 3*cm, 2*cm, 4*cm])
                t.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),AZUL), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),8),
                    ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey), ('PADDING',(0,0),(-1,-1),5),
                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, GRIS]),
                    ('ALIGN',(2,0),(2,-1),'CENTER'),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("Sin herramienta en almacén.", normal))
            elements.append(Spacer(1, 0.3*cm))

        # ── EN OBRA ───────────────────────────────────────────────────────────
        if seccion in ('obra', 'todo'):
            cur.execute("""
                SELECT herramienta_nombre, nombre_operario, obra, fecha_alta
                FROM herramienta_obra WHERE activo=TRUE
                ORDER BY obra, herramienta_nombre
            """)
            rows = cur.fetchall()
            elements.append(Paragraph("🏗️ EN OBRA", sec_style))
            elements.append(Spacer(1, 0.1*cm))
            if rows:
                filas = [['Herramienta', 'Operario', 'Obra', 'Desde']]
                for r in rows:
                    fecha = r[3].strftime('%d/%m/%Y') if r[3] else ''
                    filas.append([r[0], r[1] or '', r[2] or '', fecha])
                t = Table(filas, colWidths=[6*cm, 4*cm, 4*cm, 3*cm])
                t.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),NARANJA), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),8),
                    ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey), ('PADDING',(0,0),(-1,-1),5),
                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, GRIS]),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("No hay herramienta actualmente en obra.", normal))
            elements.append(Spacer(1, 0.3*cm))

        # ── PERSONAL ──────────────────────────────────────────────────────────
        if seccion in ('personal', 'todo'):
            cur.execute("""
                SELECT nombre, propietario, observaciones
                FROM herramienta WHERE tipo='personal'
                ORDER BY propietario, nombre
            """)
            rows = cur.fetchall()
            elements.append(Paragraph("👷 HERRAMIENTA PERSONAL", sec_style))
            elements.append(Spacer(1, 0.1*cm))
            if rows:
                filas = [['Herramienta', 'Asignada a', 'Observaciones']]
                for r in rows:
                    filas.append([r[0], r[1] or '', r[2] or ''])
                t = Table(filas, colWidths=[7*cm, 5*cm, 5*cm])
                t.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),VERDE), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),8),
                    ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey), ('PADDING',(0,0),(-1,-1),5),
                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, GRIS]),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("Sin registros de herramienta personal.", normal))
            elements.append(Spacer(1, 0.3*cm))

        # ── EPIs ──────────────────────────────────────────────────────────────
        if seccion in ('epis', 'todo'):
            cur.execute("""
                SELECT nombre_operario, epi, talla, fecha_entrega, observaciones
                FROM herramienta_epis
                ORDER BY nombre_operario, epi
            """)
            rows = cur.fetchall()
            elements.append(Paragraph("🦺 EPIs", sec_style))
            elements.append(Spacer(1, 0.1*cm))
            if rows:
                filas = [['Trabajador', 'EPI', 'Talla', 'Entrega', 'Obs.']]
                for r in rows:
                    fecha = r[3].strftime('%d/%m/%Y') if r[3] else ''
                    filas.append([r[0] or '', r[1], r[2] or '', fecha, r[4] or ''])
                t = Table(filas, colWidths=[4.5*cm, 4*cm, 2*cm, 3*cm, 3.5*cm])
                t.setStyle(TableStyle([
                    ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#6c3483')), ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                    ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'), ('FONTSIZE',(0,0),(-1,-1),8),
                    ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey), ('PADDING',(0,0),(-1,-1),5),
                    ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, GRIS]),
                ]))
                elements.append(t)
            else:
                elements.append(Paragraph("Sin registros de EPIs.", normal))

        cur.close(); conn.close()
    except Exception as e:
        elements.append(Paragraph(f"Error generando listado: {e}", normal))

    elements.append(Spacer(1, 0.5*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Paragraph(f"Instapalma · Control de Herramienta — {fecha_str}", pie_style))
    doc.build(elements)
    return buffer.getvalue()


def get_material_by_nombre(nombre):
    """Busca material por nombre exacto o aproximado. Devuelve (id, nombre, unidad, stock_actual, stock_minimo)."""
    import unicodedata as _ud
    def _norm(t):
        # Quitar tildes y pasar a minúsculas para comparación tolerante
        return ''.join(c for c in _ud.normalize('NFD', t.lower()) if _ud.category(c) != 'Mn')

    conn = None
    try:
        conn = get_db(); cur = conn.cursor()
        # Exacto (sin tildes)
        cur.execute("SELECT id, nombre, unidad, stock_actual, stock_minimo, precio_unitario FROM stock_materiales WHERE LOWER(nombre)=LOWER(%s)", (nombre,))
        r = cur.fetchone()
        if r:
            return r
        # Aproximado con ILIKE sobre cada palabra del texto buscado (tolerante a tildes)
        palabras = [p for p in _norm(nombre).split() if len(p) > 1]
        # Buscar todos y filtrar en Python para tolerancia de tildes
        cur.execute("SELECT id, nombre, unidad, stock_actual, stock_minimo, precio_unitario FROM stock_materiales ORDER BY nombre")
        todos = cur.fetchall()
        cur.close(); conn.close()
        resultados = []
        for row in todos:
            nombre_norm = _norm(row[1])
            if all(p in nombre_norm for p in palabras):
                resultados.append(row)
        if resultados:
            return resultados[:5]
        # Fallback: cualquier palabra coincide (búsqueda laxa)
        resultados_laxa = []
        for row in todos:
            nombre_norm = _norm(row[1])
            if any(p in nombre_norm for p in palabras if len(p) > 2):
                resultados_laxa.append(row)
        return resultados_laxa[:5]
    except Exception as e:
        print(f"Error get_material: {e}")
        return None
    finally:
        try:
            if conn: conn.close()
        except: pass

def ajustar_stock(material_id, delta):
    """Suma delta al stock (delta negativo = salida)."""
    conn = None
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            UPDATE stock_materiales SET stock_actual = stock_actual + %s, updated_at=NOW()
            WHERE id=%s RETURNING stock_actual, stock_minimo, nombre, unidad
        """, (delta, material_id))
        r = cur.fetchone()
        conn.commit(); cur.close(); conn.close()
        return r  # (stock_actual, stock_minimo, nombre, unidad)
    except Exception as e:
        print(f"Error ajustar_stock: {e}")
        if conn: conn.rollback()
        return None
    finally:
        try:
            if conn: conn.close()
        except: pass

def registrar_movimiento(tipo, material_id, material_nombre, cantidad, unidad, operario, nombre_op, obra='', albaran_id=None, notas=''):
    conn = None
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO stock_movimientos (tipo, material_id, material_nombre, cantidad, unidad, operario, nombre_operario, obra, albaran_id, notas)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (tipo, material_id, material_nombre, cantidad, unidad, operario, nombre_op, obra, albaran_id, notas))
        conn.commit(); cur.close(); conn.close()
    except Exception as e:
        print(f"Error registrar_movimiento: {e}")
        if conn: conn.rollback()
    finally:
        try:
            if conn: conn.close()
        except: pass

def crear_albaran(numero, operario, nombre_op, obra, lineas):
    """Crea albarán y devuelve su id."""
    conn = None
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("""
            INSERT INTO stock_albaranes (numero, operario, nombre_operario, obra, lineas)
            VALUES (%s,%s,%s,%s,%s) RETURNING id
        """, (numero, operario, nombre_op, obra, json.dumps(lineas)))
        aid = cur.fetchone()[0]
        conn.commit(); cur.close(); conn.close()
        return aid
    except Exception as e:
        print(f"Error crear_albaran: {e}")
        if conn: conn.rollback()
        return None
    finally:
        try:
            if conn: conn.close()
        except: pass

def siguiente_numero_albaran():
    conn = None
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM stock_albaranes")
        n = cur.fetchone()[0] + 1
        cur.close(); conn.close()
        from datetime import datetime as dt
        return f"ALB-{dt.now().strftime('%Y%m')}-{n:04d}"
    except:
        from datetime import datetime as dt
        return f"ALB-{dt.now().strftime('%Y%m%d%H%M%S')}"
    finally:
        try:
            if conn: conn.close()
        except: pass

# ── PDF albarán interno ───────────────────────────────────────────────────────

def generar_pdf_stock(titulo="LISTADO DE STOCK", familia_filtro=None):
    """Genera un PDF con el listado de stock, agrupado por familia. Si familia_filtro se filtra por esa familia."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    import io as _io
    from datetime import datetime as _dt

    AZUL = colors.HexColor('#1a3a5c')
    GRIS = colors.HexColor('#f5f5f5')
    ROJO = colors.HexColor('#c0392b')

    buffer = _io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=1.5*cm)
    elements = []

    # Estilos
    titulo_style = ParagraphStyle('t', fontSize=16, textColor=AZUL, alignment=TA_CENTER,
        fontName='Helvetica-Bold', spaceAfter=4)
    sub_style   = ParagraphStyle('s', fontSize=9, textColor=colors.grey, alignment=TA_CENTER, spaceAfter=10)
    sec_style   = ParagraphStyle('fam', fontSize=10, textColor=colors.white, backColor=AZUL,
        fontName='Helvetica-Bold', spaceBefore=8, spaceAfter=2, borderPad=4)
    pie_style   = ParagraphStyle('p', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    fecha_str = _dt.now().strftime('%d/%m/%Y %H:%M')
    filtro_txt = f" — Familia: {familia_filtro}" if familia_filtro else " — Todo el stock"
    elements.append(Paragraph(titulo, titulo_style))
    elements.append(Paragraph(f"Generado: {fecha_str}{filtro_txt}", sub_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=AZUL))
    elements.append(Spacer(1, 0.3*cm))

    # Obtener materiales de BD
    try:
        conn = get_db(); cur = conn.cursor()
        if familia_filtro:
            cur.execute("""
                SELECT nombre, unidad, stock_actual, stock_minimo, precio_unitario,
                       COALESCE(familia, 'General') as familia
                FROM stock_materiales
                WHERE LOWER(COALESCE(familia,'General')) = LOWER(%s) AND stock_actual >= 0
                ORDER BY nombre
            """, (familia_filtro,))
        else:
            cur.execute("""
                SELECT nombre, unidad, stock_actual, stock_minimo, precio_unitario,
                       COALESCE(familia, 'General') as familia
                FROM stock_materiales
                WHERE stock_actual >= 0
                ORDER BY COALESCE(familia,'General'), nombre
            """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        rows = []

    if not rows:
        elements.append(Paragraph("No hay artículos en stock.", sub_style))
    else:
        # Agrupar por familia
        from collections import OrderedDict
        familias = OrderedDict()
        for r in rows:
            fam = r[5] or 'General'
            if fam not in familias:
                familias[fam] = []
            familias[fam].append(r)

        total_valorado = 0.0
        for fam, arts in familias.items():
            elements.append(Paragraph(f"  {fam.upper()}", sec_style))
            filas = [['Artículo', 'Stock', 'Ud.', 'Mín.', 'P.Unit (€)', 'Valor (€)']]
            subtotal_fam = 0.0
            for art in arts:
                nombre, unidad, stock, minimo, precio = art[0], art[1], float(art[2]), float(art[3]), float(art[4] or 0)
                valor = stock * precio
                subtotal_fam += valor
                alerta = ' ⚠️' if stock <= minimo and minimo > 0 else ''
                filas.append([
                    nombre + alerta,
                    fmt_cant(stock),
                    unidad,
                    fmt_cant(minimo) if minimo > 0 else '—',
                    fmt_cant(precio) if precio > 0 else '—',
                    fmt_cant(valor) if precio > 0 else '—',
                ])
            total_valorado += subtotal_fam
            sub_txt = fmt_cant(subtotal_fam) + ' €' if subtotal_fam > 0 else '—'
            filas.append(['', '', '', '', 'Subtotal', sub_txt])

            col_w = [6.5*cm, 2*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.5*cm]
            t = Table(filas, colWidths=col_w)
            estilo = [
                ('BACKGROUND',(0,0),(-1,0),AZUL),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),8),
                ('GRID',(0,0),(-1,-1),0.3,colors.lightgrey),
                ('PADDING',(0,0),(-1,-1),5),
                ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white, GRIS]),
                ('ALIGN',(1,0),(-1,-1),'CENTER'),
                ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
                ('BACKGROUND',(0,-1),(-1,-1),GRIS),
            ]
            t.setStyle(TableStyle(estilo))
            elements.append(t)
            elements.append(Spacer(1, 0.2*cm))

        # Total general
        elements.append(Spacer(1, 0.3*cm))
        t_total = Table([['', '', '', '', 'TOTAL STOCK', fmt_cant(total_valorado) + ' €']], colWidths=[6.5*cm, 2*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.5*cm])
        t_total.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1),AZUL),
            ('TEXTCOLOR',(0,0),(-1,-1),colors.white),
            ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),10),
            ('PADDING',(0,0),(-1,-1),7),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ]))
        elements.append(t_total)

    elements.append(Spacer(1, 0.5*cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Paragraph(f"Generado el {fecha_str} — Instapalma · Almacén", pie_style))
    doc.build(elements)
    return buffer.getvalue()


def generar_pdf_albaran(albaran):
    """albaran: dict con numero, nombre_operario, obra, lineas, fecha, tipo ('salida'|'devolucion')."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    AZUL = colors.HexColor('#1a3a5c')
    GRIS = colors.HexColor('#f5f5f5')

    titulo_style = ParagraphStyle('t', fontSize=20, textColor=AZUL,
        alignment=TA_CENTER, fontName='Helvetica-Bold')
    pie_style = ParagraphStyle('p', fontSize=7, textColor=colors.grey, alignment=TA_CENTER)

    import os as _os
    LOGO_PATH = _os.path.join(_os.path.dirname(__file__), 'logo.jpg')
    if _os.path.exists(LOGO_PATH):
        _lw = 4*cm; _lh = _lw / (1024/219)
        logo = RLImage(LOGO_PATH, width=_lw, height=_lh)
    else:
        logo = Paragraph("INSTAPALMA", titulo_style)

    es_devol = albaran.get('tipo','') == 'devolucion' or 'DEVOLUCI' in albaran.get('obra','').upper()
    titulo_alb = '📥 ALBARÁN DEVOLUCIÓN A ALMACÉN' if es_devol else '📤 ALBARÁN INTERNO DE ALMACÉN'
    titulo_sec = 'MATERIALES DEVUELTOS' if es_devol else 'MATERIALES RETIRADOS'

    cab_style = ParagraphStyle('c', fontName='Helvetica-Bold', fontSize=16,
        textColor=AZUL, alignment=1, leading=20)
    cab = Table([[logo, Paragraph(titulo_alb, cab_style)]], colWidths=[5*cm, 12*cm])
    cab.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),0),('RIGHTPADDING',(0,0),(-1,-1),0)]))
    elements.append(cab)
    elements.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=8))
    elements.append(Spacer(1, 0.2*cm))

    from datetime import datetime as dt
    fecha_str = albaran.get('fecha', dt.now().strftime('%d/%m/%Y %H:%M'))
    obra_label = 'Devuelto por' if es_devol else 'Obra / Cliente'
    t_cab = Table([
        ['Nº Albarán', albaran.get('numero','')],
        ['Fecha', fecha_str],
        ['Operario', albaran.get('nombre_operario','')],
        [obra_label, albaran.get('nombre_operario','') if es_devol else albaran.get('obra','')],
    ], colWidths=[4*cm, 13*cm])
    t_cab.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(0,-1),AZUL),
        ('TEXTCOLOR',(0,0),(0,-1),colors.white),
        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),10),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('PADDING',(0,0),(-1,-1),7),
        ('ROWBACKGROUNDS',(1,0),(1,-1),[colors.white, GRIS]),
    ]))
    elements.append(t_cab)
    elements.append(Spacer(1, 0.5*cm))

    # Líneas de material
    sec_style = ParagraphStyle('s', fontSize=9, textColor=colors.white,
        backColor=AZUL, fontName='Helvetica-Bold', spaceAfter=0, spaceBefore=4, borderPad=4)
    elements.append(Paragraph(titulo_sec, sec_style))
    elements.append(Spacer(1, 0.1*cm))

    lineas = albaran.get('lineas', [])
    # Signo negativo en devoluciones
    signo = -1 if es_devol else 1

    # Albarán SIEMPRE valorado
    filas = [['Material', 'Cant.', 'Ud.', 'P.Unit (€)', 'Total (€)']]
    total_general = 0
    for l in lineas:
        cant = float(l.get('cantidad', 0) or 0)
        precio = float(l.get('precio', 0) or 0)
        subtotal = cant * precio * signo
        total_general += subtotal
        cant_str = ('-' if es_devol else '') + fmt_cant(abs(cant))
        precio_str = fmt_cant(precio) if precio > 0 else '—'
        subtotal_str = ('-' if es_devol else '') + fmt_cant(abs(subtotal)) if precio > 0 else '—'
        filas.append([
            l.get('material',''),
            cant_str,
            l.get('unidad',''),
            precio_str,
            subtotal_str
        ])
    total_str = ('-' if es_devol else '') + fmt_cant(abs(total_general)) + ' €'
    filas.append(['', '', '', 'TOTAL', total_str])
    col_widths = [7.5*cm, 2.5*cm, 1.5*cm, 3*cm, 2.5*cm]

    t_lin = Table(filas, colWidths=col_widths)
    estilo_tabla = [
        ('BACKGROUND',(0,0),(-1,0),AZUL),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('PADDING',(0,0),(-1,-1),7),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),[colors.white, GRIS]),
        ('ALIGN',(1,0),(-1,-1),'CENTER'),
        ('BACKGROUND',(0,-1),(-1,-1),AZUL),
        ('TEXTCOLOR',(0,-1),(-1,-1),colors.white),
        ('FONTNAME',(0,-1),(-1,-1),'Helvetica-Bold'),
        ('SPAN',(0,-1),(2,-1)),
    ]
    t_lin.setStyle(TableStyle(estilo_tabla))
    elements.append(t_lin)
    elements.append(Spacer(1, 1*cm))

    # Firma
    if not es_devol:
        firma_table = Table([
            ['Firma del operario:', ''],
            ['', ''],
        ], colWidths=[8*cm, 9*cm])
        firma_table.setStyle(TableStyle([
            ('FONTNAME',(0,0),(0,0),'Helvetica-Bold'),
            ('FONTSIZE',(0,0),(-1,-1),9),
            ('LINEBELOW',(1,1),(1,1),1,colors.black),
            ('BOTTOMPADDING',(0,0),(-1,-1),18),
        ]))
        elements.append(firma_table)
        elements.append(Spacer(1, 0.3*cm))

    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    elements.append(Paragraph(f"Generado el {fecha_str} — Instapalma · Almacén", pie_style))
    doc.build(elements)
    return buffer.getvalue()

def subir_pdf_albaran(pdf_bytes, numero):
    """Guarda el PDF en la BD (upsert) y devuelve la URL pública del bot."""
    try:
        BOT_URL = os.environ.get('RAILWAY_PUBLIC_DOMAIN', 'bot-production-66b8.up.railway.app')
        numero_safe = numero.replace('/', '_').replace(' ', '_')
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO stock_albaranes (numero, pdf_bytes)
            VALUES (%s, %s)
            ON CONFLICT (numero) DO UPDATE SET pdf_bytes = EXCLUDED.pdf_bytes
        """, (numero, psycopg2.Binary(pdf_bytes)))
        conn.commit(); cur.close(); conn.close()
        return f"https://{BOT_URL}/albaran/{numero_safe}.pdf"
    except Exception as e:
        print(f"Error guardar PDF albarán en BD: {e}")
        return ''

# ── Flujo conversacional stock ────────────────────────────────────────────────
# Pasos salida: stock_salida_obra → stock_salida_material → stock_salida_cantidad → (loop) → stock_salida_confirmar
# Pasos devol:  stock_devol_material → stock_devol_cantidad → (loop) → stock_devol_confirmar
# Consulta:     directa, sin pasos

def buscar_material_msg(texto):
    """Devuelve (material_row, mensaje_error_o_None)."""
    r = get_material_by_nombre(texto)
    if r is None:
        return None, "❌ Error al consultar el almacén. Intenta de nuevo."
    if isinstance(r, tuple):
        return r, None  # exacto
    if len(r) == 0:
        return None, f"❌ No encontré *{texto}* en el catálogo. Escribe el nombre exacto o parte de él."
    if len(r) == 1:
        return r[0], None  # único resultado aproximado
    # Varios candidatos — detectar si son retales del mismo cable
    if detectar_retales(r):
        return None, ('RETALES', r)  # señal especial: es un grupo de retales
    lista = '\n'.join([f"*{i}.* {x[1]} ({x[3]} {x[2]})" for i, x in enumerate(r, 1)])
    return None, ('MULTIPLES', r, f"🔍 Encontré varios materiales:\n{lista}\n\nResponde con el *número* o escribe el nombre más completo.")


def detectar_retales(candidatos):
    """
    Detecta si una lista de candidatos son retales del mismo material
    (misma raíz de nombre, unidad m/ml/metros, o stock=1 en todos).
    """
    if len(candidatos) < 2:
        return False
    unidades_longitud = {'m', 'ml', 'metros', 'metro'}
    for c in candidatos:
        if str(c[2]).lower() in unidades_longitud:
            return True
    if all(c[3] == 1 for c in candidatos):
        return True
    return False


def sugerir_retales(retales, metros_necesarios):
    """
    Dado un listado de retales (id, nombre, unidad, stock, minimo)
    y los metros necesarios, devuelve sugerencias de combinación óptima.
    """
    import re
    def extraer_metros(r):
        nombre = r[1]
        m = re.search(r'(\d+[.,]?\d*)\s*(ml?|metros?)\b', nombre.lower())
        if m:
            return float(m.group(1).replace(',', '.'))
        m2 = re.search(r'(\d+[.,]?\d*)$', nombre.strip())
        if m2:
            return float(m2.group(1).replace(',', '.'))
        return float(r[3])  # fallback: stock_actual como longitud

    retales_con_m = []
    for r in retales:
        if r[3] <= 0:
            continue
        metros = extraer_metros(r)
        retales_con_m.append({
            'id': r[0], 'nombre': r[1], 'unidad': r[2],
            'metros': metros, 'stock': r[3],
            'precio': float(r[5]) if len(r) > 5 and r[5] else 0
        })

    total_disponible = sum(x['metros'] for x in retales_con_m)

    # Retales individuales que cubren solos (ordenados de menor a mayor desperdicio)
    individuales = sorted(
        [r for r in retales_con_m if r['metros'] >= metros_necesarios],
        key=lambda x: x['metros']
    )

    # Combinación greedy: los más grandes hasta cubrir
    combinacion = []
    if not individuales:
        disponibles = sorted(retales_con_m, key=lambda x: -x['metros'])
        acumulado = 0
        for r in disponibles:
            combinacion.append(r)
            acumulado += r['metros']
            if acumulado >= metros_necesarios:
                break

    return {
        'individuales': individuales,
        'combinacion': combinacion,
        'total_disponible': total_disponible,
        'retales': retales_con_m
    }



# ── Panel web Almacén ─────────────────────────────────────────────────────────

@app.route('/almacen')
def panel_almacen():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, nombre, unidad, stock_actual, stock_minimo, updated_at FROM stock_materiales ORDER BY nombre")
        materiales = cur.fetchall()
        cur.execute("SELECT tipo, material_nombre, cantidad, unidad, nombre_operario, obra, created_at FROM stock_movimientos ORDER BY created_at DESC LIMIT 100")
        movimientos = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500

    filas_mat = ''
    for m in materiales:
        mid, nombre, unidad, stock, minimo, upd = m
        alerta = ' style="background:#fff3e0"' if minimo > 0 and stock <= minimo else ''
        badge = f'<span style="background:#e65100;color:white;padding:2px 8px;border-radius:8px;font-size:11px">⚠️ Bajo mínimo</span>' if minimo > 0 and stock <= minimo else ''
        filas_mat += (
            f"<tr{alerta}>"
            f"<td>{nombre}</td><td style='text-align:center'>{stock}</td>"
            f"<td style='text-align:center'>{unidad}</td>"
            f"<td style='text-align:center'>{minimo}</td>"
            f"<td>{badge}</td>"
            f"<td><a href='/almacen/material/{mid}/editar' style='color:#1a3a5c;font-size:12px'>✏️</a></td>"
            f"</tr>"
        )
    if not filas_mat:
        filas_mat = "<tr><td colspan=6 class='empty'>Sin materiales — sube tu Excel para cargar el catálogo</td></tr>"

    filas_mov = ''
    iconos = {'salida': '📤', 'devolucion': '📥', 'entrada': '➕'}
    for mv in movimientos:
        tipo, mat, cant, unidad, op, obra, fecha = mv
        icono = iconos.get(tipo, '•')
        filas_mov += (
            f"<tr><td>{icono} {tipo.upper()}</td><td>{mat}</td>"
            f"<td style='text-align:center'>{cant} {unidad}</td>"
            f"<td>{op or ''}</td><td>{obra or ''}</td><td>{str(fecha)[:16]}</td></tr>"
        )
    if not filas_mov:
        filas_mov = "<tr><td colspan=6 class='empty'>Sin movimientos</td></tr>"

    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'>
    <title>Almacén — Instapalma</title>
    <style>{CSS_BASE}
    .upload-box{{background:white;border:2px dashed #1a3a5c;border-radius:10px;padding:24px;margin:0 0 20px;text-align:center}}
    .upload-box input[type=file]{{margin:10px 0}}
    .btn-up{{background:#1a3a5c;color:white;padding:10px 24px;border:none;border-radius:8px;font-weight:700;cursor:pointer;font-size:14px}}
    </style></head><body>
    <header>
      <div><h1>📦 Almacén</h1><p>Instapalma</p></div>
      <a href='/almacen/material/nuevo' style='background:white;color:#1a3a5c;padding:8px 16px;border-radius:8px;font-weight:700;text-decoration:none;font-size:13px'>+ Añadir material</a>
    </header>
    <div class='wrap' style='padding-top:20px'>

    <div class='upload-box'>
      <b style='color:#1a3a5c'>📂 Carga masiva desde Excel</b><br>
      <small>Columnas requeridas: <b>nombre</b>, <b>unidad</b>, <b>stock_actual</b>, <b>stock_minimo</b></small><br>
      <form method='POST' action='/almacen/importar' enctype='multipart/form-data'>
        <input type='file' name='archivo' accept='.xlsx,.xls,.csv' required>
        <button class='btn-up' type='submit'>⬆ Importar</button>
      </form>
    </div>

    <h3 style='color:#1a3a5c;margin-bottom:12px'>Stock de Materiales</h3>
    <table><thead><tr><th>Material</th><th>Stock</th><th>Unidad</th><th>Mínimo</th><th>Estado</th><th></th></tr></thead>
    <tbody>{filas_mat}</tbody></table>

    <h3 style='color:#1a3a5c;margin:24px 0 12px'>Últimos Movimientos</h3>
    <table><thead><tr><th>Tipo</th><th>Material</th><th>Cantidad</th><th>Operario</th><th>Obra</th><th>Fecha</th></tr></thead>
    <tbody>{filas_mov}</tbody></table>

    <div style='margin-top:16px'><a href='/almacen/albaranes' style='color:#1a3a5c;font-weight:700'>📋 Ver albaranes →</a></div>
    </div>
    <div style='padding:0 30px'><a href='/partes' class='back'>← Partes de Trabajo</a></div>
    </body></html>"""


@app.route('/almacen/material/nuevo', methods=['GET','POST'])
@app.route('/almacen/material/<int:mid>/editar', methods=['GET','POST'])
def editar_material(mid=None):
    from flask import request as req
    if req.method == 'POST':
        nombre = req.form.get('nombre','').strip()
        unidad = req.form.get('unidad','ud').strip()
        stock  = float(req.form.get('stock_actual', 0) or 0)
        minimo = float(req.form.get('stock_minimo', 0) or 0)
        precio = float(req.form.get('precio_unitario', 0) or 0)
        familia = req.form.get('familia', 'General').strip() or 'General'
        try:
            conn = get_db(); cur = conn.cursor()
            if mid:
                cur.execute("UPDATE stock_materiales SET nombre=%s, unidad=%s, stock_actual=%s, stock_minimo=%s, precio_unitario=%s, familia=%s, updated_at=NOW() WHERE id=%s",
                    (nombre, unidad, stock, minimo, precio, familia, mid))
            else:
                cur.execute("INSERT INTO stock_materiales (nombre, unidad, stock_actual, stock_minimo, precio_unitario, familia) VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (nombre) DO UPDATE SET unidad=%s, stock_actual=%s, stock_minimo=%s, precio_unitario=%s, familia=%s, updated_at=NOW()",
                    (nombre, unidad, stock, minimo, precio, familia, unidad, stock, minimo, precio, familia))
            conn.commit(); cur.close(); conn.close()
        except Exception as e:
            return f"Error: {e}", 500
        from flask import redirect
        return redirect('/almacen')

    datos = {'nombre':'','unidad':'ud','stock_actual':0,'stock_minimo':0,'precio_unitario':0,'familia':'General'}
    if mid:
        try:
            conn = get_db(); cur = conn.cursor()
            cur.execute("SELECT nombre, unidad, stock_actual, stock_minimo, precio_unitario, familia FROM stock_materiales WHERE id=%s", (mid,))
            r = cur.fetchone(); cur.close(); conn.close()
            if r: datos = {'nombre':r[0],'unidad':r[1],'stock_actual':r[2],'stock_minimo':r[3],'precio_unitario':r[4] or 0,'familia':r[5] or 'General'}
        except: pass
    titulo = 'Editar material' if mid else 'Nuevo material'
    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>{titulo}</title>
    <style>{CSS_BASE} label{{display:block;margin-bottom:4px;font-size:13px;font-weight:600;color:#555}}
    input{{width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px;margin-bottom:16px;box-sizing:border-box}}
    .btn{{background:#1a3a5c;color:white;padding:12px 28px;border:none;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer}}</style></head><body>
    <header><div><h1>📦 {titulo}</h1><p>Instapalma</p></div></header>
    <div style='max-width:500px;margin:30px auto;background:white;padding:28px;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.1)'>
    <form method='POST'>
    <label>Nombre del material</label><input name='nombre' value='{datos["nombre"]}' required>
    <label>Familia / Categoría</label><input name='familia' value='{datos["familia"]}' placeholder='Ej: Cables, Tubos, Herramientas...'>
    <label>Unidad (ud, m, ml, kg, rollo...)</label><input name='unidad' value='{datos["unidad"]}' required>
    <label>Stock actual</label><input name='stock_actual' type='number' step='0.001' value='{datos["stock_actual"]}' required>
    <label>Stock mínimo (alerta)</label><input name='stock_minimo' type='number' step='0.001' value='{datos["stock_minimo"]}'>
    <label>Precio unitario (€)</label><input name='precio_unitario' type='number' step='0.0001' value='{datos["precio_unitario"]}'>
    <button class='btn' type='submit'>Guardar</button>
    </form></div>
    <div style='padding:0 30px'><a href='/almacen' class='back'>← Volver</a></div>
    </body></html>"""


@app.route('/almacen/importar', methods=['POST'])
def importar_excel():
    from flask import request as req
    archivo = req.files.get('archivo')
    if not archivo:
        return "No se recibió archivo", 400
    try:
        import openpyxl, io as _io, csv
        content = archivo.read()
        nombre_arch = archivo.filename.lower()
        filas = []
        if nombre_arch.endswith('.csv'):
            reader = csv.DictReader(_io.StringIO(content.decode('utf-8-sig')))
            for row in reader:
                filas.append(row)
        else:
            wb = openpyxl.load_workbook(_io.BytesIO(content))
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                filas.append(dict(zip(headers, row)))

        conn = get_db(); cur = conn.cursor()
        importados = 0
        for f in filas:
            nombre = str(f.get('nombre',f.get('material',''))).strip()
            if not nombre or nombre.lower() in ['none','nan','']: continue
            unidad = str(f.get('unidad','ud')).strip() or 'ud'
            try: stock = float(str(f.get('stock_actual', f.get('stock',0))).replace(',','.'))
            except: stock = 0
            try: minimo = float(str(f.get('stock_minimo', f.get('minimo',0))).replace(',','.'))
            except: minimo = 0
            try: precio = float(str(f.get('precio_unitario', f.get('precio', f.get('pvp',0)))).replace(',','.'))
            except: precio = 0
            cur.execute("""
                INSERT INTO stock_materiales (nombre, unidad, stock_actual, stock_minimo, precio_unitario)
                VALUES (%s,%s,%s,%s,%s)
                ON CONFLICT (nombre) DO UPDATE SET unidad=%s, stock_actual=%s, stock_minimo=%s, precio_unitario=%s, updated_at=NOW()
            """, (nombre, unidad, stock, minimo, precio, unidad, stock, minimo, precio))
            importados += 1
        conn.commit(); cur.close(); conn.close()
        from flask import redirect
        return redirect(f'/almacen?importados={importados}')
    except Exception as e:
        return f"Error importando: {e}", 500


@app.route('/almacen/albaranes')
def panel_albaranes():
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT id, numero, nombre_operario, obra, created_at FROM stock_albaranes ORDER BY created_at DESC LIMIT 200")
        rows = cur.fetchall(); cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500

    filas = ''
    for r in rows:
        aid, num, op, obra, fecha = r
        filas += (
            f"<tr><td>{num}</td><td>{op or ''}</td><td>{obra or ''}</td>"
            f"<td>{str(fecha)[:16]}</td>"
            f"<td><a href='/almacen/albaranes/{aid}/pdf' style='color:#1a3a5c;font-size:12px'>⬇ PDF</a></td></tr>"
        )
    if not filas:
        filas = "<tr><td colspan=5 class='empty'>Sin albaranes</td></tr>"

    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>Albaranes — Instapalma</title>
    <style>{CSS_BASE}</style></head><body>
    <header><div><h1>📋 Albaranes de Almacén</h1><p>Instapalma</p></div></header>
    <div class='wrap' style='padding-top:20px'>
    <table><thead><tr><th>Nº Albarán</th><th>Operario</th><th>Obra</th><th>Fecha</th><th>PDF</th></tr></thead>
    <tbody>{filas}</tbody></table>
    </div>
    <div style='padding:0 30px'><a href='/almacen' class='back'>← Almacén</a></div>
    </body></html>"""


@app.route('/almacen/albaranes/<int:aid>/pdf')
def pdf_albaran(aid):
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT numero, nombre_operario, obra, lineas, created_at FROM stock_albaranes WHERE id=%s", (aid,))
        r = cur.fetchone(); cur.close(); conn.close()
    except Exception as e:
        return f"Error: {e}", 500
    if not r:
        return "No encontrado", 404
    import json as _json
    lineas = r[3] if isinstance(r[3], list) else _json.loads(r[3] or '[]')
    from datetime import datetime as dt
    fecha_str = str(r[4])[:16] if r[4] else dt.now().strftime('%d/%m/%Y %H:%M')
    pdf_bytes = generar_pdf_albaran({'numero': r[0], 'nombre_operario': r[1], 'obra': r[2], 'lineas': lineas, 'fecha': fecha_str})
    nombre_f = f"{r[0]}-{(r[2] or 'OBRA').replace(' ','_')[:30].upper()}.pdf"
    from flask import Response
    return Response(pdf_bytes, mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{nombre_f}"'})


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

# ══════════════════════════════════════════════════════════════════════════════
# WEB — Gestión de Herramienta /herramienta
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/herramienta')
def web_herramienta():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id, nombre, tipo, stock_almacen, observaciones FROM herramienta ORDER BY tipo, nombre")
    items = cur.fetchall()
    cur.execute("""
        SELECT ho.id, h.nombre, ho.operario_nombre, ho.obra, ho.fecha_alta
        FROM herramienta_obra ho
        JOIN herramienta h ON h.id = ho.herramienta_id
        ORDER BY ho.fecha_alta DESC
    """)
    en_obra = cur.fetchall()
    cur.execute("""
        SELECT hp.id, h.nombre, hp.operario_nombre, hp.descripcion
        FROM herramienta_personal hp
        JOIN herramienta h ON h.id = hp.herramienta_id
        ORDER BY hp.operario_nombre, h.nombre
    """)
    personal = cur.fetchall()
    cur.close(); conn.close()

    filas_almacen = "".join(
        f"<tr><td>{i[1]}</td><td><span class='tag tag-{i[2].lower()}'>{i[2]}</span></td><td style='text-align:center'><b>{i[3]}</b></td><td>{i[4] or ''}</td>"
        f"<td><a href='/herramienta/editar/{i[0]}'>✏️</a></td></tr>"
        for i in items
    )
    filas_obra = "".join(
        f"<tr><td>{o[1]}</td><td>{o[2] or ''}</td><td>{o[3]}</td><td>{str(o[4])[:10]}</td>"
        f"<td><a href='/herramienta/baja_obra/{o[0]}' onclick='return confirm(\"¿Confirmar baja?\")'>🔙 Baja</a></td></tr>"
        for o in en_obra
    )
    filas_pers = "".join(
        f"<tr><td>{p[1]}</td><td>{p[2] or ''}</td><td>{p[3] or ''}</td></tr>"
        for p in personal
    )

    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>Herramienta — Instapalma</title>
    <style>{CSS_BASE}
    .tabs{{display:flex;gap:8px;margin:16px 0}}
    .tab{{padding:8px 18px;border-radius:20px;background:#eee;cursor:pointer;font-weight:600;font-size:13px;border:none}}
    .tab.active{{background:#1a3a5c;color:white}}
    .section{{display:none}}.section.active{{display:block}}
    table{{width:100%;border-collapse:collapse;font-size:13px}}
    th{{background:#1a3a5c;color:white;padding:8px;text-align:left}}
    td{{padding:7px 8px;border-bottom:1px solid #eee}}
    tr:hover td{{background:#f5f7ff}}
    .tag{{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:700}}
    .tag-almacen{{background:#d4f0d4;color:#1a5c1a}}
    .tag-personal{{background:#d4e8f0;color:#1a3a5c}}
    .tag-epis{{background:#f0ecd4;color:#5c4a1a}}
    .btn-add{{background:#1a3a5c;color:white;padding:9px 20px;border-radius:8px;border:none;font-weight:700;cursor:pointer;text-decoration:none;display:inline-block;margin-bottom:12px}}
    </style>
    <script>
    function showTab(t){{document.querySelectorAll('.section').forEach(s=>s.classList.remove('active'));document.querySelectorAll('.tab').forEach(b=>b.classList.remove('active'));document.getElementById(t).classList.add('active');event.target.classList.add('active');}}
    </script></head><body>
    <header><div><h1>🔧 Herramienta</h1><p>Instapalma</p></div></header>
    <div class='container'>
    <div class='tabs'>
      <button class='tab active' onclick='showTab("almacen")'>📦 Almacén ({len(items)})</button>
      <button class='tab' onclick='showTab("obra")'>🏗️ En Obra ({len(en_obra)})</button>
      <button class='tab' onclick='showTab("pers")'>👷 Personal ({len(personal)})</button>
    </div>

    <div id='almacen' class='section active'>
      <a href='/herramienta/nuevo' class='btn-add'>+ Nueva herramienta</a>
      <table><tr><th>Nombre</th><th>Tipo</th><th>Stock</th><th>Observaciones</th><th></th></tr>
      {filas_almacen if filas_almacen else "<tr><td colspan='5' style='text-align:center;color:#999'>Sin herramienta registrada</td></tr>"}
      </table>
    </div>

    <div id='obra' class='section'>
      <table><tr><th>Herramienta</th><th>Operario</th><th>Obra</th><th>Fecha alta</th><th></th></tr>
      {filas_obra if filas_obra else "<tr><td colspan='5' style='text-align:center;color:#999'>Nada en obra ahora mismo</td></tr>"}
      </table>
    </div>

    <div id='pers' class='section'>
      <table><tr><th>Herramienta / EPI</th><th>Operario</th><th>Descripción</th></tr>
      {filas_pers if filas_pers else "<tr><td colspan='3' style='text-align:center;color:#999'>Sin asignaciones personales</td></tr>"}
      </table>
    </div>

    <div style='margin-top:24px'>
      <a href='https://bot-production-66b8.up.railway.app/herramienta/pdf/todo' class='btn-add' style='background:#2d7a2d'>📄 PDF completo</a>
      <a href='https://bot-production-66b8.up.railway.app/herramienta/pdf/almacen' class='btn-add' style='background:#1a3a5c;margin-left:8px'>📦 PDF almacén</a>
      <a href='https://bot-production-66b8.up.railway.app/herramienta/pdf/obra' class='btn-add' style='background:#5c3a1a;margin-left:8px'>🏗️ PDF obra</a>
    </div>
    </div></body></html>"""


@app.route('/herramienta/nuevo', methods=['GET','POST'])
@app.route('/herramienta/editar/<int:mid>', methods=['GET','POST'])
def web_herramienta_form(mid=None):
    from flask import request as req
    if req.method == 'POST':
        nombre  = req.form.get('nombre','').strip()
        tipo    = req.form.get('tipo','almacen').strip()
        stock   = int(req.form.get('stock_almacen', 1) or 1)
        obs     = req.form.get('observaciones','').strip()
        conn = get_db(); cur = conn.cursor()
        if mid:
            cur.execute("UPDATE herramienta SET nombre=%s, tipo=%s, stock_almacen=%s, observaciones=%s, updated_at=NOW() WHERE id=%s",
                (nombre, tipo, stock, obs, mid))
        else:
            cur.execute("INSERT INTO herramienta (nombre, tipo, stock_almacen, observaciones) VALUES (%s,%s,%s,%s) ON CONFLICT (nombre) DO UPDATE SET tipo=%s, stock_almacen=%s, observaciones=%s, updated_at=NOW()",
                (nombre, tipo, stock, obs, tipo, stock, obs))
        conn.commit(); cur.close(); conn.close()
        from flask import redirect
        return redirect('/herramienta')

    datos = {'nombre':'','tipo':'almacen','stock_almacen':1,'observaciones':''}
    if mid:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT nombre, tipo, stock_almacen, observaciones FROM herramienta WHERE id=%s", (mid,))
        r = cur.fetchone(); cur.close(); conn.close()
        if r: datos = {'nombre':r[0],'tipo':r[1],'stock_almacen':r[2],'observaciones':r[3] or ''}
    titulo = 'Editar herramienta' if mid else 'Nueva herramienta'
    return f"""<!DOCTYPE html><html><head><meta charset='utf-8'><title>{titulo}</title>
    <style>{CSS_BASE} label{{display:block;margin:12px 0 4px;font-size:13px;font-weight:600;color:#555}}
    input,select,textarea{{width:100%;padding:10px;border:1px solid #ddd;border-radius:8px;font-size:14px;box-sizing:border-box}}
    .btn{{background:#1a3a5c;color:white;padding:12px 28px;border:none;border-radius:8px;font-size:15px;font-weight:700;cursor:pointer;margin-top:16px}}</style></head><body>
    <header><div><h1>🔧 {titulo}</h1><p>Instapalma</p></div></header>
    <div style='max-width:500px;margin:30px auto;background:white;padding:28px;border-radius:12px;box-shadow:0 2px 8px rgba(0,0,0,.1)'>
    <form method='POST'>
    <label>Nombre</label><input name='nombre' value='{datos["nombre"]}' required>
    <label>Tipo</label>
    <select name='tipo'>
      <option value='almacen' {'selected' if datos["tipo"]=='almacen' else ''}>📦 Almacén (compartida)</option>
      <option value='personal' {'selected' if datos["tipo"]=='personal' else ''}>👷 Personal</option>
      <option value='epis' {'selected' if datos["tipo"]=='epis' else ''}>🦺 EPIs</option>
    </select>
    <label>Unidades en almacén</label><input name='stock_almacen' type='number' value='{datos["stock_almacen"]}' min='0'>
    <label>Observaciones</label><textarea name='observaciones' rows='3'>{datos["observaciones"]}</textarea>
    <button class='btn' type='submit'>Guardar</button>
    </form></div>
    <div style='padding:0 30px'><a href='/herramienta' class='back'>← Volver</a></div></body></html>"""


@app.route('/herramienta/baja_obra/<int:oid>')
def web_baja_obra(oid):
    """Devuelve la herramienta al almacén desde la web."""
    try:
        conn = get_db(); cur = conn.cursor()
        cur.execute("SELECT herramienta_id FROM herramienta_obra WHERE id=%s", (oid,))
        row = cur.fetchone()
        if row:
            cur.execute("UPDATE herramienta SET stock_almacen = stock_almacen + 1, updated_at=NOW() WHERE id=%s", (row[0],))
            cur.execute("DELETE FROM herramienta_obra WHERE id=%s", (oid,))
            conn.commit()
        cur.close(); conn.close()
    except Exception as e:
        print(f"Error baja_obra web: {e}")
    from flask import redirect
    return redirect('/herramienta')


@app.route('/herramienta/pdf/<modo>')
def web_herramienta_pdf(modo):
    """Genera y sirve PDF de herramienta: todo/almacen/obra/personal."""
    try:
        pdf_bytes = generar_pdf_herramienta(modo)
        from flask import Response
        return Response(pdf_bytes, mimetype='application/pdf',
            headers={'Content-Disposition': f'inline; filename="herramienta_{modo}.pdf"'})
    except Exception as e:
        return f"Error: {e}", 500
